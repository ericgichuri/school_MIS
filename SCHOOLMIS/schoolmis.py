from time import sleep
from tkinter import  *
from tkinter import font
from turtle import bgcolor, color, up, width
import customtkinter
from tkinter import ttk,messagebox,filedialog
import mysql.connector
from tkinter.ttk import Style
from PIL import Image,ImageTk,ImageDraw,ImageFont
from datetime import datetime
import time,os,qrcode,datetime,webbrowser,win32api
from tkcalendar import DateEntry
from fpdf import FPDF
import openpyxl
import bcrypt

customtkinter.set_appearance_mode("light")
customtkinter.set_default_color_theme("dark-blue")
#========COLORS===================
fg1="#FFD900"
bg1="#000066"
fg2="#E600AC"
fg3="aliceblue"
bg2="aliceblue"
coldark="#3F3F3F"

#=================================
conn=""
myusername=""
today=time.strftime("%Y/%m/%d")
curtime=time.strftime("%H:%M:%S")
stdid=""
stfid=""
bookid=""
sysuid=""
currenttermid=""
currenttermname=""
mytermfee=0
salarystaffid=""
salarystfid=""
issue_id_clearance=""
book_no_clearance=""
entry_width=150
template_f=""
template_b=""
student_pic=""
school_logo=""
list_school_details=[]
lbfont3=("times",14,"bold")
sel_pt_cash_id=""
student_class=""

#===================system for================
school_type="Primary_school"
Primary_school=["PP_1","PP_1","Grade_1","Grade_2","Grade_3","Grade_4","Grade_5","Grade_6","Grade_7","Grade_8","Grade_9"]
Secondary_school=["Form_1","Form_2","Form_3","Form_4"]
Collage=["Year_1","Year_2","Year_3","Year_4"]
class_values=Primary_school
#=============================================


#========database connections=========
def databaseconnections():
    global conn
    conn=mysql.connector.connect(host="127.0.0.1",user="root",password="")
    cursor=conn.cursor()
    # ---------create database-----------
    dbname="schoolmis"
    sql="CREATE DATABASE IF NOT EXISTS %s"%(dbname)
    try:
        cursor.execute(sql)
        conn.commit()
    except IOError:
        pass
    conn=mysql.connector.connect(host="127.0.0.1",user="root",password="",database=dbname)
    #--------------create table staffs---------------------------------
    sql1="""CREATE TABLE IF NOT EXISTS staffs(
        staffid int NOT NULL AUTO_INCREMENT,
        fname varchar(20) NOT NULL,
        lname varchar(20) NOT NULL,
        sname varchar(20) NOT NULL,
        idno varchar(15) NOT NULL,
        phoneno varchar(15) NOT NULL,
        email varchar(50) NOT NULL,
        stafftype varchar(20) NOT NULL,
        staffno varchar(20) NOT NULL,
        occupation varchar(20) NOT NULL,
        employer varchar(20) NOT NULL,
        religion varchar(20) NOT NULL,
        homelocation varchar(40) NOT NULL,
        recordby varchar(20) NOT NULL,
        recorddate date,
        PRIMARY KEY(staffid),
        UNIQUE KEY(idno),
        UNIQUE KEY(phoneno)
    )"""
    try:
        cursor=conn.cursor()
        cursor.execute(sql1)
    except IOError:
        pass
    #-----------------------create table system users----------------
    sql2="""CREATE TABLE IF NOT EXISTS systemusers(
        sysuserid int AUTO_INCREMENT NOT NULL,
        sysstaffid int NOT NULL,
        username varchar(20) NOT NULL,
        password varchar(100) NOT NULL,
        role varchar(20) NOT NULL,
        lastlogindate date,
        lastlogintime time,
        activestatus int,
        recordby varchar(20) NOT NULL,
        PRIMARY KEY(sysuserid),
        UNIQUE KEY(sysstaffid),
        UNIQUE KEY(username),
        FOREIGN KEY(sysstaffid) REFERENCES staffs(staffid)
    )"""
    try:
        cursor=conn.cursor()
        cursor.execute(sql2)
    except IOError:
        pass

    #---------------create table students------------------------------
    sql3="""CREATE TABLE IF NOT EXISTS students(
        studentid int AUTO_INCREMENT NOT NULL,
        fname varchar(20) NOT NULL,
        lname varchar(20) NOT NULL,
        sname varchar(20) NOT NULL,
        gender  varchar(10) NOT NULL,
        dob date,
        studymode varchar(15) NOT NULL,
        admno varchar(10) NOT NULL,
        form varchar(10) NOT NULL,
        parentname varchar(30) NOT NULL,
        phoneno varchar(20) NOT NULL,
        religion varchar(20) NOT NULL,
        homelocation varchar(40) NOT NULL,
        recordby varchar(20) NOT NULL,
        recorddate date,    
        PRIMARY KEY(studentid),
        UNIQUE KEY(admno)
    )"""
    try:
        cursor=conn.cursor()
        cursor.execute(sql3)
    except IOError:
        pass

    #-----------------create table Term-------------------------------
    sql4="""CREATE TABLE IF NOT EXISTS terms(
        termid int NOT NULL AUTO_INCREMENT,
        yearcreated varchar(20) NOT NULL,
        term varchar(20) NOT NULL,
        termname varchar(20) NOT NULL,
        startdate date NOT NULL,
        enddate date NOT NULL,
        termstatus int NOT NULL,
        recordby varchar(20),
        PRIMARY KEY(termid),
        UNIQUE KEY(termname)
    )"""
    try:
        cursor=conn.cursor()
        cursor.execute(sql4)
    except IOError:
        pass

    #-------------------create table reporting----------------------------
    sql5="""CREATE TABLE IF NOT EXISTS reporting(
        reportingid int NOT NULL AUTO_INCREMENT,
        reporttermid int NOT NULL,
        stdadmno varchar(20) NOT NULL,
        reportdate date,
        class varchar(20) NOT NULL,
        PRIMARY KEY(reportingid)
    )
    """
    try:
        cursor=conn.cursor()
        cursor.execute(sql5)
    except IOError:
        pass

    #--------------create table fees-----------------------------------------
    sql6="""CREATE TABLE IF NOT EXISTS fees(
        paymentid int NOT NULL AUTO_INCREMENT,
        stadmno varchar(20) NOT NULL,
        amount double NOT NULL,
        paidvia varchar(20) NOT NULL,
        remarks varchar(20) NOT NULL,
        paymentdate date,
        recordby varchar(20),
        PRIMARY KEY(paymentid)
    )"""
    try:
        cursor=conn.cursor()
        cursor.execute(sql6)
    except IOError:
        pass

    #----------------create table term fee. store every set term fee-------------------
    sql7="""CREATE TABLE IF NOT EXISTS termfee(
        feeid int NOT NULL AUTO_INCREMENT,
        termid int NOT NULL,
        amount double NOT NULL,
        createddate date,
        recordby varchar(20),
        PRIMARY KEY(feeid),
        UNIQUE KEY(termid)
    )"""
    try:
        cursor=conn.cursor()
        cursor.execute(sql7)
    except IOError:
        pass

    #----------------create table salaries for non government/BOM staffs-----------------
    sql8="""CREATE TABLE IF NOT EXISTS salaries(
        salaryid int NOT NULL AUTO_INCREMENT,
        staffid int NOT NULL,
        staffno varchar(20) NOT NULL,
        amount double NOT NULL,
        recordby varchar(20) NOT NULL,
        PRIMARY KEY(salaryid),
        UNIQUE KEY(staffid),
        UNIQUE KEY(staffno)
    )"""
    try:
        cursor=conn.cursor()
        cursor.execute(sql8)
    except IOError:
        pass

    #---------------create table projects. all planned projects registered here-----------
    sql9="""CREATE TABLE IF NOT EXISTS projects(
        projectid int NOT NULL AUTO_INCREMENT,
        projectname varchar(30) NOT NULL,
        projectdescription varchar(500) NOT NULL,
        projectcost double NOT NULL,
        projectstart date NOT NULL,
        projectexpectedend date NOT NULL,
        recordby varchar(20),
        PRIMARY KEY(projectid)
    )"""
    try:
        cursor=conn.cursor()
        cursor.execute(sql9)
    except IOError:
        pass

    #-------------create table petty cash for allowances and other school resources------
    sql10="""CREATE TABLE IF NOT EXISTS pettycash(
        ptcashid int NOT NULL AUTO_INCREMENT,
        ptcategory varchar(20) NOT NULL,
        ptnarration varchar(200),
        ptamount double NOT NULL,
        ptrecordby varchar(20) NOT NULL,
        ptapproved int NOT NULL,
        ptapprovedby varchar(20) NOT NULL,
        ptapproveddate date NOT NULL,
        PRIMARY KEY(ptcashid)
    )"""
    try:
        cursor=conn.cursor()
        cursor.execute(sql10)
    except IOError:
        pass

    #--------------create table books for library management----------------------------
    sql11="""CREATE TABLE IF NOT EXISTS books(
        bookid int NOT NULL AUTO_INCREMENT,
        bookcategory varchar(20) NOT NULL,
        booktitle varchar(50) NOT NULL,
        bookauthor varchar(30) NOT NULL,
        bookpublisher varchar(20) NOT NULL,
        publishdate date NOT NULL,
        quantity int NOT NULL,
        bookprice double NOT NULL,
        bookno varchar(20),
        recordby varchar(20),
        recorddate date NOT NULL,
        PRIMARY KEY(bookid),
        UNIQUE KEY(bookno)
    )"""
    try:
        cursor=conn.cursor()
        cursor.execute(sql11)
    except IOError:
        pass

    #-----------------create tabke issue book for capturing students issued books-----------
    sql12="""CREATE TABLE IF NOT EXISTS issuebook(
        issueid int NOT NULL AUTO_INCREMENT,
        bookno varchar(30) NOT NULL,
        stadmno varchar(20) NOT NULL,
        studentname varchar(20) NOT NULL,
        issuedate date,
        returndate date,
        cleared varchar(20) NOT NULL,
        recordby varchar(20) NOT NULL,
        PRIMARY KEY(issueid)
    )"""
    try:
        cursor=conn.cursor()
        cursor.execute(sql12)
    except IOError:
        pass

    #---------------create table table school information like name,logo,contact-------------
    sql13="""CREATE TABLE IF NOT EXISTS schoolinfo(
            sch_id int NOT NULL AUTO_INCREMENT,
            sch_name varchar(100) NOT NULL,
            sch_motto varchar(150) NOT NULL,
            sch_contact varchar(30) NOT NULL,
            sch_box varchar(50) NOT NULL,
            sch_email varchar(100) NOT NULL,
            sch_sign varchar(20) NOT NULL,
            sch_logo varchar(100) NOT NULL,
            school_type varchar(30) NOT NULL,
            PRIMARY KEY(sch_id)
        )"""
    try:
        cursor=conn.cursor()
        cursor.execute(sql13)
    except IOError:
        pass

    #---------------insert these in school info for the first time installation but updated later----------
    sql14="INSERT INTO schoolinfo(sch_name,sch_motto,sch_contact,sch_box,sch_email,sch_sign,sch_logo,school_type) VALUES('%s','%s','%s','%s','%s','%s','%s','%s')"
    try:
        #check if school details exists
        cursor=conn.cursor()
        sql15="SELECT * FROM schoolinfo"
        cursor.execute(sql15)
        schooldts=cursor.fetchall()
        if schooldts:
            pass
        else:
            cursor=conn.cursor()
            cursor.execute(sql14%('School MIS','Smart Technologies Smart Service','+254707273244','12-Nyeri','ericsoftwaresolutions@gmail.com','ESS','schoolmis.png','Primary_school'))
            conn.commit()
    except IOError:
        pass

    #-------------create table salary payments for staffs--------------------------------------------------
    sql15="""CREATE TABLE IF NOT EXISTS salary_payment(
            pay_id int NOT NULL AUTO_INCREMENT,
            staff_no varchar(20) NOT NULL,
            staff_name varchar(40) NOT NULL,
            amount_paid double NOT NULL,
            datepaid date NOT NULL,
            processedby varchar(20) NOT NULL,
            PRIMARY KEY(pay_id)
        )"""
    try:
        cursor=conn.cursor()
        cursor.execute(sql15)
    except IOError:
        pass
#=====================================
#------------call DB connection to create all database info----------------------
databaseconnections()
#check if table staff has something
try:
    cursor=conn.cursor()
    sql="SELECT * FROM staffs"
    cursor.execute(sql)
    res=cursor.fetchall()
    if res:
        pass
    else:
        sql2="INSERT INTO staffs(fname,lname,sname,occupation,staffno,recordby,recorddate) VALUES('%s','%s','%s','%s','%s','%s','%s')"
        cursor.execute(sql2%('Admin','Admin','Admin','ICT','ST_001','Admin',today))
        conn.commit()

        x_pass=str('Admin').encode('utf-8')
        hashed=bcrypt.hashpw(x_pass,bcrypt.gensalt())
        hashedpw=hashed.decode('utf-8')
        sql3="INSERT INTO systemusers(sysstaffid,username,password,role) VALUES(%s,'%s','%s','%s')"
        val=(1,'Admin',hashedpw,'ICT')
        cursor.execute(sql3%val)
        conn.commit()
except IOError:
    pass
    messagebox.showerror("Error",str(e))
    conn.rollback()
#========fonts====================
fonttitle=("times",21,"bold")
fontentries=("times",14)
fontlbl=("times",14)
fontbtn=("times",14)
fontlbl2=("times",15,"bold")

#=================================

#-----------------------system info and help--------------------------
txt_for_system="""Welcome To School Management Information System (School MIS)
\nThe System is Developed to manage school data and information by:
\tAutomating data\n\tEasy of retrival\n\tAvoiding too much staffs\n\tTo limit use of papers.
The system has the following modules:
\t\t1: Student module\n\t\t2: Staff module\n\t\t3: Reporting\n\t\t4: Finance\n\t\t5: Library\n\t\t6: System users\n\t\t7: Examination
The Dashboard:
i. School information Name,motto,contact,PO BOX,email,sign.
ii. School Finance where Non government staff are paid off and cash approval.
iii. School Analysis. Here Get analyzed Data for finance, reporting, and resources.
\nThe Student Module:
i: Register student, Update student,delete and viewing student data.
ii: Student ID generating
\nThe Staff Moldule:
i: Register staff, Update staff,delete and viewing staff data.
\nReporting Module:
i: Add Term(Register the term students to report Select Term 1,2,3 then save)
ii: Student Reporting(Student Report For current term).
iii: Viewing Reported students
\nFinance Module:
i: If current Term Fee is set Dont change the Term ID then save.
ii: Student Fee payment: search student, enter amount, remarks, and paid via.
iii: To search student enter admission number the press tab
iv: View statement click view statement when student is found.
v: Print statement.
vi: View fees payment also filtering available.
vii: Staff salaries only non-government staff allowed.
viii: Projects registration.
ix: Petty cash or allowances for staffs or cash spent for minor things
\n Libray Module:
i: Add book,Update,delete and view books.
ii: issue book, view issued book.
iii: make clearance double click the record the clearance
\n System users Module:
i: Add System user and must be a staff, Select a role on the system.
ii: update system user roles and deleting user.
iii: View User status 0 offline status 1 online
\n Examination:  
"""

#-------------------------main window function------------------------
#------ Main Modules available dashboard,students,staff,examination,reporting,finances,library,system users
def tomainwindow():
    global icon,icon1
    #-----------------------system info---------------------------------------------------
    def to_system_infomation():
        info_window=customtkinter.CTkToplevel(fg_color=bg1)
        info_window.title("School MIS")
        info_window.iconphoto(False,icon)
        info_scr_w=info_window.winfo_screenwidth()
        info_scr_h=info_window.winfo_screenheight()
        info_window_w=500
        info_window_h=400
        info_x_cord=(info_scr_w-info_window_w)/2
        info_y_cord=(info_scr_h-info_window_h)/2
        info_window.geometry("%dx%d+%d+%d"%(info_window_w,info_window_h,info_x_cord,info_y_cord))
        info_window.resizable(False,False)
        txt_system_info=customtkinter.CTkTextbox(info_window,fg_color=fg3,text_color=bg1,font=fontentries)
        txt_system_info.pack(side=LEFT,fill=BOTH,expand=True)
        
        
        txt_system_info.insert(1.0,txt_for_system)
        txt_system_info.configure(state="disabled")
        info_window.mainloop()
    def to_system_settings():
        
        def tocall_dev():
            webbrowser.open("tel:+254707273244")
        def toemail_dev():
            webbrowser.open("mailto:ericsoftwaresolutions@gmail.com")
        def towebsite():
            webbrowser.open("https://ericgichuri.github.io/")
        def togithub():
            webbrowser.open("https://github.com/ericgichuri")
        settings_window=customtkinter.CTkToplevel()
        settings_window.title("School MIS")
        settings_window.iconphoto(False,icon)
        settings_scr_w=settings_window.winfo_screenwidth()
        settings_scr_h=settings_window.winfo_screenheight()
        settings_window_w=500
        settings_window_h=400
        settings_x_cord=(settings_scr_w-settings_window_w)/2
        settings_y_cord=(settings_scr_h-settings_window_h)/2
        settings_window.geometry("%dx%d+%d+%d"%(settings_window_w,settings_window_h,settings_x_cord,settings_y_cord))
        settings_window.resizable(False,False)
        lb=customtkinter.CTkLabel(settings_window,text="About Developer",font=fonttitle,text_color=fg1,fg_color=bg1)
        lb.pack(side=TOP,fill=X)
        lb=customtkinter.CTkLabel(settings_window,text="Developed By: Eric Software Solutions",text_color=bg1,font=fontlbl)
        lb.pack(side=TOP,fill=X,padx=6,pady=4)
        lb=customtkinter.CTkLabel(settings_window,text="Tel: +254707273244/ +25459091813",text_color=bg1,font=fontlbl)
        lb.pack(side=TOP,fill=X,padx=6,pady=4)
        lb=customtkinter.CTkLabel(settings_window,text="Email: ericsoftwaresolutions@gmail.com",text_color=bg1,font=fontlbl)
        lb.pack(side=TOP,fill=X,padx=6,pady=4)
        lb=customtkinter.CTkLabel(settings_window,text="For Any Assistance use these platforms",text_color=bg1,font=fontlbl)
        lb.pack(side=TOP,fill=X,padx=6,pady=4)
        btn_call=customtkinter.CTkButton(settings_window,text="Call",fg_color=bg1,text_color=fg1,width=100,font=fontlbl,command=tocall_dev)
        btn_call.pack(side=LEFT,padx=6,pady=4,anchor=CENTER)
        btn_email=customtkinter.CTkButton(settings_window,text="Email",fg_color=bg1,text_color=fg1,width=100,font=fontlbl,command=toemail_dev)
        btn_email.pack(side=LEFT,padx=6,pady=4,anchor=CENTER)
        btn_website=customtkinter.CTkButton(settings_window,text="Website",fg_color=bg1,text_color=fg1,width=100,font=fontlbl,command=towebsite)
        btn_website.pack(side=LEFT,padx=6,pady=4,anchor=CENTER)
        btn_github=customtkinter.CTkButton(settings_window,text="Github",fg_color=bg1,text_color=fg1,width=100,font=fontlbl,command=togithub)
        btn_github.pack(side=LEFT,padx=6,pady=4,anchor=CENTER)

    #-----------logout function check user of the user to logout---------------------------
    def tologout():
        global myusername
        msgtologout=messagebox.askyesno("Confirm message","Do you want to logout? ")
        if msgtologout==True:
            sql="UPDATE systemusers SET activestatus=0 WHERE username='%s'"%(myusername)
            try:
                cursor=conn.cursor()
                cursor.execute(sql)
                conn.commit()
                get_system_analysis()
                exit()
            except IOError:
                pass
    #------------------------hide all tabs from the main window----------
    def toremovealltabs():
        framecontainer.hide(0)
        framecontainer.hide(1)
        framecontainer.hide(2)
        framecontainer.hide(3)
        framecontainer.hide(4)
        framecontainer.hide(5)
        framecontainer.hide(6)
        framecontainer.hide(7)

    #----------------------dashboard tab to display-----------------------------
    def todashboard():
        toremovealltabs()
        framecontainer.add(framedashboard,text="Dashboard")
        framecontainer.select(0)
    
    #----------------------Students tab to display-----------------------------
    def tostudents():
        toremovealltabs()
        framecontainer.add(framestudents,text="Students")
        framecontainer.select(1)

    #----------------------Staffs tab to display-----------------------------
    def tostaffs():
        toremovealltabs()
        framecontainer.add(framestaffs,text="Staffs")
        framecontainer.select(2)

    #----------------------Examination tab to display-----------------------------
    def toexamination():
        toremovealltabs()
        framecontainer.add(frameexamination,text="Examination")
        framecontainer.select(3)

    #----------------------Reporting tab to display-----------------------------
    def toreporting():
        toremovealltabs()
        framecontainer.add(framereporting,text="Reporting")
        framecontainer.select(4)

    #----------------------finance tab to display-----------------------------
    def tofinance():
        toremovealltabs()
        framecontainer.add(framefinance,text="Finance")
        framecontainer.select(5)

    #----------------------library tab to display-----------------------------
    def tolibrary():
        toremovealltabs()
        framecontainer.add(framelibrary,text="Libary")
        framecontainer.select(6)

    #----------------------system users tab to display-----------------------------
    def tosystemusers():
        toremovealltabs()
        framecontainer.add(framesystemuser,text="System users")
        framecontainer.select(7)
    
    #----------------------Student tab functions-----------------------------
    def toaddstudent():
        ntbook1.select(0)
    def toeditstudent():
        ntbook1.select(1)
    def toviewstudent():
        ntbook1.select(2)
    def tostudentidgenerator():
        ntbook1.select(3)
    def viewstudents():
        try:
            cursor=conn.cursor()
            sql="SELECT admno,CONCAT(fname,' ',lname,' ',sname) AS name,gender,parentname,phoneno,form,studymode,religion FROM students"
            cursor.execute(sql)
            results=cursor.fetchall()
            if results:
                for records in tablestudents.get_children():
                    tablestudents.delete(records)
                for i in results:
                    tablestudents.insert('',END,values=i)
        except IOError:
            pass
    def clearstudentform():
        fname.delete(0,END)
        lname.delete(0,END)
        sname.delete(0,END)
        gender.set("")
        dob.delete(0,END)
        studymode.set("")
        admno.delete(0,END)
        form.set("")
        parent.delete(0,END)
        phoneno.delete(0,END)
        religion.set("")
        homelocation.delete(0,END)
    def addstudents():
        if fname.get()=="":
            return False
        elif lname.get()=="":
            return False
        elif sname.get()=="":
            return False
        elif gender.get()=="":
            return False
        elif dob.get()=="":
            return False
        elif studymode.get()=="":
            return False
        elif admno.get()=="":
            return False
        elif form.get()=="":
            return False
        elif parent.get()=="":
            return False
        elif phoneno.get()=="":
            return False
        elif religion.get()=="":
            return False
        elif homelocation.get()=="":
            return False
        else:
            global myusername
            try:
                sql="INSERT INTO students(fname,lname,sname,gender,dob,admno,studymode,form,parentname,phoneno,religion,homelocation,recordby,recorddate) VALUES ('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')"
                cursor=conn.cursor()
                cursor.execute(sql%(fname.get(),lname.get(),sname.get(),gender.get(),dob.get(),admno.get(),studymode.get(),form.get(),parent.get(),phoneno.get(),religion.get(),homelocation.get(),myusername,today))
                conn.commit()
                clearstudentform()
                viewstudents()
                get_all_student_analysis()
                messagebox.showinfo("Success message","Student saved successfully")
            except IOError:
                messagebox.showerror("Alert message","Unable to process.\nTry again")
                pass
    def clearstudenteditform():
        global stdid
        stdid=""
        efname.delete(0,END)
        elname.delete(0,END)
        esname.delete(0,END)
        egender.set("")
        edob.delete(0,END)
        eform.set("")
        estudymode.set("")
        eadmno.delete(0,END)
        eparent.delete(0,END)
        ephoneno.delete(0,END)
        ereligion.set("")
        ehomelocation.delete(0,END)
    def searchstudent():
        global stdid,myusername
        if searchstdby.get()=="Admission":
            try:
                sql="SELECT * FROM students WHERE admno='%s'"%(searchtext.get())
                cursor=conn.cursor()
                cursor.execute(sql)
                results=cursor.fetchall()
                if results:
                    clearstudenteditform()
                    stdid=results[0][0]
                    efname.insert(END,results[0][1])
                    elname.insert(END,results[0][2])
                    esname.insert(END,results[0][3])
                    egender.set(results[0][4])
                    edob.insert(END,results[0][5])
                    estudymode.set(results[0][6])
                    eadmno.insert(END,results[0][7])
                    eform.set(results[0][8])
                    eparent.insert(END,results[0][9])
                    ephoneno.insert(END,results[0][10])
                    ereligion.set(results[0][11])
                    ehomelocation.insert(END,results[0][12])
                    
            except IOError:
                pass
        elif searchstdby.get()=="PhoneNo":
            try:
                sql="SELECT * FROM students WHERE phoneno='%s'"%(searchtext.get())
                cursor=conn.cursor()
                cursor.execute(sql)
                results=cursor.fetchall()
                if results:
                    clearstudenteditform()
                    stdid=results[0][0]
                    efname.insert(END,results[0][1])
                    elname.insert(END,results[0][2])
                    esname.insert(END,results[0][3])
                    egender.set(results[0][4])
                    edob.insert(END,results[0][5])
                    estudymode.set(results[0][6])
                    eadmno.insert(END,results[0][7])
                    eform.set(results[0][8])
                    eparent.insert(END,results[0][9])
                    ephoneno.insert(END,results[0][10])
                    ereligion.set(results[0][11])
                    ehomelocation.insert(END,results[0][12])
            except IOError:
                pass
    def updatestudent():
        if efname.get()=="":
            return False
        elif elname.get()=="":
            return False
        elif esname.get()=="":
            return False
        elif egender.get()=="":
            return False
        elif edob.get()=="":
            return False
        elif estudymode.get()=="":
            return False
        elif eadmno.get()=="":
            return False
        elif eform.get()=="":
            return False
        elif eparent.get()=="":
            return False
        elif ephoneno.get()=="":
            return False
        elif ereligion.get()=="":
            return False
        elif ehomelocation.get()=="":
            return False
        else:
            global stdid,myusername
            if stdid!="":
                try:
                    sql="UPDATE students SET fname='%s',lname='%s',sname='%s',gender='%s',dob='%s',studymode='%s',admno='%s',form='%s',parentname='%s',phoneno='%s',religion='%s',homelocation='%s',recordby='%s' WHERE studentid=%s"%(efname.get(),elname.get(),esname.get(),egender.get(),edob.get(),estudymode.get(),eadmno.get(),eform.get(),eparent.get(),ephoneno.get(),ereligion.get(),ehomelocation.get(),myusername,stdid)
                    cursor=conn.cursor()
                    cursor.execute(sql)
                    conn.commit()
                    stdid=""
                    searchtext.delete(0,END)
                    clearstudenteditform()
                    viewstudents()
                    get_all_student_analysis()
                    messagebox.showinfo("Success","Student Updated successfully")
                except IOError:
                    messagebox.showwarning("Error","Unable to update.\n try Again")
                    pass
            else:
                messagebox.showwarning("Alert","Search student to Update")
    def deletestudent():
        
        global stdid
        if stdid=="":
            messagebox.showwarning("Alert","Search student to delete")
        else:
            msgtodelete=messagebox.askyesno("Alert question","do you want to detete student Admno "+searchtext.get())
            if msgtodelete==True:
                try:
                    sql="DELETE FROM students WHERE studentid=%s"%(stdid)
                    cursor=conn.cursor()
                    cursor.execute(sql)
                    conn.commit()
                    stdid=""
                    clearstudenteditform()
                    viewstudents()
                    get_all_student_analysis()
                    messagebox.showinfo("success","Student deleted successful")
                    searchtext.delete(0,END)
                except IOError:
                    messagebox.showwarning("Error","Unable to delete.\n try Again")
                    pass
    def chech_school_type():
        global school_type,class_values
        if school_type=="Primary_school":
            class_values=Primary_school
        elif school_type=="Secondary_school":
            class_values==Secondary_school
        elif school_type=="Collage":
            class_values=Collage
            
    global myusername
    
    win=customtkinter.CTk(fg_color="#000066")
    win.title("School Management Information system")
    icon=PhotoImage(file="icons/schoolmis.png")
    win.iconphoto(False,icon)
    win.geometry("1100x660+20+30")

    #========icons====================
    
    icon1=customtkinter.CTkImage(light_image=Image.open("icons/schoolmis.png"),dark_image=Image.open("icons/schoolmis.png"),size=(40,40))
    icon2=customtkinter.CTkImage(light_image=Image.open("icons/schoolmis.png"),dark_image=Image.open("icons/schoolmis.png"),size=(70,70))
    iconabout=customtkinter.CTkImage(light_image=Image.open("icons/help.png"),dark_image=Image.open("icons/help.png"),size=(15,15))
    iconsettings=customtkinter.CTkImage(light_image=Image.open("icons/about.png"),dark_image=Image.open("icons/about.png"),size=(15,15))
    #=================================

    #========fonts====================
    fonttitle=("times",22,"bold")
    #=================================
    
    #========COLORS===================
    fg1="#FFD900"
    bg1="#000066"
    hovbg1="#000077"
    fg2="#E600AC"
    bg2="aliceblue"
    coldark="#3F3F3F"
    bg3="#808080"
    cancelcol="#cc0000"
    #=================================
    
    s1=ttk.Style()
    s1.theme_use("default")
    s1.configure("TNotebook.Tab",background=coldark,foreground=fg1)
    s1.map("TNotebook",background=[("selected",bg1)],foreground=[("selected","aliceblue")])
    s2=ttk.Style()
    s2.theme_use("clam")
    s2.configure("TNotebook",background=bg1,foreground=fg1)
    
    #style1=ttk.Style()

    s2.configure("Treeview",
        background="aliceblue",
        foreground="black",
        rowheight="25",
        fieldbackground="aliceblue"

    )
    s2.configure("Treeview.Heading",font=("times",10,"bold"))
    s2.map("Treeview",
        background=[("selected",bg1)],
        foreground=[("selected",fg1)]
    )
    #===================================
    lbltitle=customtkinter.CTkLabel(win,text="        School Management Information System",image=icon1,compound=LEFT,text_color=fg1,font=fonttitle,bg_color=bg1)
    lbltitle.pack(side=TOP,fill=X,pady=(3,1))
    frameuserdetail=customtkinter.CTkFrame(win,fg_color=bg1,corner_radius=0)
    frameuserdetail.pack(side=TOP,fill=X)
    lbusername=customtkinter.CTkLabel(frameuserdetail,text="Admin: "+myusername,font=fontlbl,text_color=bg2)
    lbusername.pack(side=LEFT,padx=(10,5))
    lbldate=customtkinter.CTkLabel(frameuserdetail,text="Date: "+today,font=fontlbl,text_color=bg2)
    lbldate.pack(side=LEFT,padx=(20,20))
    btnsettings=customtkinter.CTkButton(frameuserdetail,image=iconsettings,cursor="hand2",text="",text_color=bg2,fg_color=bg1,hover_color=hovbg1,width=4,corner_radius=5,command=to_system_settings)
    btnsettings.pack(side=RIGHT,padx=(5,10))
    btnabout=customtkinter.CTkButton(frameuserdetail,image=iconabout,cursor="hand2",text="",text_color=bg2,fg_color=bg1,hover_color=hovbg1,width=4,corner_radius=5,command=to_system_infomation)
    btnabout.pack(side=RIGHT,padx=(5,10))
    framemain=customtkinter.CTkFrame(win,fg_color="#595959")
    framemain.pack(fill=BOTH,expand=True)
    framesidebar=customtkinter.CTkFrame(framemain,fg_color=coldark,width=200)
    framesidebar.pack(side=LEFT,fill=Y)
    logolbl=customtkinter.CTkLabel(framesidebar,text="",image=icon2,height=120)
    logolbl.pack(pady=(10,15))
    btndashboard=customtkinter.CTkButton(framesidebar,text="Dashboard",width=180,fg_color=bg1,hover_color=hovbg1,text_color=fg1,cursor="hand2",font=fontbtn,command=todashboard)
    btndashboard.pack(padx=10,pady=6)
    btnstudents=customtkinter.CTkButton(framesidebar,text="Students",width=180,fg_color=bg1,hover_color=hovbg1,text_color=fg1,cursor="hand2",font=fontbtn,command=tostudents)
    btnstudents.pack(padx=10,pady=6)
    btnstaffs=customtkinter.CTkButton(framesidebar,text="Staffs",width=180,fg_color=bg1,hover_color=hovbg1,text_color=fg1,cursor="hand2",font=fontbtn,command=tostaffs)
    btnstaffs.pack(padx=10,pady=6)
    btnexams=customtkinter.CTkButton(framesidebar,text="Examination",width=180,fg_color=bg1,hover_color=hovbg1,text_color=fg1,cursor="hand2",font=fontbtn,command=toexamination)
    btnexams.pack(padx=10,pady=6)
    btnreporting=customtkinter.CTkButton(framesidebar,text="Reporting",width=180,fg_color=bg1,hover_color=hovbg1,text_color=fg1,cursor="hand2",font=fontbtn,command=toreporting)
    btnreporting.pack(padx=10,pady=6)
    btnfees=customtkinter.CTkButton(framesidebar,text="Finance",width=180,fg_color=bg1,hover_color=hovbg1,text_color=fg1,cursor="hand2",font=fontbtn,command=tofinance)
    btnfees.pack(padx=10,pady=6)
    btnlibrary=customtkinter.CTkButton(framesidebar,text="Library",width=180,fg_color=bg1,hover_color=hovbg1,text_color=fg1,cursor="hand2",font=fontbtn,command=tolibrary)
    btnlibrary.pack(padx=10,pady=6)
    btnsysuser=customtkinter.CTkButton(framesidebar,text="System Users",width=180,fg_color=bg1,hover_color=hovbg1,text_color=fg1,cursor="hand2",font=fontbtn,command=tosystemusers)
    btnsysuser.pack(padx=10,pady=6)
    btnlogout=customtkinter.CTkButton(framesidebar,text="Logout",width=180,fg_color=bg1,hover_color=hovbg1,text_color=fg1,cursor="hand2",font=fontbtn,command=tologout)
    btnlogout.pack(padx=10,pady=6)

    framecontainer=ttk.Notebook(framemain,style="TNotebook")
    framecontainer.pack(side=LEFT,fill=BOTH,expand=True)
    framedashboard=customtkinter.CTkFrame(framecontainer,fg_color=bg3,bg_color=coldark)
    framedashboard.place(x=0,y=0,relwidth=1,relheight=1)
    framestudents=customtkinter.CTkFrame(framecontainer,fg_color=bg3,bg_color=coldark)
    framestudents.place(x=0,y=0,relwidth=1,relheight=1)
    framestaffs=customtkinter.CTkFrame(framecontainer,fg_color=bg3,bg_color=coldark)
    framestaffs.place(x=0,y=0,relwidth=1,relheight=1)
    frameexamination=customtkinter.CTkFrame(framecontainer,fg_color=bg3,bg_color=coldark)
    frameexamination.place(x=0,y=0,relwidth=1,relheight=1)
    framereporting=customtkinter.CTkFrame(framecontainer,fg_color=bg3,bg_color=coldark)
    framereporting.place(x=0,y=0,relwidth=1,relheight=1)
    framefinance=customtkinter.CTkFrame(framecontainer,fg_color=bg3,bg_color=coldark)
    framefinance.place(x=0,y=0,relwidth=1,relheight=1)
    framelibrary=customtkinter.CTkFrame(framecontainer,fg_color=bg3,bg_color=coldark)
    framelibrary.place(x=0,y=0,relwidth=1,relheight=1)
    framesystemuser=customtkinter.CTkFrame(framecontainer,fg_color=bg3,bg_color=coldark)
    framesystemuser.place(x=0,y=0,relwidth=1,relheight=1)
    framecontainer.add(framedashboard,text="Dashboard")
    framecontainer.add(framestudents,text="Students")
    framecontainer.add(framestaffs,text="Staffs")
    framecontainer.add(frameexamination,text="Examination")
    framecontainer.add(framereporting,text="Reporting")
    framecontainer.add(framefinance,text="Finance")
    framecontainer.add(framelibrary,text="Library")
    framecontainer.add(framesystemuser,text="System users")
    
    
    #============frame dashboard===============

    #---------dashboard functions--------------
    def toschoolinfo():
        ntbook0.select(0)
        get_system_analysis()
    def toschoolfinance():
        ntbook0.select(1)
    def toschoolanalysis():
        ntbook0.select(2)

    def clear_school_details():
        sch_name_txt.delete(0,END)
        sch_motto_txt.delete(0,END)
        sch_contact_txt.delete(0,END)
        sch_box_txt.delete(0,END)
        sch_email_txt.delete(0,END)
        sch_sign_txt.delete(0,END)
    def school_info_detail():
        global school_logo,list_school_details,school_type,class_values
        try:
            sql="SELECT * FROM schoolinfo"
            cursor=conn.cursor()
            cursor.execute(sql)
            results=cursor.fetchall()
            if results:
                list_school_details=results
                clear_school_details()
                sch_name_txt.insert(END,results[0][1])
                sch_motto_txt.insert(END,results[0][2])
                sch_contact_txt.insert(END,results[0][3])
                sch_box_txt.insert(END,results[0][4])
                sch_email_txt.insert(END,results[0][5])
                sch_sign_txt.insert(END,results[0][6])
                school_logo="icons/"+results[0][7]
                sch_type_txt.set(results[0][8])
                school_type=results[0][8]
                sc_logo=customtkinter.CTkImage(light_image=Image.open(school_logo),size=(80,80))
                lb_school_logo.configure(image=sc_logo)
                chech_school_type()
                if school_type=="Primary_school":
                    class_values=Primary_school
                elif school_type=="Secondary_school":
                    class_values==Secondary_school
                elif school_type=="Collage":
                    class_values=Collage
            else:
                pass
        except IOError:
            pass
    def select_logo_fun():
        global school_logo
        sel_logo=filedialog.askopenfilename(initialdir="",title="Select Logo",filetype=[
                ("Images","*.png;*.jpg;*.jpeg;*.gif;*.bmp")
            ])
        if sel_logo:
            school_logo=sel_logo
        else:
            messagebox.showwarning("Warning","School Logo not selected")
    def update_school_info_details():
        global school_logo
        if sch_name_txt.get()=="" and sch_motto_txt.get()=="" and sch_contact_txt.get()=="" and sch_box_txt.get()=="" and sch_email_txt.get()=="" and sch_sign_txt.get()=="":
            messagebox.showerror("Error","All Fields Must be filled")
        else:
            try:
                sql="UPDATE schoolinfo SET sch_name='%s',sch_motto='%s',sch_contact='%s',sch_box='%s',sch_email='%s',sch_sign='%s',sch_logo='schoollogo.png',school_type='%s'"
                cursor=conn.cursor()
                cursor.execute(sql%(sch_name_txt.get(),sch_motto_txt.get(),sch_contact_txt.get(),sch_box_txt.get(),sch_email_txt.get(),sch_sign_txt.get(),sch_type_txt.get()))
                conn.commit()
                school_l=Image.open(school_logo)
                school_l.save("icons/schoollogo.png")
                messagebox.showinfo("Success","School Details Updated successfully\n Restart")
                school_info_detail()
                chech_school_type()
            except IOError:
                conn.rollback()
                pass
    def get_all_student_analysis():
        try:
            cursor=conn.cursor()
            sql="SELECT COUNT(studentid) FROM students"
            cursor.execute(sql)
            t_students=cursor.fetchall()
            if t_students:
                lb_t_student.configure(text=f"Total Students: {t_students[0][0]}")
            else:
                lb_t_student.configure(text="Total Students: 0")

            sql2="SELECT COUNT(studentid) FROM students WHERE gender='Male'"
            cursor.execute(sql2)
            t_Mstudents=cursor.fetchall()
            if t_Mstudents:
                lb_t_Boystudent.configure(text=f"Boys Students: {t_Mstudents[0][0]}")
            else:
                lb_t_Boystudent.configure(text="Boys Students: 0")

            sql3="SELECT COUNT(studentid) FROM students WHERE gender='Female'"
            cursor.execute(sql3)
            t_Fstudents=cursor.fetchall()
            if t_Fstudents:
                lb_t_Girlstudent.configure(text=f"Girls Students: {t_Fstudents[0][0]}")
            else:
                lb_t_Girlstudent.configure(text="Girls Students: 0")
        except IOError:
            pass
            
    def get_all_staff_analysis():
        try:
            cursor=conn.cursor()
            sql="SELECT COUNT(staffid) FROM staffs"
            cursor.execute(sql)
            t_staff=cursor.fetchall()
            if t_staff:
                lb_t_staff.configure(text=f"Total Staffs: {t_staff[0][0]}")
            else:
                lb_t_staff.configure(text=f"Total Staffs: 0")

            sql2="SELECT COUNT(staffid) FROM staffs WHERE stafftype='Teaching'"
            cursor.execute(sql2)
            t_Tstaff=cursor.fetchall()
            if t_Tstaff:
                lb_t_Tstaff.configure(text=f"Teaching Staffs: {t_Tstaff[0][0]}")
            else:
                lb_t_Tstaff.configure(text="Teaching Staffs: 0")

            sql3="SELECT COUNT(staffid) FROM staffs WHERE stafftype='Non-Teaching'"
            cursor.execute(sql3)
            t_NTstaff=cursor.fetchall()
            if t_NTstaff:
                lb_t_NTstaff.configure(text=f"Non-Teaching Staffs: {t_NTstaff[0][0]}")
            else:
                lb_t_NTstaff.configure(text="Non-Teaching Staffs: 0")
        except IOError:
            pass
    def get_system_analysis():
        try:
            cursor=conn.cursor()
            sql="SELECT COUNT(sysuserid) FROM systemusers"
            cursor.execute(sql)
            t_Sysuser=cursor.fetchall()
            if t_Sysuser:
                lb_t_Sysuser.configure(text=f"Total Users: {t_Sysuser[0][0]}")
            else:
                lb_t_Sysuser.configure(text="Total Users: 0")

            sql2="SELECT COUNT(sysuserid) FROM systemusers WHERE activestatus=1"
            cursor.execute(sql2)
            t_ONusers=cursor.fetchall()
            if t_ONusers:
                lb_t_ONuser.configure(text=f"Online Users: {t_ONusers[0][0]}")
            else:
                lb_t_ONuser.configure(text=f"Online Users: 0")

            sql3="SELECT COUNT(sysuserid) FROM systemusers WHERE activestatus=0"
            cursor.execute(sql3)
            t_OFFusers=cursor.fetchall()
            if t_OFFusers:
                lb_t_OFFuser.configure(text=f"Offline Users: {t_OFFusers[0][0]}")
            else:
                lb_t_OFFuser.configure(text=f"Offline Users: 0")
            
            
        except IOError:
            pass


    #------------------------------------------

    framesubmenu0=customtkinter.CTkFrame(framedashboard,width=150,fg_color=coldark)
    framesubmenu0.pack(side=LEFT,fill=Y)
    btnschool_info=customtkinter.CTkButton(framesubmenu0,text="School Info",text_color=fg1,fg_color=bg1,hover_color=hovbg1,cursor="hand2",command=toschoolinfo)
    btnschool_info.pack(pady=6,padx=6)
    btnschool_finance=customtkinter.CTkButton(framesubmenu0,text="School Finance",text_color=fg1,fg_color=bg1,hover_color=hovbg1,cursor="hand2",command=toschoolfinance)
    btnschool_finance.pack(pady=6,padx=6)
    btnschool_analysis=customtkinter.CTkButton(framesubmenu0,text="School Analysis",text_color=fg1,fg_color=bg1,hover_color=hovbg1,cursor="hand2",command=toschoolanalysis)
    btnschool_analysis.pack(pady=6,padx=6)
    ntbook0=ttk.Notebook(framedashboard)
    ntbook0.pack(side=LEFT,fill=BOTH,expand=True)
    fmschool_info=customtkinter.CTkFrame(ntbook0,fg_color=bg3,bg_color=coldark)
    fmschool_info.place(x=0,y=0,relheight=1,relwidth=1)
    fmschool_finance=customtkinter.CTkFrame(ntbook0,fg_color=bg3,bg_color=coldark)
    fmschool_finance.place(x=0,y=0,relheight=1,relwidth=1)
    fmschool_analysis=customtkinter.CTkFrame(ntbook0,fg_color=bg3,bg_color=coldark)
    fmschool_analysis.place(x=0,y=0,relheight=1,relwidth=1)
    
    ntbook0.add(fmschool_info,text="School Info")
    ntbook0.add(fmschool_finance,text="School Finance")
    ntbook0.add(fmschool_analysis,text="School Analysis")

    lbframe_sc_details=LabelFrame(fmschool_info,bg=bg3,text="School Detail",fg=bg1)
    lbframe_sc_details.pack(side=LEFT,fill=BOTH,expand=True)
    lb_school_logo=customtkinter.CTkLabel(lbframe_sc_details,text="",fg_color=bg1)
    lb_school_logo.grid(column=0,row=0,columnspan=2,pady=2,ipadx=2,ipady=1)
    #--------------display school info-------------
    lb=customtkinter.CTkLabel(lbframe_sc_details,text="School Name",font=fontlbl,anchor=W,text_color=bg1)
    lb.grid(column=0,row=1,sticky=W,padx=5)
    sch_name_txt=customtkinter.CTkEntry(lbframe_sc_details,border_width=1,width=200)
    sch_name_txt.grid(column=0,row=2,columnspan=2,padx=5)
    lb=customtkinter.CTkLabel(lbframe_sc_details,text="School Motto",font=fontlbl,anchor=W,text_color=bg1)
    lb.grid(column=0,row=3,sticky=W,padx=5)
    sch_motto_txt=customtkinter.CTkEntry(lbframe_sc_details,border_width=1,width=200)
    sch_motto_txt.grid(column=0,row=4,columnspan=2,padx=5)
    lb=customtkinter.CTkLabel(lbframe_sc_details,text="School Contact",font=fontlbl,anchor=W,text_color=bg1)
    lb.grid(column=0,row=5,sticky=W,padx=5)
    sch_contact_txt=customtkinter.CTkEntry(lbframe_sc_details,border_width=1,width=200)
    sch_contact_txt.grid(column=0,row=6,columnspan=2,padx=5)
    lb=customtkinter.CTkLabel(lbframe_sc_details,text="P.O BOX",font=fontlbl,anchor=W,text_color=bg1)
    lb.grid(column=0,row=7,sticky=W,padx=5)
    sch_box_txt=customtkinter.CTkEntry(lbframe_sc_details,border_width=1,width=200)
    sch_box_txt.grid(column=0,row=8,columnspan=2,padx=5)
    lb=customtkinter.CTkLabel(lbframe_sc_details,text="School Email",font=fontlbl,anchor=W,text_color=bg1)
    lb.grid(column=0,row=9,sticky=W,padx=5)
    sch_email_txt=customtkinter.CTkEntry(lbframe_sc_details,border_width=1,width=200)
    sch_email_txt.grid(column=0,row=10,columnspan=2,padx=5)
    lb=customtkinter.CTkLabel(lbframe_sc_details,text="Sign/Initials",font=fontlbl,anchor=W,text_color=bg1)
    lb.grid(column=0,row=11,sticky=W,padx=5)
    sch_sign_txt=customtkinter.CTkEntry(lbframe_sc_details,border_width=1,width=200)
    sch_sign_txt.grid(column=0,row=12,columnspan=2,padx=5)
    lb=customtkinter.CTkLabel(lbframe_sc_details,text="Schoo Type",font=fontlbl,anchor=W,text_color=bg1)
    lb.grid(column=0,row=13,sticky=W,padx=5)
    sch_type_txt=customtkinter.CTkOptionMenu(lbframe_sc_details,button_color=bg1,button_hover_color=bg1,text_color="black",fg_color="white",width=200,values=["Primary_school","Secondary_school","Collage"])
    sch_type_txt.grid(column=0,row=14,columnspan=2,padx=5)

    btn_change_logo=customtkinter.CTkButton(lbframe_sc_details,text="Change Logo",text_color=fg1,fg_color=bg1,hover_color=bg1,width=110,command=select_logo_fun)
    btn_change_logo.grid(column=0,row=15,padx=3,pady=(5,2))
    btn_update_sch_details=customtkinter.CTkButton(lbframe_sc_details,text="Update",text_color=fg1,fg_color=bg1,hover_color=bg1,width=110,command=update_school_info_details)
    btn_update_sch_details.grid(column=1,row=15,padx=3,pady=(5,2))
    #----------------------------------------------
    fmschool_info1=customtkinter.CTkFrame(fmschool_info,fg_color=bg3)
    fmschool_info1.pack(side=LEFT,fill=BOTH,expand=True)
    lbframe_sc_details1=LabelFrame(fmschool_info1,bg=bg3,text="Student Info",fg=bg1)
    lbframe_sc_details1.pack(side=TOP,fill=BOTH,expand=True)
    lb_t_student=customtkinter.CTkLabel(lbframe_sc_details1,text="Student",justify=RIGHT,text_color=bg1,font=lbfont3)
    lb_t_student.pack(expand=True)
    lb_t_Boystudent=customtkinter.CTkLabel(lbframe_sc_details1,text="Student Boys",justify=RIGHT,text_color=bg1,font=lbfont3)
    lb_t_Boystudent.pack(expand=True)
    lb_t_Girlstudent=customtkinter.CTkLabel(lbframe_sc_details1,text="Student Girls",justify=RIGHT,text_color=bg1,font=lbfont3)
    lb_t_Girlstudent.pack(expand=True)

    lbframe_sc_details2=LabelFrame(fmschool_info1,bg=bg3,text="Staff Info",fg=bg1)
    lbframe_sc_details2.pack(side=TOP,fill=BOTH,expand=True)
    lb_t_staff=customtkinter.CTkLabel(lbframe_sc_details2,text="Staff",justify=RIGHT,text_color=bg1,font=lbfont3)
    lb_t_staff.pack(expand=True)
    lb_t_Tstaff=customtkinter.CTkLabel(lbframe_sc_details2,text="Staff",justify=RIGHT,text_color=bg1,font=lbfont3)
    lb_t_Tstaff.pack(expand=True)
    lb_t_NTstaff=customtkinter.CTkLabel(lbframe_sc_details2,text="Staff",justify=RIGHT,text_color=bg1,font=lbfont3)
    lb_t_NTstaff.pack(expand=True)
    
    lbframe_sc_details3=LabelFrame(fmschool_info1,bg=bg3,text="System Info",fg=bg1)
    lbframe_sc_details3.pack(side=TOP,fill=BOTH,expand=True)
    lb_t_Sysuser=customtkinter.CTkLabel(lbframe_sc_details3,text="System User",justify=RIGHT,text_color=bg1,font=lbfont3)
    lb_t_Sysuser.pack(expand=True)
    lb_t_ONuser=customtkinter.CTkLabel(lbframe_sc_details3,text="Online",justify=RIGHT,text_color=bg1,font=lbfont3)
    lb_t_ONuser.pack(expand=True)
    lb_t_OFFuser=customtkinter.CTkLabel(lbframe_sc_details3,text="Offline",justify=RIGHT,text_color=bg1,font=lbfont3)
    lb_t_OFFuser.pack(expand=True)

    #----------------------school finance approvals-------------------
    #get_payment
    def get_payment_list():
        try:
            cursor=conn.cursor()
            sql="SELECT * FROM salary_payment ORDER BY pay_id DESC"
            cursor.execute(sql)
            results=cursor.fetchall()
            if results:
                for records in table_payments.get_children():
                    table_payments.delete(records)
                for i in results:
                    table_payments.insert('',END,values=i)
        except IOError:
            pass
    def get_staff_paydetails(*args):

        if txt_staff_no.get()=="":
            pass
        else:
            sql="SELECT * FROM salaries WHERE staffno='%s'"%(txt_staff_no.get())
            cursor=conn.cursor()
            cursor.execute(sql)
            results=cursor.fetchall()
            if results:
                sql2="SELECT CONCAT(fname,' ',lname,' ',sname) AS name FROM staffs WHERE staffid='%s'"%(results[0][1])
                cursor.execute(sql2)
                results2=cursor.fetchall()
                if results2:
                    txt_staff_no.delete(0,END)
                    txt_staff_name.delete(0,END)
                    txt_amount_no.delete(0,END)
                    txt_staff_no.insert(END,results[0][2])
                    txt_staff_name.insert(END,results2[0][0])
                    txt_amount_no.insert(END,results[0][3])
                else:
                    pass
            else:
                messagebox.showerror("Error","No Staff with this staff No")
                txt_staff_no.delete(0,END)
                txt_staff_name.delete(0,END)
                txt_amount_no.delete(0,END)
    def pay_off_staffs():
        global myusername
        if txt_staff_no.get()=="" and txt_staff_name.get()=="" and txt_amount_no.get()=="":
            messagebox.showwarning("Warning","User Not selected")
        else:
            msg=messagebox.askyesno("Confirm message","Do you want to pay off staff No: " +txt_staff_no.get())
            if msg==1:
                try:
                    cursor=conn.cursor()
                    sql="INSERT INTO salary_payment(staff_no,staff_name,amount_paid,datepaid,processedby) VALUES('%s','%s',%s,'%s','%s')"
                    values=(txt_staff_no.get(),txt_staff_name.get(),txt_amount_no.get(),today,myusername)
                    cursor.execute(sql%values)
                    conn.commit()
                    messagebox.showinfo("Success","Staff Payment successful")
                    cancel_pay_off()
                    get_payment_list()
                except IOError:
                    conn.rollback()
                    pass
            else:
                cancel_pay_off()
    def cancel_pay_off():
        txt_staff_no.delete(0,END)
        txt_staff_name.delete(0,END)
        txt_amount_no.delete(0,END)
    def get_petty_cash_toapprove():
        for records in table_approvals.get_children():
            table_approvals.delete(records)
        try:
            cursor=conn.cursor()
            sql="SELECT ptcashid,ptcategory,ptamount,ptrecordby FROM pettycash WHERE ptapproved=0"
            cursor.execute(sql)
            results=cursor.fetchall()
            if results:
                for records in table_approvals.get_children():
                    table_approvals.delete(records)
                for i in results:
                    table_approvals.insert('',END,values=i)
            else:
                pass
        except IOError:
            pass
    
    def get_petty_cash_record(event):
        global sel_pt_cash_id
        try:
            selected_item=table_approvals.focus()
            items=table_approvals.item(selected_item,'values')
            sel_pt_cash_id=items[0]
            try:
                cursor=conn.cursor()
                sql="SELECT * FROM pettycash WHERE ptcashid=%s"%(sel_pt_cash_id)
                cursor.execute(sql)
                results=cursor.fetchall()
                if results:
                    txt_narration.delete(1.0,END)
                    txt_narration.insert(END,results[0][2])
            except IOError:
                pass
        except:
            pass
    def approve_petty_cash():
        global sel_pt_cash_id,myusername,today
        if sel_pt_cash_id=="":
            messagebox.showwarning("Warning","Record not selected")
        else:
            msg=messagebox.askyesno("Confirm message","Do you want to Approve this cash? ")
            if msg==1:
                try:
                    sql="UPDATE pettycash SET ptapproved=1,ptapprovedby='%s',ptapproveddate='%s' WHERE ptcashid=%s"
                    cursor=conn.cursor()
                    cursor.execute(sql%(myusername,today,sel_pt_cash_id))
                    conn.commit()
                    sel_pt_cash_id=""
                    messagebox.showinfo("Success","Approved successfully")
                    get_petty_cash_toapprove()
                    txt_narration.delete(1.0,END)
                    viewpettycash()
                except IOError:
                    pass
    def reject_petty_cash():
        global sel_pt_cash_id
        if sel_pt_cash_id=="":
            messagebox.showwarning("Warning","Record not selected")
        else:
            msg=messagebox.askyesno("Confirm message","Do you want to reject/delete this cash record")
            if msg==1:
                try:
                    sql="DELETE FROM pettycash WHERE ptcashid=%s"
                    cursor=conn.cursor()
                    cursor.execute(sql%(sel_pt_cash_id))
                    conn.commit()
                    sel_pt_cash_id=""
                    get_petty_cash_toapprove()
                    txt_narration.delete(1.0,END)
                    viewpettycash()
                except IOError:
                    pass

    #get payee
    fm_payments=LabelFrame(fmschool_finance,text="Payments",bg=bg3,fg=bg1)
    fm_payments.pack(side=LEFT,fill=BOTH,expand=True)
    fm_payments1=customtkinter.CTkFrame(fm_payments,fg_color=bg3)
    fm_payments1.pack(side=TOP,fill=BOTH,expand=True)
    table_payments=ttk.Treeview(fm_payments1,height=6)
    table_payments.pack(side=LEFT,fil=BOTH,expand=True)
    table_payments['show']="headings"
    table_payments['columns']=(0,1,2,3,4,5)
    table_payments.heading(0,text="Id")
    table_payments.heading(1,text="Staff No")
    table_payments.heading(2,text="Name")
    table_payments.heading(3,text="Amount")
    table_payments.heading(4,text="Date")
    table_payments.heading(5,text="process By")
    table_payments.column(0,width=30,anchor=CENTER)
    table_payments.column(1,width=50,anchor=CENTER)
    table_payments.column(2,width=70,anchor=CENTER)
    table_payments.column(3,width=60,anchor=CENTER)
    table_payments.column(4,width=60,anchor=CENTER)
    table_payments.column(5,width=60,anchor=CENTER)

    scroll_table_payments=customtkinter.CTkScrollbar(fm_payments1,command=table_payments.yview)
    scroll_table_payments.pack(side=LEFT,fill=Y)
    table_payments.configure(yscrollcommand=scroll_table_payments)

    fm_payments2=customtkinter.CTkFrame(fm_payments,fg_color=bg3)
    fm_payments2.pack(side=TOP,fill=BOTH,expand=True,pady=(10,5),padx=6)
    lb=customtkinter.CTkLabel(fm_payments2,text="Staff No: ",font=fontlbl,text_color=bg1,justify=LEFT,anchor=W)
    lb.grid(column=0,row=0,sticky=W)
    txt_staff_no=customtkinter.CTkEntry(fm_payments2,border_width=1,border_color=bg1,width=150)
    txt_staff_no.grid(column=1,row=0)
    txt_staff_no.bind('<FocusOut>',get_staff_paydetails)
    lb=customtkinter.CTkLabel(fm_payments2,text="Staff Name: ",font=fontlbl,text_color=bg1,justify=LEFT,anchor=W)
    lb.grid(column=2,row=0,sticky=W)
    txt_staff_name=customtkinter.CTkEntry(fm_payments2,border_width=1,border_color=bg1,width=150)
    txt_staff_name.grid(column=3,row=0)
    lb=customtkinter.CTkLabel(fm_payments2,text="Amount: ",font=fontlbl,text_color=bg1,justify=LEFT,anchor=W)
    lb.grid(column=0,row=1,sticky=W)
    txt_amount_no=customtkinter.CTkEntry(fm_payments2,border_width=1,border_color=bg1,width=150)
    txt_amount_no.grid(column=1,row=1)
    btn_fm_payments2=customtkinter.CTkFrame(fm_payments2,fg_color=bg3)
    btn_fm_payments2.grid(column=0,row=2,columnspan=3,pady=5)
    btn_cancel_payment=customtkinter.CTkButton(btn_fm_payments2,text="Cancel",fg_color=coldark,hover_color=bg1,text_color=fg1,cursor="hand2",command=cancel_pay_off)
    btn_cancel_payment.grid(column=0,row=0,padx=4)
    btn_payoff=customtkinter.CTkButton(btn_fm_payments2,text="Pay out",fg_color=bg1,hover_color=bg1,text_color=fg1,cursor="hand2",command=pay_off_staffs)
    btn_payoff.grid(column=1,row=0,padx=4)


    fm_approvals=LabelFrame(fmschool_finance,text="Approvals",bg=bg3,fg=bg1)
    fm_approvals.pack(side=LEFT,fill=BOTH,expand=True)
    fm_approvals1=customtkinter.CTkFrame(fm_approvals,fg_color=bg3)
    fm_approvals1.pack(side=TOP,fill=BOTH,expand=True)
    table_approvals=ttk.Treeview(fm_approvals1,height=6)
    table_approvals.pack(side=LEFT,fil=BOTH,expand=True)
    table_approvals['show']="headings"
    table_approvals['columns']=(0,1,2,3)
    table_approvals.heading(0,text="Id")
    table_approvals.heading(1,text="Category")
    table_approvals.heading(2,text="Amount")
    table_approvals.heading(3,text="By")
    table_approvals.column(0,width=30)
    table_approvals.column(1,width=50)
    table_approvals.column(2,width=50)
    table_approvals.column(3,width=40)
    scroll_table_approvals=customtkinter.CTkScrollbar(fm_approvals1,command=table_approvals.yview)
    scroll_table_approvals.pack(side=LEFT,fill=Y)
    table_approvals.configure(yscrollcommand=scroll_table_approvals)
    lb=customtkinter.CTkLabel(fm_approvals,text="Narration: ",font=fontlbl,text_color=bg1,justify=LEFT,anchor=W)
    lb.pack(side=TOP,fill=X,pady=(5,3),padx=4,anchor=W)
    txt_narration=customtkinter.CTkTextbox(fm_approvals,fg_color="white",border_width=1,border_color=bg1,text_color="black",height=80)
    txt_narration.pack(side=TOP,fill=X,pady=5,padx=(4,11))
    btn_approve_cash=customtkinter.CTkButton(fm_approvals,text="Approve",fg_color=bg1,hover_color=bg1,text_color=fg1,cursor="hand2",command=approve_petty_cash)
    btn_approve_cash.pack(side=TOP,pady=(3,3))
    btn_reject_cash=customtkinter.CTkButton(fm_approvals,text="Reject",fg_color=coldark,hover_color=bg1,text_color=fg1,cursor="hand2",command=reject_petty_cash)
    btn_reject_cash.pack(side=TOP,pady=(3,3))

    def get_school_finance_analysis():
        try:
            cursor=conn.cursor()
            sql1="SELECT SUM(amount) FROM fees"
            cursor.execute(sql1)
            results=cursor.fetchall()
            if results:
                lbl_total_money.configure(text=f"Total Cash\n {results[0][0]}")
            else:
                lbl_total_money.configure(text=f"Total Cash\n 0.0")

            sql2="SELECT SUM(amount_paid) FROM salary_payment"
            cursor.execute(sql2)
            results2=cursor.fetchall()
            if results2:
                lbl_total_salaries.configure(text=f"Total Salaries\n {results2[0][0]}")
            else:
                lbl_total_salaries.configure(text=f"Total Salaries\n 0.0")

            sql3="SELECT SUM(ptamount) FROM pettycash"
            cursor.execute(sql3)
            results3=cursor.fetchall()
            if results3:
                lbl_total_pettycash.configure(text=f"Total PettyCash\n {results3[0][0]}")
            else:
                lbl_total_pettycash.configure(text=f"Total PettyCash\n 0.0")
            try:
                available_amount=results[0][0]-(results2[0][0]+results3[0][0])
                lbl_total_availablecash.configure(text=f"Available Amount\n {available_amount}")
            except:
                lbl_total_availablecash.configure(text=f"Available Amount\n 0.0")
        except IOError:
            pass
    def get_reporting_analysis():
        try:
            r_termid=""
            r_reported=""
            cursor=conn.cursor()
            sql="SELECT termid,termname FROM terms WHERE termstatus=1"
            cursor.execute(sql)
            results=cursor.fetchall()
            if results:
                r_termid=results[0][0]
                lbl_Current_term.configure(text=f"Current Term\n {results[0][1]}")
            else:
                lbl_Current_term.configure(text=f"Current Term\n Term Not created")

            if r_termid!="":
                sql2="SELECT COUNT(reportingid) FROM reporting WHERE reporttermid=%s"
                cursor.execute(sql2%(r_termid))
                results2=cursor.fetchall()
                if results2:
                    r_reported=results2[0][0]
                    lbl_total_reported.configure(text=f"Total Reporting\n {results2[0][0]}")
                else:
                    lbl_total_reported.configure(text=f"Total Reporting\n 0")

                sql3="SELECT SUM(amount) FROM termfee WHERE termid=%s"
                cursor.execute(sql3%(r_termid))
                results3=cursor.fetchall()
                if results3:
                    expected_amount=int(r_reported)*int(results3[0][0])
                    lbl_total_expected_fee.configure(text=f"Amount Expected\n {expected_amount}")
                else:
                    lbl_total_expected_fee.configure(text=f"Amount Expected\n 0.0")


        except IOError:
            pass
    def get_resource_analysis():
        try:
            cursor=conn.cursor()
            sql="SELECT SUM(quantity) FROM books"
            cursor.execute(sql)
            results=cursor.fetchall()
            if results:
                lb_cur_total_books.configure(text=f"Total Books\n {results[0][0]}")
            else:
                lb_cur_total_books.configure(text=f"Total Books\n 0")

            sql2="SELECT COUNT(projectid),SUM(projectcost) FROM projects"
            cursor.execute(sql2)
            results2=cursor.fetchall()
            if results2:
                lb_cur_projects.configure(text=f"Total Projects\n {results2[0][0]}")
                lb_cur_total_projects.configure(text=f"Projects Cost\n {results2[0][1]}")
            else:
                lb_cur_projects.configure(text=f"Total Projects\n 0")
                lb_cur_total_projects.configure(text=f"Projects Cost\n 0.0")

            total_book_costs=0    
            sql3="SELECT bookprice,quantity FROM books"
            cursor.execute(sql3)
            results3=cursor.fetchall()
            if results3:
                i=0
                
                while i<len(results3):
                    total_book_costs=total_book_costs+(results3[i][0]*results3[i][1])
                    
                    if i==len(results3):
                        break
                    i=i+1
                
                lb_total_bookscost.configure(text=f"Total Books Cost\n {total_book_costs}")
            else:
                lb_total_bookscost.configure(text=f"Total Books Cost\n {total_book_costs}")
        except IOError:
            pass

    lb_finance_analysis=LabelFrame(fmschool_analysis,text="Finance Analysis")
    lb_finance_analysis.pack(side=TOP,fill=BOTH,expand=True)
    lbl_total_money=customtkinter.CTkLabel(lb_finance_analysis,text="Total cash",text_color=bg1,font=lbfont3)
    lbl_total_money.pack(side=LEFT,expand=True)
    lbl_total_salaries=customtkinter.CTkLabel(lb_finance_analysis,text="Total Salaries",text_color=bg1,font=lbfont3)
    lbl_total_salaries.pack(side=LEFT,expand=True)
    lbl_total_pettycash=customtkinter.CTkLabel(lb_finance_analysis,text="Total Pettycash",text_color=bg1,font=lbfont3)
    lbl_total_pettycash.pack(side=LEFT,expand=True)
    lbl_total_availablecash=customtkinter.CTkLabel(lb_finance_analysis,text="Available Amount",text_color=bg1,font=lbfont3)
    lbl_total_availablecash.pack(side=LEFT,expand=True)

    lb_reporting_analysis=LabelFrame(fmschool_analysis,text="Reporting Analysis")
    lb_reporting_analysis.pack(side=TOP,fill=BOTH,expand=True)
    lbl_Current_term=customtkinter.CTkLabel(lb_reporting_analysis,text="Current term",text_color=bg1,font=lbfont3)
    lbl_Current_term.pack(side=LEFT,expand=True)
    lbl_total_reported=customtkinter.CTkLabel(lb_reporting_analysis,text="Total Reporting",text_color=bg1,font=lbfont3)
    lbl_total_reported.pack(side=LEFT,expand=True)
    lbl_total_expected_fee=customtkinter.CTkLabel(lb_reporting_analysis,text="Total Expected",text_color=bg1,font=lbfont3)
    lbl_total_expected_fee.pack(side=LEFT,expand=True)
    lbl_total_col_fee=customtkinter.CTkLabel(lb_reporting_analysis,text="",text_color=bg1,font=lbfont3)
    lbl_total_col_fee.pack(side=LEFT,expand=True)    

    lb_resource_analysis=LabelFrame(fmschool_analysis,text="Resources Analysis")
    lb_resource_analysis.pack(side=TOP,fill=BOTH,expand=True)
    lb_cur_total_books=customtkinter.CTkLabel(lb_resource_analysis,text="Total Books",text_color=bg1,font=lbfont3)
    lb_cur_total_books.pack(side=LEFT,expand=True)
    lb_total_bookscost=customtkinter.CTkLabel(lb_resource_analysis,text="Books Cost",text_color=bg1,font=lbfont3)
    lb_total_bookscost.pack(side=LEFT,expand=True)
    lb_cur_projects=customtkinter.CTkLabel(lb_resource_analysis,text="Total Projects",text_color=bg1,font=lbfont3)
    lb_cur_projects.pack(side=LEFT,expand=True)
    lb_cur_total_projects=customtkinter.CTkLabel(lb_resource_analysis,text="Projects cost",text_color=bg1,font=lbfont3)
    lb_cur_total_projects.pack(side=LEFT,expand=True)

    school_info_detail()
    get_all_student_analysis()
    get_all_staff_analysis()
    get_system_analysis()
    get_payment_list()
    get_petty_cash_toapprove()
    table_approvals.bind('<Double-1>',get_petty_cash_record)
    get_school_finance_analysis()
    get_reporting_analysis()
    get_resource_analysis()
    chech_school_type()
    
    #=========framestudents=====================================================================
    framesubmenu1=customtkinter.CTkFrame(framestudents,width=150,fg_color=coldark)
    framesubmenu1.pack(side=LEFT,fill=Y)
    btnaddstudent=customtkinter.CTkButton(framesubmenu1,text="Add student",text_color=fg1,fg_color=bg1,hover_color=hovbg1,cursor="hand2",command=toaddstudent)
    btnaddstudent.pack(pady=6,padx=6)
    btneditstudent=customtkinter.CTkButton(framesubmenu1,text="Edit student",text_color=fg1,fg_color=bg1,hover_color=hovbg1,cursor="hand2",command=toeditstudent)
    btneditstudent.pack(pady=6,padx=6)
    btnviewstudent=customtkinter.CTkButton(framesubmenu1,text="View student",text_color=fg1,fg_color=bg1,hover_color=hovbg1,cursor="hand2",command=toviewstudent)
    btnviewstudent.pack(pady=6,padx=6)
    btngenstudentid=customtkinter.CTkButton(framesubmenu1,text="Student ID",text_color=fg1,fg_color=bg1,hover_color=hovbg1,cursor="hand2",command=tostudentidgenerator)
    btngenstudentid.pack(pady=6,padx=6)
    ntbook1=ttk.Notebook(framestudents)
    ntbook1.pack(side=LEFT,fill=BOTH,expand=True)
    fmaddstudent=customtkinter.CTkFrame(ntbook1,fg_color=bg3,bg_color=coldark)
    fmaddstudent.place(x=0,y=0,relheight=1,relwidth=1)
    fmeditstudent=customtkinter.CTkFrame(ntbook1,fg_color=bg3,bg_color=coldark)
    fmeditstudent.place(x=0,y=0,relheight=1,relwidth=1)
    fmviewstudent=customtkinter.CTkFrame(ntbook1,fg_color=bg3,bg_color=coldark)
    fmviewstudent.place(x=0,y=0,relheight=1,relwidth=1)
    fmgenstudentid=customtkinter.CTkFrame(ntbook1,fg_color=bg3,bg_color=coldark)
    fmgenstudentid.place(x=0,y=0,relheight=1,relwidth=1)
    ntbook1.add(fmaddstudent,text="Add student")
    ntbook1.add(fmeditstudent,text="Edit students")
    ntbook1.add(fmviewstudent,text="View students")
    ntbook1.add(fmgenstudentid,text="Student ID")
    
    textboxwidth=200
    #-----------------add student------------------------
    
    lb=customtkinter.CTkLabel(fmaddstudent,text="Add Student Form",text_color=bg1,font=fontlbl2,justify=CENTER,anchor=CENTER)
    lb.grid(column=0,row=0,pady=(5,5),columnspan=4)
    lb=customtkinter.CTkLabel(fmaddstudent,text="First Name: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=0,row=1,pady=(5,5))
    fname=customtkinter.CTkEntry(fmaddstudent,border_width=1,border_color=bg1,font=fontentries,width=textboxwidth)
    fname.grid(column=1,row=1,pady=(5,5))
    lb=customtkinter.CTkLabel(fmaddstudent,text="Last Name: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=2,row=1,pady=(5,5))
    lname=customtkinter.CTkEntry(fmaddstudent,border_width=1,border_color=bg1,font=fontentries,width=textboxwidth)
    lname.grid(column=3,row=1,pady=(5,5))
    lb=customtkinter.CTkLabel(fmaddstudent,text="Surname: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=0,row=2,pady=(5,5))
    sname=customtkinter.CTkEntry(fmaddstudent,border_width=1,border_color=bg1,font=fontentries,width=textboxwidth)
    sname.grid(column=1,row=2,pady=(5,5))
    lb=customtkinter.CTkLabel(fmaddstudent,text="Gender: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=2,row=2,pady=(5,5))
    gender=customtkinter.CTkOptionMenu(fmaddstudent,fg_color="white",button_color=bg1,button_hover_color=bg1,font=fontentries,text_color="black",width=textboxwidth,values=["Male","Female","Others"])
    gender.grid(column=3,row=2,pady=(5,5))
    lb=customtkinter.CTkLabel(fmaddstudent,text="D.O.B: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=0,row=3,pady=(5,5))
    dob=customtkinter.CTkEntry(fmaddstudent,border_width=1,border_color=bg1,font=fontentries,width=textboxwidth)
    dob.grid(column=1,row=3,pady=(5,5))
    lb=customtkinter.CTkLabel(fmaddstudent,text="Study Mode: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=2,row=3,pady=(5,5))
    studymode=customtkinter.CTkOptionMenu(fmaddstudent,fg_color="white",button_color=bg1,button_hover_color=bg1,font=fontentries,text_color="black",width=textboxwidth,values=["Border","Day-Scholer"])
    studymode.grid(column=3,row=3,pady=(5,5))
    lb=customtkinter.CTkLabel(fmaddstudent,text="Admission No: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=0,row=4,pady=(5,5))
    admno=customtkinter.CTkEntry(fmaddstudent,border_width=1,border_color=bg1,font=fontentries,width=textboxwidth)
    admno.grid(column=1,row=4,pady=(5,5))
    lb=customtkinter.CTkLabel(fmaddstudent,text="Grade: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=2,row=4,pady=(5,5))
    form=customtkinter.CTkOptionMenu(fmaddstudent,fg_color="white",button_color=bg1,button_hover_color=bg1,font=fontentries,text_color="black",width=textboxwidth,values=class_values)
    form.grid(column=3,row=4,pady=(5,5))
    lb=customtkinter.CTkLabel(fmaddstudent,text="Parent: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=0,row=5,pady=(5,5))
    parent=customtkinter.CTkEntry(fmaddstudent,border_width=1,border_color=bg1,font=fontentries,width=textboxwidth)
    parent.grid(column=1,row=5,pady=(5,5))
    lb=customtkinter.CTkLabel(fmaddstudent,text="Phone No: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=2,row=5,pady=(5,5))
    phoneno=customtkinter.CTkEntry(fmaddstudent,border_width=1,border_color=bg1,font=fontentries,width=textboxwidth)
    phoneno.grid(column=3,row=5,pady=(5,5))
    lb=customtkinter.CTkLabel(fmaddstudent,text="Religion: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=0,row=6,pady=(5,5))
    religion=customtkinter.CTkOptionMenu(fmaddstudent,fg_color="white",button_color=bg1,button_hover_color=bg1,font=fontentries,text_color="black",width=textboxwidth,values=["Christian","Muslim","Others"])
    religion.grid(column=1,row=6,pady=(5,5))
    lb=customtkinter.CTkLabel(fmaddstudent,text="County/Location: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=2,row=6,pady=(5,5))
    homelocation=customtkinter.CTkEntry(fmaddstudent,border_width=1,border_color=bg1,font=fontentries,width=textboxwidth)
    homelocation.grid(column=3,row=6,pady=(5,5))
    fmstudentbtnholder=customtkinter.CTkFrame(fmaddstudent)
    fmstudentbtnholder.grid(column=1,row=7,pady=(20,20),columnspan=3)
    btnsavestudent=customtkinter.CTkButton(fmstudentbtnholder,text="Save",fg_color=bg1,text_color=fg1,hover_color=hovbg1,cursor="hand2",command=addstudents)
    btnsavestudent.grid(column=0,row=0,pady=(5,5),padx=5)
    btncancelstudent=customtkinter.CTkButton(fmstudentbtnholder,text="Cancel",fg_color=cancelcol,text_color=fg1,hover_color=hovbg1,cursor="hand2",command=clearstudentform)
    btncancelstudent.grid(column=1,row=0,pady=(5,5),padx=5)
    #----------------------------------------------------
    #--------------edit student--------------------------
    framesearchstudent=customtkinter.CTkFrame(fmeditstudent,fg_color=bg3,bg_color=coldark)
    framesearchstudent.pack(side=TOP,fill=X)
    customtkinter.CTkLabel(framesearchstudent,text="Search student",text_color=bg1,font=fontlbl2,justify=CENTER,anchor=CENTER).grid(column=0,row=0,columnspan=5)
    customtkinter.CTkLabel(framesearchstudent,text="Search By: ",font=fontlbl,text_color=bg1).grid(column=0,row=1)
    searchstdby=customtkinter.CTkOptionMenu(framesearchstudent,fg_color="white",button_color=bg1,button_hover_color=bg1,text_color="black",font=fontentries,width=120,values=["Admission","PhoneNo"])
    searchstdby.grid(column=1,row=1)
    customtkinter.CTkLabel(framesearchstudent,text="Search Input: ",font=fontlbl,text_color=bg1).grid(column=2,row=1)
    searchtext=customtkinter.CTkEntry(framesearchstudent,border_width=1,border_color=bg1,font=fontentries,width=120)
    searchtext.grid(column=3,row=1)
    btnsearchstd=customtkinter.CTkButton(framesearchstudent,text="search",fg_color=bg1,hover_color=hovbg1,text_color=fg1,cursor="hand2",command=searchstudent)
    btnsearchstd.grid(column=4,row=1,padx=(4,3))

    framesearchedstudent=customtkinter.CTkFrame(fmeditstudent,fg_color=bg3,bg_color=coldark)
    framesearchedstudent.pack(side=TOP,fill=BOTH,expand=True)
    lb=customtkinter.CTkLabel(framesearchedstudent,text="Edit Student Form",text_color=bg1,font=fontlbl2,justify=CENTER,anchor=CENTER)
    lb.grid(column=0,row=0,pady=(5,5),columnspan=4)
    lb=customtkinter.CTkLabel(framesearchedstudent,text="First Name: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=0,row=1,pady=(5,5))
    efname=customtkinter.CTkEntry(framesearchedstudent,border_width=1,border_color=bg1,font=fontentries,width=textboxwidth)
    efname.grid(column=1,row=1,pady=(5,5))
    lb=customtkinter.CTkLabel(framesearchedstudent,text="Last Name: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=2,row=1,pady=(5,5))
    elname=customtkinter.CTkEntry(framesearchedstudent,border_width=1,border_color=bg1,font=fontentries,width=textboxwidth)
    elname.grid(column=3,row=1,pady=(5,5))
    lb=customtkinter.CTkLabel(framesearchedstudent,text="Surname: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=0,row=2,pady=(5,5))
    esname=customtkinter.CTkEntry(framesearchedstudent,border_width=1,border_color=bg1,font=fontentries,width=textboxwidth)
    esname.grid(column=1,row=2,pady=(5,5))
    lb=customtkinter.CTkLabel(framesearchedstudent,text="Gender: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=2,row=2,pady=(5,5))
    egender=customtkinter.CTkOptionMenu(framesearchedstudent,fg_color="white",button_color=bg1,button_hover_color=bg1,text_color="black",font=fontentries,width=textboxwidth,values=["Male","Female","Others"])
    egender.grid(column=3,row=2,pady=(5,5))
    lb=customtkinter.CTkLabel(framesearchedstudent,text="D.O.B: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=0,row=3,pady=(5,5))
    edob=customtkinter.CTkEntry(framesearchedstudent,border_width=1,border_color=bg1,font=fontentries,width=textboxwidth)
    edob.grid(column=1,row=3,pady=(5,5))
    lb=customtkinter.CTkLabel(framesearchedstudent,text="Study Mode: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=2,row=3,pady=(5,5))
    estudymode=customtkinter.CTkOptionMenu(framesearchedstudent,fg_color="white",button_color=bg1,button_hover_color=bg1,text_color="black",font=fontentries,width=textboxwidth,values=["Border","Day-Scholer"])
    estudymode.grid(column=3,row=3,pady=(5,5))
    lb=customtkinter.CTkLabel(framesearchedstudent,text="Admission No: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=0,row=4,pady=(5,5))
    eadmno=customtkinter.CTkEntry(framesearchedstudent,border_width=1,border_color=bg1,font=fontentries,width=textboxwidth)
    eadmno.grid(column=1,row=4,pady=(5,5))
    lb=customtkinter.CTkLabel(framesearchedstudent,text="Grade: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=2,row=4,pady=(5,5))
    eform=customtkinter.CTkOptionMenu(framesearchedstudent,fg_color="white",button_color=bg1,button_hover_color=bg1,text_color="black",font=fontentries,width=textboxwidth,values=class_values)
    eform.grid(column=3,row=4,pady=(5,5))
    lb=customtkinter.CTkLabel(framesearchedstudent,text="Parent: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=0,row=5,pady=(5,5))
    eparent=customtkinter.CTkEntry(framesearchedstudent,border_width=1,border_color=bg1,font=fontentries,width=textboxwidth)
    eparent.grid(column=1,row=5,pady=(5,5))
    lb=customtkinter.CTkLabel(framesearchedstudent,text="Phone No: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=2,row=5,pady=(5,5))
    ephoneno=customtkinter.CTkEntry(framesearchedstudent,border_width=1,border_color=bg1,font=fontentries,width=textboxwidth)
    ephoneno.grid(column=3,row=5,pady=(5,5))
    lb=customtkinter.CTkLabel(framesearchedstudent,text="Religion: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=0,row=6,pady=(5,5))
    ereligion=customtkinter.CTkOptionMenu(framesearchedstudent,fg_color="white",button_color=bg1,button_hover_color=bg1,text_color="black",font=fontentries,width=textboxwidth,values=["Christian","Muslim","Others"])
    ereligion.grid(column=1,row=6,pady=(5,5))
    lb=customtkinter.CTkLabel(framesearchedstudent,text="County/Location: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=2,row=6,pady=(5,5))
    ehomelocation=customtkinter.CTkEntry(framesearchedstudent,border_width=1,border_color=bg1,font=fontentries,width=textboxwidth)
    ehomelocation.grid(column=3,row=6,pady=(5,5))
    efmstudentbtnholder=customtkinter.CTkFrame(framesearchedstudent)
    efmstudentbtnholder.grid(column=1,row=7,pady=(20,20),columnspan=3)
    ebtnsavestudent=customtkinter.CTkButton(efmstudentbtnholder,text="Update",fg_color=bg1,text_color=fg1,hover_color=hovbg1,cursor="hand2",command=updatestudent)
    ebtnsavestudent.grid(column=0,row=0,pady=(5,5),padx=5)
    ebtndeletestudent=customtkinter.CTkButton(efmstudentbtnholder,text="Delete",fg_color=cancelcol,text_color=fg1,hover_color=hovbg1,cursor="hand2",command=deletestudent)
    ebtndeletestudent.grid(column=1,row=0,pady=(5,5),padx=5)
    ebtncancelstudent=customtkinter.CTkButton(efmstudentbtnholder,text="Cancel",fg_color=coldark,text_color=fg1,hover_color=hovbg1,cursor="hand2",command=clearstudenteditform)
    ebtncancelstudent.grid(column=2,row=0,pady=(5,5),padx=5)
    #----------------------------------------------------
    #-----------------view students----------------------
    tablestudents=ttk.Treeview(fmviewstudent)
    tablestudents.pack(side=TOP,fill=BOTH,expand=True)
    tablestudents['columns']=(0,1,2,3,4,5,6,7)
    tablestudents['show']="headings"
    tablestudents.heading(0,text="Admno")
    tablestudents.heading(1,text="Full Name")
    tablestudents.heading(2,text="Gender")
    tablestudents.heading(3,text="Parent")
    tablestudents.heading(4,text="Phone No")
    tablestudents.heading(5,text="Form")
    tablestudents.heading(6,text="Study Mode")
    tablestudents.heading(7,text="Religion")
    tablestudents.column(0,width=60)
    tablestudents.column(1,width=130)
    tablestudents.column(2,width=60)
    tablestudents.column(3,width=100)
    tablestudents.column(4,width=90)
    tablestudents.column(5,width=70)
    tablestudents.column(6,width=70)
    tablestudents.column(7,width=70)
    viewstudents()
    #----------------generate student id-----------------
    #function

    ft=ImageFont.truetype('times.ttf',50)
    ft2=ImageFont.truetype('times.ttf',35)
    front_template_folder="id_templates/front_template/"
    front_template_list=[]
    for file in os.listdir(front_template_folder):
        front_template_list.append(file)
    
    custom_student_photo=customtkinter.CTkImage(light_image=Image.open("icons/custom_user.png"),size=(100,100))
    def select_photo_fun():
        global student_pic
        sel_photo=filedialog.askopenfilename(initialdir="",title="Select Photo",filetype=[
                ("Images","*.png;*.jpg;*.jpeg;*.gif;*.bmp")
            ])
        if sel_photo:
            student_pic=sel_photo
            change_pic=customtkinter.CTkImage(light_image=Image.open(sel_photo),size=(100,100))
            lbl_student_photo.configure(image=change_pic)
        else:
            messagebox.showwarning("Warning","Student Photo not selected")
    def select_logo_fun():
        global school_logo
        sel_logo=filedialog.askopenfilename(initialdir="",title="Select Logo",filetype=[
                ("Images","*.png;*.jpg;*.jpeg;*.gif;*.bmp")
            ])
        if sel_logo:
            school_logo=sel_logo
        else:
            messagebox.showwarning("Warning","School Logo not selected")
    def display_front_template():
        global my_front_template
        if option_templates.get()=="":
            messagebox.showwarning("Warning","No template on this system")
        else:
            my_front_template=os.path.join(front_template_folder,option_templates.get())
            front_temp_preview=customtkinter.CTkImage(light_image=Image.open(my_front_template),size=(300,190))
            lbl_front_view.configure(image=front_temp_preview)
    def display_front_template1(*args):
        global my_front_template
        if option_templates.get()=="":
            messagebox.showwarning("Warning","No template on this system")
        else:
            my_front_template=os.path.join(front_template_folder,option_templates.get())
            front_temp_preview=customtkinter.CTkImage(light_image=Image.open(my_front_template),size=(300,190))
            lbl_front_view.configure(image=front_temp_preview)
            display_back_template()
    def display_back_template():
        global my_back_template
        my_back_template="id_templates/back_template/STUDENT_ID_CARD_BACK.png"
        back_temp_preview=customtkinter.CTkImage(light_image=Image.open(my_back_template),size=(300,190))
        lbl_back_view.configure(image=back_temp_preview)
    def apply_school_info():
        global school_logo,my_front_template,my_back_template,template_f,template_b,my_sch_sign,tmp_color
        if sel_id_color.get()=="black":
            tmp_color=(0,0,0,255)
        elif sel_id_color.get()=="blue":
            tmp_color=(0,0,255,255)
        elif sel_id_color.get()=="darkblue":
            tmp_color=(0,0,136,255)
        elif sel_id_color.get()=="orange":
            tmp_color=(255,165,0,255)
        elif sel_id_color.get()=="yellow":
            tmp_color=(255,255,0,255)
        elif sel_id_color.get()=="white":
            tmp_color=(255,255,255,255)
        elif sel_id_color.get()=="lightblue":
            tmp_color=(173,216,230,255)
        else:
            tmp_color=(0,0,0,0)

        school_name=school_name_txt.get()
        school_motto=school_motto_txt.get()
        school_contact=school_contact_txt.get()
        school_box=school_box_txt.get()

        template_f=Image.open(my_front_template)
        d=ImageDraw.Draw(template_f)
        d.text((290,50),school_name,font=ft,fill=tmp_color)
        d.text((290,600),school_motto,font=ft2,fill=tmp_color)

        template_b=Image.open(my_back_template)
        e=ImageDraw.Draw(template_b)
        applytext=f"This card is a property of {school_name} \nand must be surrendered back at the end off service"
        e.text((220,565),school_box,font=ft2,fill=(0,0,0,255))
        e.text((620,565),school_contact,font=ft2,fill=(0,0,0,255))
        e.text((50,50),applytext,font=ft2,fill=(0,0,0,255))
        e.text((560,405),my_sch_sign,font=ft2,fill=(0,0,0,255))
        if school_logo!="":
            school_logo_r=Image.open(school_logo).resize((164,164))
            template_f.paste(school_logo_r,(5,5))
        else:
            pass
        front_temp_preview=customtkinter.CTkImage(light_image=template_f,size=(300,190))
        lbl_front_view.configure(image=front_temp_preview)
        back_temp_preview=customtkinter.CTkImage(light_image=template_b,size=(300,190))
        lbl_back_view.configure(image=back_temp_preview)
    def apply_student_info():
        global template_f,template_b,student_pic,tmp_color
        if template_b=="" and template_f=="":
            messagebox.showerror("Error","School Details Not Applied")
            return False
        else:
            student_name=student_name_txt.get()
            student_admno=student_admno_txt.get()
            student_dob=student_dob_txt.get()
            student_jyear=student_jyear_txt.get()
        if student_pic!="":
            d=ImageDraw.Draw(template_f)
            d.text((540,270),student_name,font=ft2,fill=(0,0,0,255))
            d.text((540,328),student_admno,font=ft2,fill=(0,0,0,255))
            d.text((540,390),student_dob,font=ft2,fill=(0,0,0,255))
            d.text((540,450),student_jyear,font=ft2,fill=(0,0,0,255))

            student_pic_r=Image.open(student_pic).resize((265,344))
            template_f.paste(student_pic_r,(35,197))

            qr=qrcode.QRCode(version=1,box_size=14,border=2)
            qr.add_data(student_admno)
            qr.make(fit=True)
            qr_img=qr.make_image(fill="black",back_color="white")

            e=ImageDraw.Draw(template_b)
            template_b.paste(qr_img,(80,200))

            front_temp_preview=customtkinter.CTkImage(light_image=template_f,size=(300,190))
            lbl_front_view.configure(image=front_temp_preview)
            back_temp_preview=customtkinter.CTkImage(light_image=template_b,size=(300,190))
            lbl_back_view.configure(image=back_temp_preview)
        else:
            select_photo_fun()
    def clear_form_student_id():
        student_name_txt.delete(0,END)
        student_admno_txt.delete(0,END)
        student_dob_txt.delete(0,END)
        student_jyear_txt.delete(0,END)
    def clear_details():
        global my_front_template,my_back_template,template_f,template_b,student_pic
        msg=messagebox.askyesnocancel("Clear message","Do you want To clear All Detail\n 1: YES to clear Student data and template\n 2:NO to clear only template")
        if msg==1:
            clear_form_student_id()
            lbl_student_photo.configure(image=custom_student_photo)
            my_front_template=os.path.join(front_template_folder,option_templates.get())
            front_temp_preview=customtkinter.CTkImage(light_image=Image.open(my_front_template),size=(300,190))
            lbl_front_view.configure(image=front_temp_preview)
            my_back_template="id_templates/back_template/STUDENT_ID_CARD_BACK.png"
            back_temp_preview=customtkinter.CTkImage(light_image=Image.open(my_back_template),size=(300,190))
            lbl_back_view.configure(image=back_temp_preview)
            template_f=Image.open(my_front_template)
            template_b=Image.open(my_back_template)
            student_pic=""
        elif msg==0:
            my_front_template=os.path.join(front_template_folder,option_templates.get())
            front_temp_preview=customtkinter.CTkImage(light_image=Image.open(my_front_template),size=(300,190))
            lbl_front_view.configure(image=front_temp_preview)
            my_back_template="id_templates/back_template/STUDENT_ID_CARD_BACK.png"
            back_temp_preview=customtkinter.CTkImage(light_image=Image.open(my_back_template),size=(300,190))
            lbl_back_view.configure(image=back_temp_preview)
            template_f=Image.open(my_front_template)
            template_b=Image.open(my_back_template)
    def search_studentDetails_id(*arg):
        s_admno=student_admno_txt.get()
        if s_admno=="":
            pass
        else:
            try:
                cursor=conn.cursor()
                sql="SELECT * FROM students WHERE admno='%s'"%(s_admno)
                cursor.execute(sql)
                results=cursor.fetchall()
                if results:
                    clear_form_student_id()
                    student_admno_txt.insert(END,results[0][7])
                    student_name_txt.insert(END,results[0][1]+" "+results[0][2]+" "+results[0][3])
                    student_dob_txt.insert(END,results[0][5].strftime("%d")+"/"+results[0][5].strftime("%m")+"/"+results[0][5].strftime("%Y"))
                    student_jyear_txt.insert(END,results[0][14].strftime("%Y"))
                else:
                    messagebox.showerror("Error","Invalid Student Admission No")
            except IOError:
                pass

    def save_id_generated():
        global template_f,template_b
        pdf=FPDF('P','mm',(100,80))
        pdf.add_page()
        pdf.set_title("Student ID card"+student_admno_txt.get())
        pdf.set_author("Eric Software Solutions")
        pdf.set_auto_page_break(auto=True,margin=0)
        pdf.image(template_f,x=-0.5,w=pdf.w+1)
        pdf.image(template_b,x=-0.5,w=pdf.w+1)
        pdf.output("id_card/"+student_admno_txt.get()+".pdf")
        file_to_print="\\id_card\\"+student_admno_txt.get()+".pdf"
        current_folder=os.getcwd()
        file_to_print=f'{current_folder}'+file_to_print
        #open created file
        webbrowser.open(file_to_print)
        
    def insert_id_school_details():
        global list_school_details,my_sch_sign
        school_name_txt.insert(END,list_school_details[0][1])
        school_motto_txt.insert(END,list_school_details[0][2])
        school_contact_txt.insert(END,list_school_details[0][3])
        school_box_txt.insert(END,list_school_details[0][4])
        my_sch_sign=list_school_details[0][6]

    frame_main_student_id=customtkinter.CTkFrame(fmgenstudentid,fg_color=bg3)
    frame_main_student_id.pack(fill=BOTH,side=LEFT,expand=True)
    frame_school_details=customtkinter.CTkFrame(frame_main_student_id,fg_color=bg3)
    frame_school_details.pack(side=TOP,fill=X,pady=6,padx=5)
    lb=customtkinter.CTkLabel(frame_school_details,text="School Name: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=W)
    lb.grid(column=0,row=0,sticky=E)
    school_name_txt=customtkinter.CTkEntry(frame_school_details,width=190)
    school_name_txt.grid(column=1,row=0)
    lb=customtkinter.CTkLabel(frame_school_details,text="Motto: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=W)
    lb.grid(column=2,row=0,sticky=E)
    school_motto_txt=customtkinter.CTkEntry(frame_school_details,width=190)
    school_motto_txt.grid(column=3,row=0)
    lb=customtkinter.CTkLabel(frame_school_details,text="Contact: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=0,row=1,sticky=E)
    school_contact_txt=customtkinter.CTkEntry(frame_school_details,width=190)
    school_contact_txt.grid(column=1,row=1)
    lb=customtkinter.CTkLabel(frame_school_details,text="Box: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=2,row=1,sticky=E)
    school_box_txt=customtkinter.CTkEntry(frame_school_details,width=190)
    school_box_txt.grid(column=3,row=1)
    sel_id_color=customtkinter.CTkOptionMenu(frame_school_details,button_color=bg1,button_hover_color=bg1,fg_color="white",text_color="black",values=["black","white","darkblue","blue","orange","lightblue","yellow"])
    sel_id_color.grid(column=4,row=0,padx=8,pady=4)
    btn_apply1=customtkinter.CTkButton(frame_school_details,text="Apply",fg_color=bg1,text_color=fg1,hover_color=hovbg1,cursor="hand2",command=apply_school_info)
    btn_apply1.grid(column=4,row=1,padx=8,pady=4)
    
    frame_main_stid=customtkinter.CTkFrame(frame_main_student_id,fg_color=bg3)
    frame_main_stid.pack(side=TOP,fill=BOTH,expand=True)
    
    frame_student_details=LabelFrame(frame_main_stid,bg=bg3,fg=bg1,text="Student Details",font=fontlbl)
    frame_student_details.pack(side=LEFT,fill=BOTH,expand=True,padx=6,ipadx=6)
    lb=customtkinter.CTkLabel(frame_student_details,text="Student Admno",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=W)
    lb.grid(column=0,row=0,sticky=W,padx=6)
    student_admno_txt=customtkinter.CTkEntry(frame_student_details,width=200,border_width=1,border_color=bg1)
    student_admno_txt.grid(column=0,row=1,padx=6,columnspan=2)
    student_admno_txt.bind('<FocusOut>',search_studentDetails_id)
    lb=customtkinter.CTkLabel(frame_student_details,text="Student Name",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=W)
    lb.grid(column=0,row=2,sticky=W,padx=6)
    student_name_txt=customtkinter.CTkEntry(frame_student_details,width=200,border_width=1,border_color=bg1)
    student_name_txt.grid(column=0,row=3,padx=6,columnspan=2)
    lb=customtkinter.CTkLabel(frame_student_details,text="D.O.B",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=W)
    lb.grid(column=0,row=4,sticky=W,padx=6)
    student_dob_txt=customtkinter.CTkEntry(frame_student_details,width=200,border_width=1,border_color=bg1)
    student_dob_txt.grid(column=0,row=5,padx=6,columnspan=2)
    lb=customtkinter.CTkLabel(frame_student_details,text="Year",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=W)
    lb.grid(column=0,row=6,sticky=W,padx=6)
    student_jyear_txt=customtkinter.CTkEntry(frame_student_details,width=200,border_width=1,border_color=bg1)
    student_jyear_txt.grid(column=0,row=7,padx=6,columnspan=2)
    
    frame_image_sidgen=customtkinter.CTkFrame(frame_student_details,fg_color=bg3)
    frame_image_sidgen.grid(column=0,columnspan=2,row=8,pady=(7,5),padx=6)
    lbl_student_photo=customtkinter.CTkLabel(frame_image_sidgen,text="",image=custom_student_photo)
    lbl_student_photo.pack(fill=BOTH,expand=True)

    framebtn_idgen=customtkinter.CTkFrame(frame_student_details,fg_color=bg3)
    framebtn_idgen.grid(column=0,columnspan=2,row=9,pady=(7,5),padx=6)
    btn_select_photo=customtkinter.CTkButton(framebtn_idgen,text="Select Photo",fg_color=bg1,width=100,text_color=fg1,hover_color=hovbg1,cursor="hand2",command=select_photo_fun)
    btn_select_photo.grid(column=0,row=0,padx=4,pady=3)
    btn_apply2=customtkinter.CTkButton(framebtn_idgen,text="Apply",fg_color=bg1,width=100,text_color=fg1,hover_color=hovbg1,cursor="hand2",command=apply_student_info)
    btn_apply2.grid(column=1,row=0,padx=4,pady=3)
    btn_clear=customtkinter.CTkButton(framebtn_idgen,text="Clear",fg_color=bg1,width=100,text_color=fg1,hover_color=hovbg1,cursor="hand2",command=clear_details)
    btn_clear.grid(column=0,row=1,padx=4,pady=3)
    btn_save=customtkinter.CTkButton(framebtn_idgen,text="Save",fg_color=bg1,width=100,text_color=fg1,hover_color=hovbg1,cursor="hand2",command=save_id_generated)
    btn_save.grid(column=1,row=1,padx=4,pady=3)
    #-----template previews-----------------------------
    frame_id_preview=LabelFrame(frame_main_stid,bg=bg3,fg=bg1,text="Template Preview")
    frame_id_preview.pack(side=LEFT,fill=BOTH,expand=True,padx=6)
    frame_select_template=customtkinter.CTkFrame(frame_id_preview,fg_color=bg3)
    frame_select_template.pack(side=TOP,pady=(4,4))
    lb=customtkinter.CTkLabel(frame_select_template,text="Select Template: ",text_color=bg1,font=("times",12,"bold"))
    lb.grid(column=0,row=0,padx=(6,4),pady=4)
    option_templates=customtkinter.CTkOptionMenu(frame_select_template,fg_color="white",text_color="black",font=fontlbl,button_color=bg1,button_hover_color=bg1,width=190,values=front_template_list,command=display_front_template1)
    option_templates.grid(column=1,row=0,padx=(6,4),pady=4)
    lbl_front_view=customtkinter.CTkLabel(frame_id_preview,text="")
    lbl_front_view.pack(side=TOP,fill=BOTH,expand=True,pady=5,padx=5)
    lbl_back_view=customtkinter.CTkLabel(frame_id_preview,text="")
    lbl_back_view.pack(side=TOP,fill=BOTH,expand=True,pady=5,padx=5)
    display_front_template()
    display_back_template()
    insert_id_school_details()
    #----------------------------------------------------

    #==========================================
    #=========framestaffs====================
    def toaddstaff():
        ntbook2.select(0)
    def toeditstaff():
        ntbook2.select(1)
    def toviewstaffs():
        ntbook2.select(2)
    def searchstaff():
        global stfid,myusername
        if searchstfby.get()=="StaffNo":
            try:
                sql="SELECT * FROM staffs WHERE staffno='%s'"%(ssearchtext.get())
                cursor=conn.cursor()
                cursor.execute(sql)
                results=cursor.fetchall()
                if results:
                    clearstaffeditform()
                    stfid=results[0][0]
                    esfname.insert(END,results[0][1])
                    eslname.insert(END,results[0][2])
                    essname.insert(END,results[0][3])
                    esidno.insert(END,results[0][4])
                    esphoneno.insert(END,results[0][5])
                    esemail.insert(END,results[0][6])
                    esstafftype.set(results[0][7])
                    estaffno.insert(END,results[0][8])
                    esoccupation.set(results[0][9])
                    esemployer.set(results[0][10])
                    esreligion.set(results[0][11])
                    eshomelocation.insert(END,results[0][12])
                    
            except IOError:
                pass
        elif searchstfby.get()=="IDNo":
            try:
                sql="SELECT * FROM staffs WHERE idno='%s'"%(ssearchtext.get())
                cursor=conn.cursor()
                cursor.execute(sql)
                results=cursor.fetchall()
                if results:
                    clearstaffeditform()
                    stfid=results[0][0]
                    esfname.insert(END,results[0][1])
                    eslname.insert(END,results[0][2])
                    essname.insert(END,results[0][3])
                    esidno.insert(END,results[0][4])
                    esphoneno.insert(END,results[0][5])
                    esemail.insert(END,results[0][6])
                    esstafftype.set(results[0][7])
                    estaffno.insert(END,results[0][8])
                    esoccupation.set(results[0][9])
                    esemployer.set(results[0][10])
                    esreligion.set(results[0][11])
                    eshomelocation.insert(END,results[0][12])
            except IOError:
                pass
    def viewstaffs():
        try:
            cursor=conn.cursor()
            sql="SELECT staffno,CONCAT(fname,' ',lname,' ',sname) AS name,idno,phoneno,email,stafftype,occupation,employer FROM staffs"
            cursor.execute(sql)
            results=cursor.fetchall()
            if results:
                for records in tablestaffs.get_children():
                    tablestaffs.delete(records)
                for i in results:
                    tablestaffs.insert('',END,values=i)
        except IOError:
            pass
    def addstaff():
        if sfname.get()=="":
            return False
        elif slname.get()=="":
            return False
        elif ssname.get()=="":
            return False
        elif sidno.get()=="":
            return False
        elif sphoneno.get()=="":
            return False
        elif semail.get()=="":
            return False
        elif sstafftype.get()=="":
            return False
        elif staffno.get()=="":
            return False
        elif soccupation.get()=="":
            return False
        elif semployer.get()=="":
            return False
        elif sreligion.get()=="":
            return False
        elif shomelocation.get()=="":
            return False
        else:
            global myusername
            try:
                sql="INSERT INTO staffs(fname,lname,sname,idno,phoneno,email,stafftype,staffno,occupation,employer,religion,homelocation,recordby,recorddate) VALUES ('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s')"
                cursor=conn.cursor()
                cursor.execute(sql%(sfname.get(),slname.get(),ssname.get(),sidno.get(),sphoneno.get(),semail.get(),sstafftype.get(),staffno.get(),soccupation.get(),semployer.get(),sreligion.get(),homelocation.get(),myusername,today))
                conn.commit()
                clearstaffform()
                viewstaffs()
                get_all_staff_analysis()
                messagebox.showinfo("Success message","Staff saved successfully")
            except IOError:
                messagebox.showerror("Alert message","Unable to process.\nTry again")
                pass
    def updatestaff():
        if esfname.get()=="":
            return False
        elif eslname.get()=="":
            return False
        elif essname.get()=="":
            return False
        elif esidno.get()=="":
            return False
        elif esphoneno.get()=="":
            return False
        elif esemail.get()=="":
            return False
        elif esstafftype.get()=="":
            return False
        elif estaffno.get()=="":
            return False
        elif esoccupation.get()=="":
            return False
        elif esemployer.get()=="":
            return False
        elif esreligion.get()=="":
            return False
        elif eshomelocation.get()=="":
            return False
        else:
            global stfid,myusername
            if stfid!="":
                try:
                    sql="UPDATE staffs SET fname='%s',lname='%s',sname='%s',idno='%s',phoneno='%s',email='%s',stafftype='%s',staffno='%s',occupation='%s',employer='%s',religion='%s',homelocation='%s',recordby='%s' WHERE staffid=%s"%(esfname.get(),eslname.get(),essname.get(),esidno.get(),esphoneno.get(),esemail.get(),esstafftype.get(),estaffno.get(),esoccupation.get(),esemployer.get(),esreligion.get(),eshomelocation.get(),myusername,stfid)
                    cursor=conn.cursor()
                    cursor.execute(sql)
                    conn.commit()
                    stfid=""
                    ssearchtext.delete(0,END)
                    clearstaffeditform()
                    viewstaffs()
                    get_all_staff_analysis()
                    messagebox.showinfo("Success","Staff Updated successfully")
                except IOError:
                    messagebox.showwarning("Error","Unable to update.\n try Again")
                    pass
            else:
                messagebox.showwarning("Alert","Search staff to Update")
    def deletestaff():
        global stfid
        if stfid=="":
            messagebox.showwarning("Alert","Search staff to delete")
        else:
            msgtodelete=messagebox.askyesno("Alert question","do you want to detete staff staffno "+ssearchtext.get())
            if msgtodelete==True:
                try:
                    sql="DELETE FROM staffs WHERE staffid=%s"%(stfid)
                    cursor=conn.cursor()
                    cursor.execute(sql)
                    conn.commit()
                    stfid=""
                    clearstaffeditform()
                    viewstaffs()
                    get_all_staff_analysis()
                    messagebox.showinfo("success","Staff deleted successful")
                    ssearchtext.delete(0,END)
                except IOError:
                    messagebox.showwarning("Error","Unable to delete.\n try Again")
                    pass
    def clearstaffform():
        sfname.delete(0,END)
        slname.delete(0,END)
        ssname.delete(0,END)
        sidno.delete(0,END)
        sphoneno.delete(0,END)
        semail.delete(0,END)
        sstafftype.set("")
        staffno.delete(0,END)
        soccupation.set("")
        semployer.set("")
        sreligion.set("")
        shomelocation.delete(0,END)
    def clearstaffeditform():
        ssearchtext.delete(0,END)
        global stfid
        stfid=""
        esfname.delete(0,END)
        eslname.delete(0,END)
        essname.delete(0,END)
        esidno.delete(0,END)
        esphoneno.delete(0,END)
        esemail.delete(0,END)
        esstafftype.set("")
        estaffno.delete(0,END)
        esoccupation.set("")
        esemployer.set("")
        esreligion.set("")
        eshomelocation.delete(0,END)
    framesubmenu2=customtkinter.CTkFrame(framestaffs,width=150,fg_color=coldark)
    framesubmenu2.pack(side=LEFT,fill=Y)
    btnaddstaff=customtkinter.CTkButton(framesubmenu2,text="Add staffs",text_color=fg1,fg_color=bg1,hover_color=hovbg1,cursor="hand2",command=toaddstaff)
    btnaddstaff.pack(pady=6,padx=6)
    btneditstaff=customtkinter.CTkButton(framesubmenu2,text="Edit staffs",text_color=fg1,fg_color=bg1,hover_color=hovbg1,cursor="hand2",command=toeditstaff)
    btneditstaff.pack(pady=6,padx=6)
    btnviewstaff=customtkinter.CTkButton(framesubmenu2,text="View staffs",text_color=fg1,fg_color=bg1,hover_color=hovbg1,cursor="hand2",command=toviewstaffs)
    btnviewstaff.pack(pady=6,padx=6)
    ntbook2=ttk.Notebook(framestaffs)
    ntbook2.pack(side=LEFT,fill=BOTH,expand=True)
    fmaddstaffs=customtkinter.CTkFrame(ntbook2,fg_color=bg3,bg_color=coldark)
    fmaddstaffs.place(x=0,y=0,relheight=1,relwidth=1)
    fmeditstaffs=customtkinter.CTkFrame(ntbook2,fg_color=bg3,bg_color=coldark)
    fmeditstaffs.place(x=0,y=0,relheight=1,relwidth=1)
    fmviewstaffs=customtkinter.CTkFrame(ntbook2,fg_color=bg3,bg_color=coldark)
    fmviewstaffs.place(x=0,y=0,relheight=1,relwidth=1)
    ntbook2.add(fmaddstaffs,text="Add staffs")
    ntbook2.add(fmeditstaffs,text="Edit staffs")
    ntbook2.add(fmviewstaffs,text="View staffs")
    #------------add staff---------------------
    lb=customtkinter.CTkLabel(fmaddstaffs,text="Add Staff Form",text_color=bg1,font=fontlbl2,justify=CENTER,anchor=CENTER)
    lb.grid(column=0,row=0,pady=(5,5),columnspan=4)
    lb=customtkinter.CTkLabel(fmaddstaffs,text="First Name: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=0,row=1,pady=(5,5))
    sfname=customtkinter.CTkEntry(fmaddstaffs,border_width=1,border_color=bg1,font=fontentries,width=textboxwidth)
    sfname.grid(column=1,row=1,pady=(5,5))
    lb=customtkinter.CTkLabel(fmaddstaffs,text="Last Name: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=2,row=1,pady=(5,5))
    slname=customtkinter.CTkEntry(fmaddstaffs,border_width=1,border_color=bg1,font=fontentries,width=textboxwidth)
    slname.grid(column=3,row=1,pady=(5,5))
    lb=customtkinter.CTkLabel(fmaddstaffs,text="Surname: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=0,row=2,pady=(5,5))
    ssname=customtkinter.CTkEntry(fmaddstaffs,border_width=1,border_color=bg1,font=fontentries,width=textboxwidth)
    ssname.grid(column=1,row=2,pady=(5,5))
    lb=customtkinter.CTkLabel(fmaddstaffs,text="ID No: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=2,row=2,pady=(5,5))
    sidno=customtkinter.CTkEntry(fmaddstaffs,border_width=1,border_color=bg1,font=fontentries,width=textboxwidth)
    sidno.grid(column=3,row=2,pady=(5,5))
    lb=customtkinter.CTkLabel(fmaddstaffs,text="Phone No: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=0,row=3,pady=(5,5))
    sphoneno=customtkinter.CTkEntry(fmaddstaffs,border_width=1,border_color=bg1,font=fontentries,width=textboxwidth)
    sphoneno.grid(column=1,row=3,pady=(5,5))
    lb=customtkinter.CTkLabel(fmaddstaffs,text="Email: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=2,row=3,pady=(5,5))
    semail=customtkinter.CTkEntry(fmaddstaffs,border_width=1,border_color=bg1,font=fontentries,width=textboxwidth)
    semail.grid(column=3,row=3,pady=(5,5))
    lb=customtkinter.CTkLabel(fmaddstaffs,text="Staff Type: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=0,row=4,pady=(5,5))
    sstafftype=customtkinter.CTkOptionMenu(fmaddstaffs,fg_color="white",button_color=bg1,button_hover_color=bg1,text_color="black",font=fontentries,width=textboxwidth,values=["Teaching","Non-Teaching"])
    sstafftype.grid(column=1,row=4,pady=(5,5))
    lb=customtkinter.CTkLabel(fmaddstaffs,text="Staff No: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=2,row=4,pady=(5,5))
    staffno=customtkinter.CTkEntry(fmaddstaffs,border_width=1,border_color=bg1,font=fontentries,width=textboxwidth)
    staffno.grid(column=3,row=4,pady=(5,5))
    lb=customtkinter.CTkLabel(fmaddstaffs,text="Occupation: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=0,row=5,pady=(5,5))
    soccupation=customtkinter.CTkOptionMenu(fmaddstaffs,fg_color="white",button_color=bg1,button_hover_color=bg1,text_color="black",font=fontentries,width=textboxwidth,values=["Principal","Deputy-Principle","Senior-Teacher","Teacher","ICT","Secretary","Librarian","Bursar","Lab-Technician","Driver","Cook","Security","Cleaner","Gardener"])
    soccupation.grid(column=1,row=5,pady=(5,5))
    lb=customtkinter.CTkLabel(fmaddstaffs,text="Employer: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=2,row=5,pady=(5,5))
    semployer=customtkinter.CTkOptionMenu(fmaddstaffs,fg_color="white",button_color=bg1,button_hover_color=bg1,text_color="black",font=fontentries,width=textboxwidth,values=["Government","B.O.M"])
    semployer.grid(column=3,row=5,pady=(5,5))
    lb=customtkinter.CTkLabel(fmaddstaffs,text="Religion: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=0,row=6,pady=(5,5))
    sreligion=customtkinter.CTkOptionMenu(fmaddstaffs,fg_color="white",button_color=bg1,button_hover_color=bg1,text_color="black",font=fontentries,width=textboxwidth,values=["Christian","Muslim","Others"])
    sreligion.grid(column=1,row=6,pady=(5,5))
    lb=customtkinter.CTkLabel(fmaddstaffs,text="County/Location: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=2,row=6,pady=(5,5))
    shomelocation=customtkinter.CTkEntry(fmaddstaffs,border_width=1,border_color=bg1,font=fontentries,width=textboxwidth)
    shomelocation.grid(column=3,row=6,pady=(5,5))
    fmstaffsbtnholder=customtkinter.CTkFrame(fmaddstaffs)
    fmstaffsbtnholder.grid(column=1,row=7,pady=(20,20),columnspan=3)
    btnsavestaffs=customtkinter.CTkButton(fmstaffsbtnholder,text="Save",fg_color=bg1,text_color=fg1,hover_color=hovbg1,cursor="hand2",command=addstaff)
    btnsavestaffs.grid(column=0,row=0,pady=(5,5),padx=5)
    btncancelstaffs=customtkinter.CTkButton(fmstaffsbtnholder,text="Cancel",fg_color=cancelcol,text_color=fg1,hover_color=hovbg1,cursor="hand2",command=clearstaffform)
    btncancelstaffs.grid(column=1,row=0,pady=(5,5),padx=5)
    #------------------------------------------
    #--------------edit student--------------------------
    framesearchstaff=customtkinter.CTkFrame(fmeditstaffs,fg_color=bg3,bg_color=coldark)
    framesearchstaff.pack(side=TOP,fill=X)
    customtkinter.CTkLabel(framesearchstaff,text="Search staff",text_color=bg1,font=fontlbl2,justify=CENTER,anchor=CENTER).grid(column=0,row=0,columnspan=5)
    customtkinter.CTkLabel(framesearchstaff,text="Search By: ",font=fontlbl,text_color=bg1).grid(column=0,row=1)
    searchstfby=customtkinter.CTkOptionMenu(framesearchstaff,fg_color="white",button_color=bg1,button_hover_color=bg1,text_color="black",font=fontentries,width=120,values=["StaffNo","IDNo"])
    searchstfby.grid(column=1,row=1)
    customtkinter.CTkLabel(framesearchstaff,text="Search Input: ",font=fontlbl,text_color=bg1).grid(column=2,row=1)
    ssearchtext=customtkinter.CTkEntry(framesearchstaff,border_width=1,border_color=bg1,font=fontentries,width=120)
    ssearchtext.grid(column=3,row=1)
    btnsearchstf=customtkinter.CTkButton(framesearchstaff,text="search",fg_color=bg1,hover_color=hovbg1,text_color=fg1,cursor="hand2",command=searchstaff)
    btnsearchstf.grid(column=4,row=1,padx=(4,3))
    
    framesearchedstaff=customtkinter.CTkFrame(fmeditstaffs,fg_color=bg3,bg_color=coldark)
    framesearchedstaff.pack(side=TOP,fill=BOTH,expand=True)
    lb=customtkinter.CTkLabel(framesearchedstaff,text="Edit Staff Form",text_color=bg1,font=fontlbl2,justify=CENTER,anchor=CENTER)
    lb.grid(column=0,row=0,pady=(5,5),columnspan=4)
    lb=customtkinter.CTkLabel(framesearchedstaff,text="First Name: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=0,row=1,pady=(5,5))
    esfname=customtkinter.CTkEntry(framesearchedstaff,border_width=1,border_color=bg1,font=fontentries,width=textboxwidth)
    esfname.grid(column=1,row=1,pady=(5,5))
    lb=customtkinter.CTkLabel(framesearchedstaff,text="Last Name: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=2,row=1,pady=(5,5))
    eslname=customtkinter.CTkEntry(framesearchedstaff,border_width=1,border_color=bg1,font=fontentries,width=textboxwidth)
    eslname.grid(column=3,row=1,pady=(5,5))
    lb=customtkinter.CTkLabel(framesearchedstaff,text="Surname: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=0,row=2,pady=(5,5))
    essname=customtkinter.CTkEntry(framesearchedstaff,border_width=1,border_color=bg1,font=fontentries,width=textboxwidth)
    essname.grid(column=1,row=2,pady=(5,5))
    lb=customtkinter.CTkLabel(framesearchedstaff,text="ID No: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=2,row=2,pady=(5,5))
    esidno=customtkinter.CTkEntry(framesearchedstaff,border_width=1,border_color=bg1,font=fontentries,width=textboxwidth)
    esidno.grid(column=3,row=2,pady=(5,5))
    lb=customtkinter.CTkLabel(framesearchedstaff,text="Phone No: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=0,row=3,pady=(5,5))
    esphoneno=customtkinter.CTkEntry(framesearchedstaff,border_width=1,border_color=bg1,font=fontentries,width=textboxwidth)
    esphoneno.grid(column=1,row=3,pady=(5,5))
    lb=customtkinter.CTkLabel(framesearchedstaff,text="Email: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=2,row=3,pady=(5,5))
    esemail=customtkinter.CTkEntry(framesearchedstaff,border_width=1,border_color=bg1,font=fontentries,width=textboxwidth)
    esemail.grid(column=3,row=3,pady=(5,5))
    lb=customtkinter.CTkLabel(framesearchedstaff,text="Staff Type: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=0,row=4,pady=(5,5))
    esstafftype=customtkinter.CTkOptionMenu(framesearchedstaff,fg_color="white",button_color=bg1,button_hover_color=bg1,text_color="black",font=fontentries,width=textboxwidth,values=["Teaching","Non-Teaching"])
    esstafftype.grid(column=1,row=4,pady=(5,5))
    lb=customtkinter.CTkLabel(framesearchedstaff,text="Staff No: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=2,row=4,pady=(5,5))
    estaffno=customtkinter.CTkEntry(framesearchedstaff,border_width=1,border_color=bg1,font=fontentries,width=textboxwidth)
    estaffno.grid(column=3,row=4,pady=(5,5))
    lb=customtkinter.CTkLabel(framesearchedstaff,text="Occupation: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=0,row=5,pady=(5,5))
    esoccupation=customtkinter.CTkOptionMenu(framesearchedstaff,fg_color="white",button_color=bg1,button_hover_color=bg1,text_color="black",font=fontentries,width=textboxwidth,values=["Principal","Deputy-Principle","Senior-Teacher","Teacher","ICT","Secretary","Librarian","Bursar","Lab-Technician","Driver","Cook","Security","Cleaner","Gardener"])
    esoccupation.grid(column=1,row=5,pady=(5,5))
    lb=customtkinter.CTkLabel(framesearchedstaff,text="Employer: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=2,row=5,pady=(5,5))
    esemployer=customtkinter.CTkOptionMenu(framesearchedstaff,fg_color="white",button_color=bg1,button_hover_color=bg1,text_color="black",font=fontentries,width=textboxwidth,values=["Government","B.O.M"])
    esemployer.grid(column=3,row=5,pady=(5,5))
    lb=customtkinter.CTkLabel(framesearchedstaff,text="Religion: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=0,row=6,pady=(5,5))
    esreligion=customtkinter.CTkOptionMenu(framesearchedstaff,fg_color="white",button_color=bg1,button_hover_color=bg1,text_color="black",font=fontentries,width=textboxwidth,values=["Christian","Muslim","Others"])
    esreligion.grid(column=1,row=6,pady=(5,5))
    lb=customtkinter.CTkLabel(framesearchedstaff,text="County/Location: ",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=E)
    lb.grid(column=2,row=6,pady=(5,5))
    eshomelocation=customtkinter.CTkEntry(framesearchedstaff,border_width=1,border_color=bg1,font=fontentries,width=textboxwidth)
    eshomelocation.grid(column=3,row=6,pady=(5,5))
    efmstaffsbtnholder=customtkinter.CTkFrame(framesearchedstaff)
    efmstaffsbtnholder.grid(column=1,row=7,pady=(20,20),columnspan=3)
    ebtnupdatestaffs=customtkinter.CTkButton(efmstaffsbtnholder,text="Update",fg_color=bg1,text_color=fg1,hover_color=hovbg1,cursor="hand2",command=updatestaff)
    ebtnupdatestaffs.grid(column=0,row=0,pady=(5,5),padx=5)
    ebtndeletestaffs=customtkinter.CTkButton(efmstaffsbtnholder,text="Delete",fg_color=cancelcol,text_color=fg1,hover_color=hovbg1,cursor="hand2",command=deletestaff)
    ebtndeletestaffs.grid(column=1,row=0,pady=(5,5),padx=5)
    ebtncancelstaffs=customtkinter.CTkButton(efmstaffsbtnholder,text="Cancel",fg_color=coldark,text_color=fg1,hover_color=hovbg1,cursor="hand2",command=clearstaffeditform)
    ebtncancelstaffs.grid(column=2,row=0,pady=(5,5),padx=5)
    #-----------------------------------------------------
    #-----------------view students----------------------
    tablestaffs=ttk.Treeview(fmviewstaffs)
    tablestaffs.pack(side=TOP,fill=BOTH,expand=True)
    tablestaffs['columns']=(0,1,2,3,4,5,6,7)
    tablestaffs['show']="headings"
    tablestaffs.heading(0,text="Staffno")
    tablestaffs.heading(1,text="Full Name")
    tablestaffs.heading(2,text="Idno")
    tablestaffs.heading(3,text="Phone no")
    tablestaffs.heading(4,text="Email")
    tablestaffs.heading(5,text="staff Type")
    tablestaffs.heading(6,text="Occupation")
    tablestaffs.heading(7,text="Employer")
    tablestaffs.column(0,width=60)
    tablestaffs.column(1,width=130)
    tablestaffs.column(2,width=60)
    tablestaffs.column(3,width=80)
    tablestaffs.column(4,width=110)
    tablestaffs.column(5,width=80)
    tablestaffs.column(6,width=80)
    tablestaffs.column(7,width=80)
    viewstaffs()
    #------------------------------------------------------
    #==========================================
    #=========framereport====================
    def toaddterm():
        ntbook4.select(0)
    def toreporting():
        ntbook4.select(1)
    def toreported():
        ntbook4.select(2)
    def createterm():
        if yearcreated.get()=="":
            return False
        elif term.get()=="":
            return False
        elif startdate.get()=="":
            return False
        elif enddate.get()=="":
            return False
        else:
            toconfirmmsg=messagebox.askyesno("Confirm message","Do you want to create term?\nif you create it will be the current term.\n please confirm!")
            if toconfirmmsg==True:
                global myusername
                termname="T"+str(term.get())+"Y"+str(yearcreated.get())
                termstatus=1
                try:
                    sql1="UPDATE terms SET termstatus=0"
                    cursor=conn.cursor()
                    cursor.execute(sql1)
                    sql="INSERT INTO terms(yearcreated,term,termname,startdate,enddate,termstatus,recordby) VALUES('%s','%s','%s','%s','%s',%s,'%s')"
                    cursor=conn.cursor()
                    cursor.execute(sql%(yearcreated.get(),term.get(),termname,startdate.get(),enddate.get(),1,myusername))
                    conn.commit()
                    viewterms()
                    viewcurrentterm()
                    cleartermform()
                    get_totol_reported()
                    get_school_finance_analysis()
                    get_reporting_analysis()
                    get_resource_analysis()
                except IOError:
                    pass    
    def cleartermform():
        yearcreated.set("")
        term.set("")
        startdate.delete(0,END)
        enddate.delete(0,END)
    def cleareditterm():
        ttermname.delete(0,END)
        tstartdate.delete(0,END)
        tenddate.delete(0,END)
    def viewterms():
        get_totol_reported()
        try:
            sql="SELECT termname,yearcreated,startdate,enddate,termstatus,recordby FROM terms"
            cursor=conn.cursor()
            cursor.execute(sql)
            results=cursor.fetchall()
            if results:
                for records in tableterms.get_children():
                    tableterms.delete(records)
                for i in results:
                    tableterms.insert('',END,values=i)
        except IOError:
            pass
    def viewcurrentterm():
        global currenttermid,currenttermname
        try:
            
            sql="SELECT termname,startdate,enddate,termid FROM terms WHERE termstatus=1 LIMIT 1"
            cursor=conn.cursor()
            cursor.execute(sql)
            results=cursor.fetchall()
            if results:
                cleareditterm()
                ttermname.insert(END,results[0][0])
                tstartdate.insert(END,results[0][1])
                tenddate.insert(END,results[0][2])
                currenttermid=results[0][3]
                currenttermname=results[0][0]
        except IOError:
            pass
    def get_totol_reported():
        global currenttermid
        try:
            sql="SELECT COUNT(stdadmno) FROM reporting WHERE reporttermid='%s'"
            cursor=conn.cursor()
            cursor.execute(sql%(currenttermid))
            results=cursor.fetchall()
            if results:
                lbltotalreporting.configure(text=str(results[0][0])+"\nReported")
        except IOError:
            pass
    framesubmenu4=customtkinter.CTkFrame(framereporting,width=150,fg_color=coldark)
    framesubmenu4.pack(side=LEFT,fill=Y)
    btnaddterm=customtkinter.CTkButton(framesubmenu4,text="Add Term",text_color=fg1,fg_color=bg1,hover_color=hovbg1,cursor="hand2",command=toaddterm)
    btnaddterm.pack(pady=6,padx=6)
    btnsreporting=customtkinter.CTkButton(framesubmenu4,text="Reporting",text_color=fg1,fg_color=bg1,hover_color=hovbg1,cursor="hand2",command=toreporting)
    btnsreporting.pack(pady=6,padx=6)
    btnviewsreported=customtkinter.CTkButton(framesubmenu4,text="View Reported",text_color=fg1,fg_color=bg1,hover_color=hovbg1,cursor="hand2",command=toreported)
    btnviewsreported.pack(pady=6,padx=6)
    ntbook4=ttk.Notebook(framereporting)
    ntbook4.pack(side=LEFT,fill=BOTH,expand=True)
    fmaddterm=customtkinter.CTkFrame(ntbook4,fg_color=bg3,bg_color=coldark)
    fmaddterm.place(x=0,y=0,relheight=1,relwidth=1)
    fmsreporting=customtkinter.CTkFrame(ntbook4,fg_color=bg3,bg_color=coldark)
    fmsreporting.place(x=0,y=0,relheight=1,relwidth=1)
    fmviewsreported=customtkinter.CTkFrame(ntbook4,fg_color=bg3,bg_color=coldark)
    fmviewsreported.place(x=0,y=0,relheight=1,relwidth=1)
    ntbook4.add(fmaddterm,text="Add Term")
    ntbook4.add(fmsreporting,text="Reporting")
    ntbook4.add(fmviewsreported,text="View Reported")
    #-----------------------add term-----------------------
    fmcreateterm=customtkinter.CTkFrame(fmaddterm,fg_color=bg3)
    fmcreateterm.pack(fill=X)
    lb=customtkinter.CTkLabel(fmcreateterm,text="Create term",font=fontlbl2,text_color=bg1,justify=CENTER,anchor=CENTER)
    lb.grid(column=0,row=0,columnspan=4)
    lb=customtkinter.CTkLabel(fmcreateterm,text="Year: ",font=fontlbl,text_color=bg1,justify=RIGHT)
    lb.grid(column=0,row=1)
    yearcreated=customtkinter.CTkOptionMenu(fmcreateterm,fg_color="white",text_color="black",button_color=bg1,button_hover_color=bg1,font=fontentries,width=130,values=["2022","2023","2024","2025"])
    yearcreated.grid(column=1,row=1,pady=(3,3))
    lb=customtkinter.CTkLabel(fmcreateterm,text="Term: ",font=fontlbl,text_color=bg1,justify=RIGHT)
    lb.grid(column=2,row=1)
    term=customtkinter.CTkOptionMenu(fmcreateterm,fg_color="white",text_color="black",button_color=bg1,button_hover_color=bg1,font=fontentries,width=130,values=["1","2","3"])
    term.grid(column=3,row=1,pady=(3,3))
    lb=customtkinter.CTkLabel(fmcreateterm,text="Start Date: ",font=fontlbl,text_color=bg1,justify=RIGHT)
    lb.grid(column=0,row=2)
    startdate=customtkinter.CTkEntry(fmcreateterm,border_width=1,border_color=bg1,font=fontentries,width=130)
    startdate.grid(column=1,row=2,pady=(3,3))
    lb=customtkinter.CTkLabel(fmcreateterm,text="End Date: ",font=fontlbl,text_color=bg1,justify=RIGHT)
    lb.grid(column=2,row=2)
    enddate=customtkinter.CTkEntry(fmcreateterm,border_width=1,border_color=bg1,font=fontentries,width=130)
    enddate.grid(column=3,row=2,pady=(3,3))
    btnsaveterm=customtkinter.CTkButton(fmcreateterm,text="Save",fg_color=bg1,hover_color=hovbg1,text_color=fg1,cursor="hand2",command=createterm)
    btnsaveterm.grid(column=4,row=1,padx=(3,3))
    btncancelterm=customtkinter.CTkButton(fmcreateterm,text="Cancel",fg_color=coldark,hover_color=hovbg1,text_color=fg1,cursor="hand2",command=cleartermform)
    btncancelterm.grid(column=4,row=2,padx=(3,3))
    
    fmcurrentterm=customtkinter.CTkFrame(fmaddterm,fg_color=bg3)
    fmcurrentterm.pack(fill=X)
    lb=customtkinter.CTkLabel(fmcurrentterm,text="Current term",font=fontlbl2,text_color=bg1,justify=CENTER,anchor=CENTER)
    lb.grid(column=0,row=0,columnspan=4)
    lb=customtkinter.CTkLabel(fmcurrentterm,text="Term Name: ",font=fontlbl,text_color=bg1,justify=RIGHT)
    lb.grid(column=0,row=1)
    ttermname=customtkinter.CTkEntry(fmcurrentterm,border_width=1,border_color=bg1,font=fontentries,width=130)
    ttermname.grid(column=1,row=1,pady=(3,3))
    lb=customtkinter.CTkLabel(fmcurrentterm,text="Start Date: ",font=fontlbl,text_color=bg1,justify=RIGHT)
    lb.grid(column=0,row=2)
    tstartdate=customtkinter.CTkEntry(fmcurrentterm,border_width=1,border_color=bg1,font=fontentries,width=130)
    tstartdate.grid(column=1,row=2,pady=(3,3))
    lb=customtkinter.CTkLabel(fmcurrentterm,text="End Date: ",font=fontlbl,text_color=bg1,justify=RIGHT)
    lb.grid(column=2,row=2)
    tenddate=customtkinter.CTkEntry(fmcurrentterm,border_width=1,border_color=bg1,font=fontentries,width=130)
    tenddate.grid(column=3,row=2,pady=(3,3))
    btnupdateterm=customtkinter.CTkButton(fmcurrentterm,text="Update",fg_color=bg1,hover_color=hovbg1,text_color=fg1,cursor="hand2")
    btnupdateterm.grid(column=1,row=3,pady=(3,3),padx=(3,3))
    btncancelterm=customtkinter.CTkButton(fmcurrentterm,text="Cancel",fg_color=coldark,hover_color=hovbg1,text_color=fg1,cursor="hand2")
    btncancelterm.grid(column=2,row=3,pady=(3,3),padx=(3,3))
    lbltotalreporting=customtkinter.CTkLabel(fmcurrentterm,text="0\nreported",text_color=bg1,font=('times',20,'bold'))
    lbltotalreporting.grid(column=4,row=1,rowspan=2,padx=(10,5))
    
    fmallterms=customtkinter.CTkFrame(fmaddterm,fg_color=bg3)
    fmallterms.pack(fill=X)
    tableterms=ttk.Treeview(fmallterms)
    tableterms.pack(side=LEFT,fill=BOTH,expand=True)
    tableterms['show']="headings"
    tableterms['columns']=(0,1,2,3,4,5,)
    tableterms.heading(0,text="Term Name")
    tableterms.heading(1,text="Year")
    tableterms.heading(2,text="Start Date")
    tableterms.heading(3,text="End Date")
    tableterms.heading(4,text="status")
    tableterms.heading(5,text="Created By")
    tableterms.column(0,width=100,anchor=CENTER)
    tableterms.column(1,width=100,anchor=CENTER)
    tableterms.column(2,width=100,anchor=CENTER)
    tableterms.column(3,width=100,anchor=CENTER)
    tableterms.column(4,width=100,anchor=CENTER)
    tableterms.column(5,width=100,anchor=CENTER)
    viewterms()
    viewcurrentterm()
    get_totol_reported()
    #-------------------------------------------------------
    #--------------reporting------------------------
    def searchstudenttotreport():
        global currenttermid,currenttermname
        try:
            sql="SELECT studentid,CONCAT(fname,' ',lname,' ',sname) AS name,admno,form FROM students WHERE admno='%s'"%(rsrchadmno.get())
            cursor=conn.cursor()
            cursor.execute(sql)
            results=cursor.fetchall()
            if results:
                clearstudentreportingform()
                rsadmno.insert(END,results[0][2])
                rsname.insert(END,results[0][1])
                rtermid.insert(END,currenttermid)
                rtermname.insert(END,currenttermname)
                classgrade.set(results[0][3])
        except IOError:
            pass
    def studentreporting():
        global currenttermid
        if rsadmno.get()=="":
            return False
        elif rsname.get()=="":
            return False
        elif rtermid.get()=="":
            return False
        elif rtermname.get()=="":
            return False
        elif classgrade.get()=="":
            return False
        else:
            #check if the student has already reported
            try:
                sqlcheck="SELECT * FROM reporting WHERE stdadmno='%s' AND reporttermid=%s"%(rsrchadmno.get(),currenttermid)
                cursor=conn.cursor()
                cursor.execute(sqlcheck)
                results=cursor.fetchall()
                if results:
                    messagebox.showwarning("Alert","This student has already reported")
                    clearstudentreportingform()
                else:
                    try:
                        sql="INSERT INTO reporting(reporttermid,stdadmno,reportdate,class) VALUES(%s,'%s','%s','%s')"
                        cursor=conn.cursor()
                        cursor.execute(sql%(currenttermid,rsadmno.get(),today,classgrade.get()))
                        conn.commit()
                        clearstudentreportingform()
                        viewreported()
                        get_totol_reported()
                        get_school_finance_analysis()
                        get_reporting_analysis()
                        get_resource_analysis()
                    except IOError:
                        pass
            except IOError:
                pass
    def clearstudentreportingform():
        rsadmno.delete(0,END)
        rsname.delete(0,END)
        rtermid.delete(0,END)
        rtermname.delete(0,END)
        classgrade.set("PP_1")
    def viewreported():
        try:
            sql="SELECT * FROM reporting"
            cursor=conn.cursor()
            cursor.execute(sql)
            results=cursor.fetchall()
            conn.commit()
            if results:
                for records in tablereported.get_children():
                    tablereported.delete(records)
                for i in results:
                    tablereported.insert('',END,values=i)
        except IOError:
            pass
    fmstdunetreportsearch=customtkinter.CTkFrame(fmsreporting,fg_color=bg3)
    fmstdunetreportsearch.pack(fill=X)
    lb=customtkinter.CTkLabel(fmstdunetreportsearch,text="Search Student",font=fontlbl2,text_color=bg1)
    lb.grid(column=0,row=0,columnspan=3)
    lb=customtkinter.CTkLabel(fmstdunetreportsearch,text="Enter Admno:",text_color=bg1,font=fontlbl)
    lb.grid(column=0,row=1,pady=(4,4))
    rsrchadmno=customtkinter.CTkEntry(fmstdunetreportsearch,border_color=bg1,border_width=1,width=textboxwidth)
    rsrchadmno.grid(column=1,row=1,pady=(4,4))
    btnsearch=customtkinter.CTkButton(fmstdunetreportsearch,text="Search",fg_color=bg1,text_color=fg1,hover_color=hovbg1,cursor="hand2",command=searchstudenttotreport)
    btnsearch.grid(column=2,row=1,padx=(3,3),pady=(4,4))
    fmstudentreport=customtkinter.CTkFrame(fmsreporting,fg_color=bg3)
    fmstudentreport.pack(fill=X)
    lb=customtkinter.CTkLabel(fmstudentreport,text="Search Student",font=fontlbl2,text_color=bg1)
    lb.grid(column=0,row=0,columnspan=4)
    lb=customtkinter.CTkLabel(fmstudentreport,text="Admno: ",font=fontlbl,text_color=bg1,justify=RIGHT)
    lb.grid(column=0,row=1)
    rsadmno=customtkinter.CTkEntry(fmstudentreport,border_color=bg1,border_width=1,width=textboxwidth)
    rsadmno.grid(column=1,row=1,pady=(4,4))
    lb=customtkinter.CTkLabel(fmstudentreport,text="Name: ",font=fontlbl,text_color=bg1,justify=RIGHT)
    lb.grid(column=2,row=1)
    rsname=customtkinter.CTkEntry(fmstudentreport,border_color=bg1,border_width=1,width=textboxwidth)
    rsname.grid(column=3,row=1,pady=(4,4))
    
    lb=customtkinter.CTkLabel(fmstudentreport,text="Term Id: ",font=fontlbl,text_color=bg1,justify=RIGHT)
    lb.grid(column=0,row=2)
    rtermid=customtkinter.CTkEntry(fmstudentreport,border_color=bg1,border_width=1,width=textboxwidth)
    rtermid.grid(column=1,row=2,pady=(4,4))
    lb=customtkinter.CTkLabel(fmstudentreport,text="Term: ",font=fontlbl,text_color=bg1,justify=RIGHT)
    lb.grid(column=2,row=2)
    rtermname=customtkinter.CTkEntry(fmstudentreport,border_color=bg1,border_width=2,width=textboxwidth)
    rtermname.grid(column=3,row=2,pady=(4,4))
    lb=customtkinter.CTkLabel(fmstudentreport,text="Grade: ",font=fontlbl,text_color=bg1,justify=RIGHT)
    lb.grid(column=0,row=4)
    classgrade=customtkinter.CTkOptionMenu(fmstudentreport,fg_color="white",text_color="black",width=textboxwidth,button_color=bg1,button_hover_color=bg1,values=class_values)
    classgrade.grid(column=1,row=4,pady=(4,4))
    fmreportbtn1=customtkinter.CTkFrame(fmstudentreport)
    fmreportbtn1.grid(column=0,row=5,columnspan=4,pady=(5,5))
    btnstudentreport=customtkinter.CTkButton(fmreportbtn1,text="Report",text_color=fg1,fg_color=bg1,hover_color=hovbg1,cursor="hand2",command=studentreporting)
    btnstudentreport.grid(column=0,row=0,padx=(3,3),pady=(3,3))
    btnstudentreportcancel=customtkinter.CTkButton(fmreportbtn1,text="Clear",text_color=fg1,fg_color=coldark,hover_color=hovbg1,cursor="hand2",command=clearstudentreportingform)
    btnstudentreportcancel.grid(column=1,row=0,padx=(3,3),pady=(3,3))
    #-----------------------------------------------
    #---------------------view reported-----------
    fmreportfilter=customtkinter.CTkFrame(fmviewsreported,fg_color=bg3)
    fmreportfilter.pack(fill=X,side=TOP)
    lb=customtkinter.CTkLabel(fmreportfilter,text="Filter: ",font=fontlbl,text_color=bg1,justify=RIGHT)
    lb.grid(column=0,row=0)
    lb=customtkinter.CTkLabel(fmreportfilter,text="From Date: ",font=fontlbl,text_color=bg1,justify=RIGHT)
    lb.grid(column=0,row=1)
    filterfromdate=DateEntry(fmreportfilter,date_pattern="yyyy/mm/dd")
    filterfromdate.grid(column=1,row=1)
    lb=customtkinter.CTkLabel(fmreportfilter,text="To Date: ",font=fontlbl,text_color=bg1,justify=RIGHT)
    lb.grid(column=2,row=1)
    filtertodate=DateEntry(fmreportfilter,date_pattern="yyyy/mm/dd")
    filtertodate.grid(column=3,row=1)
    lb=customtkinter.CTkLabel(fmreportfilter,text="Admno: ",font=fontlbl,text_color=bg1,justify=RIGHT)
    lb.grid(column=4,row=1)
    filteradmno=Entry(fmreportfilter)
    filteradmno.grid(column=5,row=1)
    tablereported=ttk.Treeview(fmviewsreported)
    tablereported.pack(side=LEFT,fill=BOTH,expand=True)
    tablereported['show']="headings"
    tablereported['columns']=(0,1,2,3,4)
    tablereported.heading(0,text="Report Id")
    tablereported.heading(1,text="Term Id")
    tablereported.heading(2,text="Student Admno")
    tablereported.heading(3,text="Report date")
    tablereported.heading(4,text="Grade")
    tablereported.column(0,width=60,anchor=CENTER)
    tablereported.column(1,width=80,anchor=CENTER)
    tablereported.column(2,width=130,anchor=CENTER)
    tablereported.column(3,width=60,anchor=CENTER)
    tablereported.column(4,width=100,anchor=CENTER)
    viewreported()
    #--------------------------------------------
    #==========================================
    #=========framefinance====================
    def toaddfees():
        ntbook5.select(0)
    def toviewpaid():
        ntbook5.select(1)
    def tosalaries():
        ntbook5.select(2)
    def toprojects():
        ntbook5.select(3)
    def topettycash():
        ntbook5.select(4)
    framesubmenu5=customtkinter.CTkFrame(framefinance,width=150,fg_color=coldark)
    framesubmenu5.pack(side=LEFT,fill=Y)
    btnaddfees=customtkinter.CTkButton(framesubmenu5,text="Add Fees",text_color=fg1,fg_color=bg1,hover_color=hovbg1,cursor="hand2",command=toaddfees)
    btnaddfees.pack(pady=6,padx=6)
    btnviewpaid=customtkinter.CTkButton(framesubmenu5,text="View Paid",text_color=fg1,fg_color=bg1,hover_color=hovbg1,cursor="hand2",command=toviewpaid)
    btnviewpaid.pack(pady=6,padx=6)
    btnsalaries=customtkinter.CTkButton(framesubmenu5,text="Salaries",text_color=fg1,fg_color=bg1,hover_color=hovbg1,cursor="hand2",command=tosalaries)
    btnsalaries.pack(pady=6,padx=6)
    btnprojects=customtkinter.CTkButton(framesubmenu5,text="Projects",text_color=fg1,fg_color=bg1,hover_color=hovbg1,cursor="hand2",command=toprojects)
    btnprojects.pack(pady=6,padx=6)
    btnpettycash=customtkinter.CTkButton(framesubmenu5,text="Petty Cash",text_color=fg1,fg_color=bg1,hover_color=hovbg1,cursor="hand2",command=topettycash)
    btnpettycash.pack(pady=6,padx=6)
    ntbook5=ttk.Notebook(framefinance)
    ntbook5.pack(side=LEFT,fill=BOTH,expand=True)
    fmaddfees=customtkinter.CTkFrame(ntbook5,fg_color=bg3,bg_color=coldark)
    fmaddfees.place(x=0,y=0,relheight=1,relwidth=1)
    fmviewpaid=customtkinter.CTkFrame(ntbook5,fg_color=bg3,bg_color=coldark)
    fmviewpaid.place(x=0,y=0,relheight=1,relwidth=1)
    fmsalaries=customtkinter.CTkFrame(ntbook5,fg_color=bg3,bg_color=coldark)
    fmsalaries.place(x=0,y=0,relheight=1,relwidth=1)
    fmprojects=customtkinter.CTkFrame(ntbook5,fg_color=bg3,bg_color=coldark)
    fmprojects.place(x=0,y=0,relheight=1,relwidth=1)
    fmpettycash=customtkinter.CTkFrame(ntbook5,fg_color=bg3,bg_color=coldark)
    fmpettycash.place(x=0,y=0,relheight=1,relwidth=1)
    ntbook5.add(fmaddfees,text="Add Fees")
    ntbook5.add(fmviewpaid,text="View Paid")
    ntbook5.add(fmsalaries,text="Salaries")
    ntbook5.add(fmprojects,text="Projects")
    ntbook5.add(fmpettycash,text="Petty Cash")
    #-------------to add fees------------------
    def tosettermfee():
        global currenttermid,myusername
        if mycurrentterm.get()=="":
            return False
        elif mycurrenttermfee.get()=="":
            return False
        else:
            msg1=messagebox.askyesno("Confirm","Do you want to set term fee for current term?")
            if msg1==True:
                #check if termfee is set
                sql1="SELECT * FROM termfee WHERE termid=%s"%(currenttermid)
                cursor=conn.cursor()
                cursor.execute(sql1)
                available=cursor.fetchall()
                if available:
                    messagebox.showwarning("Error","This Term Fee is already Set")
                else:
                    try:
                        sql="INSERT INTO termfee(termid,amount,createddate,recordby) VALUES(%s,%s,'%s','%s')"
                        cursor=conn.cursor()
                        cursor.execute(sql%(currenttermid,mycurrenttermfee.get(),today,myusername))
                        conn.commit()
                        messagebox.showinfo("Success","Term Fee set successful")
                        gettermfee()
                        get_school_finance_analysis()
                        get_reporting_analysis()
                        get_resource_analysis()
                    except IOError:
                        messagebox.showwarning("Error","Try Again")
    def gettermfee():
        global currenttermid,mytermfee
        try:
            sql="SELECT amount FROM termfee WHERE termid=%s"%(currenttermid)
            cursor=conn.cursor()
            cursor.execute(sql)
            results=cursor.fetchall()
            if results:
                mytermfee=results[0][0]
                mycurrenttermfee.delete(0,END)
                mycurrenttermfee.insert(END,results[0][0])
        except:
            pass
    framesettermfee=customtkinter.CTkFrame(fmaddfees,fg_color=bg3)
    framesettermfee.pack(side=TOP,pady=(2,4))
    lb=customtkinter.CTkLabel(framesettermfee,text="Set Current Term fees",text_color=bg1,font=fontlbl2)
    lb.grid(column=0,row=0,columnspan=5)
    lb=customtkinter.CTkLabel(framesettermfee,text="Current ID:",text_color=bg1,font=fontlbl)
    lb.grid(column=0,row=1,pady=(2,2))
    mycurrentterm=customtkinter.CTkEntry(framesettermfee,border_width=1,border_color=bg1,width=entry_width)
    mycurrentterm.grid(column=1,row=1,padx=(2,2),pady=(2,2))
    mycurrentterm.insert(END,currenttermid)
    lb=customtkinter.CTkLabel(framesettermfee,text="Current Fees:",text_color=bg1,font=fontlbl)
    lb.grid(column=2,row=1,pady=(2,2))
    mycurrenttermfee=customtkinter.CTkEntry(framesettermfee,border_width=1,border_color=bg1,width=entry_width)
    mycurrenttermfee.grid(column=3,row=1,padx=(2,2),pady=(2,2))
    btnsettermfee=customtkinter.CTkButton(framesettermfee,text="Save",text_color=fg1,fg_color=bg1,cursor="hand2",hover_color=hovbg1,command=tosettermfee)
    btnsettermfee.grid(column=4,row=1,padx=(4,2),pady=(2,2))
    gettermfee()
    #-----student fees-----------
    def getstudentname(*arg):
        global student_class
        try:
            sql="SELECT admno,CONCAT(fname,' ',lname,' ',sname),form AS name FROM students WHERE admno='%s'"%(feestadmno.get())
            cursor=conn.cursor()
            cursor.execute(sql)
            results=cursor.fetchall()
            if results:
                clearallpaymentdetails()
                feestname.insert(END,results[0][1])
                feestadmno.insert(END,results[0][0])
                student_class=results[0][2]
            else:
                messagebox.showerror("Invalid Admno")
        except IOError:
            pass
    def clearallpaymentdetails():
        feestadmno.delete(0,END)
        feestname.delete(0,END)
        feestamount.delete(0,END)
        feeremarks.delete(0,END)
        paidvia.set("Cash")
    def addfees():
        if feestadmno.get()=="":
            return False
        elif feestname.get()=="":
            return False
        elif feestamount.get()=="":
            return False
        elif feeremarks.get()=="":
            return False
        else:
            try:
                sql="INSERT INTO fees(stadmno,amount,paidvia,remarks,paymentdate,recordby) VALUES('%s',%s,'%s','%s','%s','%s')"
                cursor=conn.cursor()
                cursor.execute(sql%(feestadmno.get(),feestamount.get(),paidvia.get(),feeremarks.get(),today,myusername))
                conn.commit()
                viewpaid()
                messagebox.showinfo("Success","Saved successful")
                #clearallpaymentdetails()
                checkstatement()
                get_school_finance_analysis()
                get_reporting_analysis()
                get_resource_analysis()
            except IOError:
                pass
    def checkstatement():
        global mytermfee,total_reported_termfee,total_balance
        try:
            sql="SELECT stadmno,amount,paidvia,paymentdate FROM fees WHERE stadmno='%s'"%(feestadmno.get())
            cursor=conn.cursor()
            cursor.execute(sql)
            results=cursor.fetchall()
            
            if results:
                total_reported_termfee=0
                sql2="SELECT reporttermid FROM reporting WHERE stdadmno='%s'"%(feestadmno.get())
                cursor.execute(sql2)
                t_res=cursor.fetchall()
                if t_res:
                    i=0
                    while i<=len(t_res):
                        try:
                            sql3="SELECT amount FROM termfee WHERE termid='%s'"
                            #cursor.execute(sql3)
                            cursor.execute(sql3%(t_res[i][0]))
                            my_res=cursor.fetchall()
                            if my_res:
                                total_reported_termfee=total_reported_termfee+my_res[0][0]
                        except IOError:
                            pass
                        #print(t_res[i][0])
                        i=i+1
                        if i==len(t_res):
                            break
                    
                else:
                    messagebox.showwarning("Warning","Student not reported")
                for records in tablestatement.get_children():
                    tablestatement.delete(records)
                for i in results:
                    tablestatement.insert('',END,values=i)
                amount=0
                j=0
                x=len(results)
                while j<=x:
                    amount=amount+results[j][1]
                    j=j+1
                    if j==int(x):
                        break

                total_balance=float(total_reported_termfee)-float(amount)
                lbl_total_fee_amount.configure(text="Total amount: "+str(amount))
                lbl_total_expected_amount.configure(text="Expected: "+str(total_reported_termfee))
                lbl_total_balance_amount.configure(text="Balance: "+str(total_balance))
            else:
                for records in tablestatement.get_children():
                    tablestatement.delete(records)
                lbl_total_fee_amount.configure(text="Total amount: ")
                lbl_total_expected_amount.configure(text="Expected: ")
                lbl_total_balance_amount.configure(text="Balance: ")
                messagebox.showinfo("No details","No student fee details\n Pay 0 to the student")
        except:
            pass
    def print_statement():
        #get student data
        global school_logo,list_school_details,total_balance,total_reported_termfee,student_class

        sc_n=list_school_details[0][1]
        sc_c=list_school_details[0][3]
        sc_l="icons/"+list_school_details[0][7]
        if feestadmno.get()=="":
            messagebox.showinfo("Error","Enter Student admno")
        else:
            st_admno=feestadmno.get()
            st_name=feestname.get()
            #get fee statement
            try:
                sql="SELECT stadmno,amount,paidvia,paymentdate FROM fees WHERE stadmno='%s'"%(feestadmno.get())
                cursor=conn.cursor()
                cursor.execute(sql)
                results=cursor.fetchall()
                
                if results:
                    #print(sc_c+sc_l+sc_n)
                    try:
                        class PDF(FPDF):
                            try:
                                cursor=conn.cursor()
                                cursor.execute("SELECT * FROM schoolinfo")
                                results=cursor.fetchall()
                                if results:
                                    sc_l="icons/"+results[0][7]
                                    sc_n=results[0][1]
                                    sc_c=results[0][3]
                                else:
                                    pass
                            except IOError:
                                pass
                            def header(self):
                                self.image(f"{self.sc_l}",10,4,16)
                                self.set_font('helvetica','B',16)
                                self.cell(0,6,f'{self.sc_n}',border=False,ln=1,align='C')
                                self.set_font('helvetica','B',13)
                                self.cell(0,6,'Fee statement',border=False,ln=1,align='C')
                                self.ln(6)
                            def footer(self):
                                self.set_y(-15)
                                self.set_font('helvetica','I',8)
                                self.cell(0,6,f"School Contact: {self.sc_c}",align='C')
                        pdf=PDF('P','mm','A5')
                        pdf.add_page()
                        
                        pdf.set_author('Eric Software Solutions')
                        pdf.set_title(st_admno+" statement")
                        pdf.set_auto_page_break(auto=1,margin=10)
                        pdf.set_font('helvetica','B',10)
                        pdf.cell(40,6,f'Name: {st_name}',border=False,align='C')
                        pdf.cell(40,6,f'Adm No: {st_admno}',border=False,align='C')
                        pdf.cell(40,6,f"Class: {student_class}",border=False,ln=1,align='C')
                        pdf.cell(15,6,'No',border=1,align='C')
                        pdf.cell(32,6,'Paid Via',border=1,align='C')
                        pdf.cell(33,6,'Date',border=1,align='C')
                        pdf.cell(35,6,'Paid Amount(KES)',border=1,align='C',ln=1)
                        pdf.set_font('helvetica','I',9)
                        amount=0
                        j=0
                        x=len(results)
                        while j<=x:
                            no=j+1
                            pdf.cell(15,6,str(no),border=1,align='C')
                            pdf.cell(32,6,str(results[j][2]),border=1,align='C')
                            pdf.cell(33,6,str(results[j][3]),border=1,align='C')
                            pdf.cell(35,6,str(results[j][1]),border=1,align='C',ln=1)
                            amount=amount+results[j][1]
                            j=j+1
                            if j==int(x):
                                break
                        pdf.set_font('helvetica','BI',9)
                        pdf.cell(15,7,border=1)
                        pdf.cell(65,7,'Total (KES)',border=1,align='C')
                        pdf.cell(35,7,str(amount),border=1,align='C',ln=1)
                        pdf.cell(15,7,border=1)
                        pdf.cell(65,7,'Expected Pay (KES)',border=1,align='C')
                        pdf.cell(35,7,str(total_reported_termfee),border=1,align='C',ln=1)
                        pdf.cell(15,7,border=1)
                        pdf.cell(65,7,'Total Balance (KES)',border=1,align='C')
                        pdf.cell(35,7,str(total_balance),border=1,align='C',ln=1)
                        pdf.output('statements/'+st_admno+".pdf")
                        messagebox.showinfo("Success","Statement Generated")
                        file_to_print="\\statements\\"+st_admno+".pdf"
                        current_folder=os.getcwd()
                        file_to_print=f'{current_folder}'+file_to_print
                        #open created file
                        webbrowser.open(file_to_print)
                    except IOError:
                        pass
                    
                else:
                    messagebox.showinfo("No details","No student detail")
            except:
                pass
    framestfeepay=customtkinter.CTkFrame(fmaddfees,fg_color=bg3)
    framestfeepay.pack(side=TOP,pady=(2,4))
    lb=customtkinter.CTkLabel(framestfeepay,text="Student Fee Payment",text_color=bg1,font=fontlbl2)
    lb.grid(column=0,row=0,columnspan=5)
    lb=customtkinter.CTkLabel(framestfeepay,text="Admn No:",text_color=bg1,font=fontlbl)
    lb.grid(column=0,row=1,pady=(2,2))
    feestadmno=customtkinter.CTkEntry(framestfeepay,border_width=1,border_color=bg1,width=entry_width)
    feestadmno.grid(column=1,row=1,padx=(2,2),pady=(2,2))
    feestadmno.bind('<FocusOut>',getstudentname)
    lb=customtkinter.CTkLabel(framestfeepay,text="Name :",text_color=bg1,font=fontlbl)
    lb.grid(column=2,row=1,pady=(2,2))
    feestname=customtkinter.CTkEntry(framestfeepay,border_width=1,border_color=bg1,width=entry_width)
    feestname.grid(column=3,row=1,padx=(2,2),pady=(2,2),columnspan=3,sticky=W)
    lb=customtkinter.CTkLabel(framestfeepay,text="Amount:",text_color=bg1,font=fontlbl)
    lb.grid(column=0,row=2,pady=(2,2))
    feestamount=customtkinter.CTkEntry(framestfeepay,border_width=1,border_color=bg1,width=entry_width)
    feestamount.grid(column=1,row=2,padx=(2,2),pady=(2,2),sticky=W)
    lb=customtkinter.CTkLabel(framestfeepay,text="Paid Via:",text_color=bg1,font=fontlbl)
    lb.grid(column=2,row=2,pady=(2,2))
    paidvia=customtkinter.CTkOptionMenu(framestfeepay,fg_color="white",text_color="black",button_color=bg1,button_hover_color=bg1,width=entry_width,values=["Cash","Cheque","Equity","KCB"])
    paidvia.grid(column=3,row=2,padx=(2,2),pady=(2,2),sticky=W)
    lb=customtkinter.CTkLabel(framestfeepay,text="Remarks:",text_color=bg1,font=fontlbl)
    lb.grid(column=0,row=3,pady=(2,2))
    feeremarks=customtkinter.CTkEntry(framestfeepay,border_width=1,border_color=bg1,width=entry_width)
    feeremarks.grid(column=1,row=3,padx=(2,2),pady=(2,2),sticky=W)
    feepaybtnf1=customtkinter.CTkFrame(framestfeepay)
    feepaybtnf1.grid(column=0,row=4,columnspan=4,pady=(5,5))
    btnpayfee=customtkinter.CTkButton(feepaybtnf1,text="Pay",text_color=fg1,fg_color=bg1,hover_color=hovbg1,command=addfees)
    btnpayfee.grid(column=0,row=0,padx=(3,3),pady=(2,2))
    btnfeestatement=customtkinter.CTkButton(feepaybtnf1,text="Statement",text_color=fg1,fg_color=bg1,hover_color=hovbg1,command=checkstatement)
    btnfeestatement.grid(column=1,row=0,padx=(3,3),pady=(2,2))
    btnpayfeecancel=customtkinter.CTkButton(feepaybtnf1,text="Cancel",text_color=fg1,fg_color=coldark,hover_color=hovbg1,command=clearallpaymentdetails)
    btnpayfeecancel.grid(column=2,row=0,padx=(3,3),pady=(2,2))
    frameTablestatement=customtkinter.CTkFrame(fmaddfees,fg_color=bg3)
    frameTablestatement.pack(side=TOP,fill=X)
    tablestatement=ttk.Treeview(frameTablestatement,height=8)
    tablestatement.pack(fill=BOTH,side=LEFT,expand=1,padx=(4,0),pady=(4,4))
    tablestatement['show']="headings"
    tablestatement['columns']=(0,1,2,3)
    tablestatement.heading(0,text="Adm No")
    tablestatement.heading(1,text="Amount")
    tablestatement.heading(2,text="Paid via")
    tablestatement.heading(3,text="Date")
    tablestatement.column(0,width=80,anchor=CENTER)
    tablestatement.column(1,width=100,anchor=CENTER)
    tablestatement.column(2,width=100,anchor=CENTER)
    tablestatement.column(3,width=100,anchor=CENTER)
    scroll_statement=customtkinter.CTkScrollbar(frameTablestatement,command=tablestatement.yview)
    scroll_statement.pack(fill=Y,side=LEFT)
    tablestatement.configure(yscrollcommand=scroll_statement)
    frame_btn_state=customtkinter.CTkFrame(fmaddfees,fg_color=bg3)
    frame_btn_state.pack(fill=X,side=TOP)
    lbl_total_expected_amount=customtkinter.CTkLabel(frame_btn_state,text="Expected Pay",text_color=bg1,font=("times",14,"bold"))
    lbl_total_expected_amount.grid(column=0,row=0,padx=(10,20),pady=4)
    lbl_total_fee_amount=customtkinter.CTkLabel(frame_btn_state,text="Paid amount",text_color=bg1,font=("times",14,"bold"))
    lbl_total_fee_amount.grid(column=1,row=0,padx=(10,20),pady=4)
    lbl_total_balance_amount=customtkinter.CTkLabel(frame_btn_state,text="Balance",text_color=bg1,font=("times",14,"bold"))
    lbl_total_balance_amount.grid(column=2,row=0,padx=(10,20),pady=4)
    btn_print_statement=customtkinter.CTkButton(frame_btn_state,text="Print",text_color=fg1,fg_color=bg1,hover_color=hovbg1,command=print_statement)
    btn_print_statement.grid(column=3,row=0,padx=4,pady=4)
    #------------------------------------------
    #------------------view paid---------------
    def viewpaid():
        try:
            sql="SELECT * FROM fees"
            cursor=conn.cursor()
            cursor.execute(sql)
            results=cursor.fetchall()
            if results:
                for records in tablepaid.get_children():
                    tablepaid.delete(records)
                for i in results:
                    tablepaid.insert('',END,values=i)
        except IOError:
            pass
    def filterfeedb(event,*arg):
        if filterfeeadmno.get()=="":
            try:
                sql="SELECT * FROM fees WHERE paymentdate BETWEEN '%s' AND '%s'"%(filterfeefromdate.get(),filterfeetodate.get())
                cursor=conn.cursor()
                cursor.execute(sql)
                results=cursor.fetchall()
                if results:
                    for records in tablepaid.get_children():
                        tablepaid.delete(records)
                    for i in results:
                        tablepaid.insert('',END,values=i)
                else:
                    for records in tablepaid.get_children():
                        tablepaid.delete(records)
            except IOError:
                pass
        else:
            try:
                sql="SELECT * FROM fees WHERE stadmno='%s' AND paymentdate BETWEEN '%s' AND '%s'"%(filterfeeadmno.get(),filterfeefromdate.get(),filterfeetodate.get())
                cursor=conn.cursor()
                cursor.execute(sql)
                results=cursor.fetchall()
                if results:
                    for records in tablepaid.get_children():
                        tablepaid.delete(records)
                    for i in results:
                        tablepaid.insert('',END,values=i)
                else:
                    for records in tablepaid.get_children():
                        tablepaid.delete(records)
            except IOError:
                pass
    fmfeefilter=customtkinter.CTkFrame(fmviewpaid,fg_color=bg3)
    fmfeefilter.pack(fill=X,side=TOP)
    lb=customtkinter.CTkLabel(fmfeefilter,text="Filter: ",font=fontlbl,text_color=bg1,justify=RIGHT)
    lb.grid(column=0,row=0)
    lb=customtkinter.CTkLabel(fmfeefilter,text="From Date: ",font=fontlbl,text_color=bg1,justify=RIGHT)
    lb.grid(column=0,row=1)
    filterfeefromdate=DateEntry(fmfeefilter,date_pattern="yyyy/mm/dd")
    filterfeefromdate.grid(column=1,row=1)
    filterfeefromdate.bind('<<DateEntrySelected>>',filterfeedb)
    lb=customtkinter.CTkLabel(fmfeefilter,text="To Date: ",font=fontlbl,text_color=bg1,justify=RIGHT)
    lb.grid(column=2,row=1)
    filterfeetodate=DateEntry(fmfeefilter,date_pattern="yyyy/mm/dd")
    filterfeetodate.grid(column=3,row=1)
    filterfeetodate.bind('<<DateEntrySelected>>',filterfeedb)
    lb=customtkinter.CTkLabel(fmfeefilter,text="Admno: ",font=fontlbl,text_color=bg1,justify=RIGHT)
    lb.grid(column=4,row=1)
    filterfeeadmno=Entry(fmfeefilter)
    filterfeeadmno.grid(column=5,row=1)
    filterfeeadmno.bind('<FocusOut>',filterfeedb)
    tablepaid=ttk.Treeview(fmviewpaid)
    tablepaid.pack(side=LEFT,fill=BOTH,expand=True)
    tablepaid['show']="headings"
    tablepaid['columns']=(0,1,2,3,4,5,6)
    tablepaid.heading(0,text="Id")
    tablepaid.heading(1,text="Admno")
    tablepaid.heading(2,text="Amount")
    tablepaid.heading(3,text="paidvia")
    tablepaid.heading(4,text="paidvia")
    tablepaid.heading(5,text="Date")
    tablepaid.heading(6,text="Record By")
    tablepaid.column(0,width=60,anchor=CENTER)
    tablepaid.column(1,width=80,anchor=CENTER)
    tablepaid.column(2,width=130,anchor=CENTER)
    tablepaid.column(3,width=60,anchor=CENTER)
    tablepaid.column(4,width=100,anchor=CENTER)
    tablepaid.column(5,width=100,anchor=CENTER)
    tablepaid.column(6,width=100,anchor=CENTER)
    scroll_table_paid=customtkinter.CTkScrollbar(fmviewpaid,command=tablepaid.yview)
    scroll_table_paid.pack(fill=Y,side=LEFT)
    tablepaid.configure(yscrollcommand=scroll_table_paid)
    viewpaid()
    #-------------fmsalarie--------------------
    def viewsalaries():
        try:
            sql="SELECT * FROM salaries"
            cursor=conn.cursor()
            cursor.execute(sql)
            results=cursor.fetchall()
            if results:
                for records in tablesalaries.get_children():
                    tablesalaries.delete(records)
                for i in results:
                    tablesalaries.insert('',END,values=i)
        except IOError:
            pass
    def getsalstaffdetails(*arg):
        global salarystfid
        try:
            sql="SELECT staffid,staffno,CONCAT(fname,' ',lname,' ',sname) AS name FROM staffs WHERE staffno='%s'"%(salstaffno.get())
            cursor=conn.cursor()
            cursor.execute(sql)
            results=cursor.fetchall()
            if results:
                clearsalstaffdetails()
                salarystfid=results[0][0]
                salstaffno.insert(END,results[0][1])
                salstaffid.insert(END,results[0][0])
                salstaffname.insert(END,results[0][2])
            else:
                messagebox.showwarning("Error","Invalid staff no")
        except IOError:
            pass
    def clearsalstaffdetails():
        global salarystfid
        salarystfid=""
        salstaffno.delete(0,END)
        salstaffid.delete(0,END)
        salstaffname.delete(0,END)
        salamount.delete(0,END)
    def addsalaries():
        global myusername,salarystfid
        if salstaffid.get()=="":
            return False
        elif salstaffno.get()=="":
            return False
        elif salstaffname.get()=="":
            return False
        elif salamount.get()=="":
            return False
        else:
            #check if staff is gov/bom salaries
            #print(salarystfid)
            try:
                sql1="SELECT employer FROM staffs WHERE staffid=%s"%(salarystfid)
                cursor=conn.cursor()
                cursor.execute(sql1)
                sqlcheck1=cursor.fetchall()
                if sqlcheck1:
                    if sqlcheck1[0][0]=="Government":
                        messagebox.showerror("Error","The staff is Paid by Government")
                        clearsalstaffdetails()
                    elif sqlcheck1[0][0]=="B.O.M":
                        #check if staff is in salary sheet
                        try:
                            sql2="SELECT * FROM salaries WHERE staffid=%s"%(salarystfid)
                            cursor=conn.cursor()
                            cursor.execute(sql2)
                            sqlcheck2=cursor.fetchall()
                            if sqlcheck2:
                                msg=messagebox.askyesno("Update message","Do you want to update this salary of staff no"+salstaffno.get())
                                if msg==True:
                                    sql3="UPDATE salaries SET amount=%s,recordby='%s' WHERE staffid=%s"%(salamount.get(),myusername,salarystfid)
                                    cursor=conn.cursor()
                                    cursor.execute(sql3)
                                    conn.commit()
                                    messagebox.showinfo("success","salary updated successfully")
                                    viewsalaries()
                                    clearsalstaffdetails()
                                    get_school_finance_analysis()
                                    get_reporting_analysis()
                                    get_resource_analysis()
                            else:
                                sql="INSERT INTO salaries(staffid,staffno,amount,recordby) VALUES(%s,'%s',%s,'%s')"
                                cursor=conn.cursor()
                                cursor.execute(sql%(salarystfid,salstaffno.get(),salamount.get(),myusername))
                                conn.commit()
                                messagebox.showinfo("success","salary saved successfully")
                                viewsalaries()
                                clearsalstaffdetails()
                                get_school_finance_analysis()
                                get_reporting_analysis()
                                get_resource_analysis()
                        except IOError:
                            pass
            except IOError:
                pass 
    frameaddstaffsalary=customtkinter.CTkFrame(fmsalaries,fg_color=bg3)
    frameaddstaffsalary.pack(fill=X,side=TOP)
    lb=customtkinter.CTkLabel(frameaddstaffsalary,text="Add staff salary",font=fontlbl2,text_color=bg1)
    lb.grid(column=0,row=0,columnspan=4)
    lb=customtkinter.CTkLabel(frameaddstaffsalary,text="Staff No",font=fontlbl,text_color=bg1)
    lb.grid(column=0,row=1)
    salstaffno=customtkinter.CTkEntry(frameaddstaffsalary,border_width=1,border_color=bg1,width=entry_width)
    salstaffno.grid(column=1,row=1)
    salstaffno.bind('<FocusOut>',getsalstaffdetails)
    lb=customtkinter.CTkLabel(frameaddstaffsalary,text="Name",font=fontlbl,text_color=bg1)
    lb.grid(column=2,row=1)
    salstaffname=customtkinter.CTkEntry(frameaddstaffsalary,border_width=1,border_color=bg1,width=entry_width)
    salstaffname.grid(column=3,row=1)
    lb=customtkinter.CTkLabel(frameaddstaffsalary,text="Staff ID",font=fontlbl,text_color=bg1)
    lb.grid(column=0,row=2)
    salstaffid=customtkinter.CTkEntry(frameaddstaffsalary,border_width=1,border_color=bg1,width=entry_width)
    salstaffid.grid(column=1,row=2)
    lb=customtkinter.CTkLabel(frameaddstaffsalary,text="Salary",font=fontlbl,text_color=bg1)
    lb.grid(column=2,row=2)
    salamount=customtkinter.CTkEntry(frameaddstaffsalary,border_width=1,border_color=bg1,width=entry_width)
    salamount.grid(column=3,row=2)
    btnsavesalary=customtkinter.CTkButton(frameaddstaffsalary,text="Save",fg_color=bg1,text_color=fg1,cursor="hand2",hover_color=hovbg1,command=addsalaries)
    btnsavesalary.grid(column=4,row=1,padx=(5,5),pady=(5,5))
    btncancelsalary=customtkinter.CTkButton(frameaddstaffsalary,text="Cancel",fg_color=coldark,text_color=fg1,cursor="hand2",hover_color=hovbg1,command=clearsalstaffdetails)
    btncancelsalary.grid(column=4,row=2,padx=(5,5),pady=(5,5))
    frameviewsalaries=customtkinter.CTkFrame(fmsalaries,fg_color=bg3)
    frameviewsalaries.pack(fill=X,expand=True)
    tablesalaries=ttk.Treeview(frameviewsalaries)
    tablesalaries.pack(side=LEFT,fill=BOTH,expand=True)
    tablesalaries['show']="headings"
    tablesalaries['columns']=(0,1,2,3,4)
    tablesalaries.heading(0,text="Id")
    tablesalaries.heading(1,text="Staff ID")
    tablesalaries.heading(2,text="Staff No")
    tablesalaries.heading(3,text="Amount")
    tablesalaries.heading(4,text="Record By")
    tablesalaries.column(0,width=60,anchor=CENTER)
    tablesalaries.column(1,width=60,anchor=CENTER)
    tablesalaries.column(2,width=80,anchor=CENTER)
    tablesalaries.column(3,width=100,anchor=CENTER)
    tablesalaries.column(4,width=100,anchor=CENTER)
    viewsalaries()
    #------------------------------------------
    #---------------projects-------------------
    def addproject():
        global myusername
        if projectname.get()=="":
            return False
        elif projectdescription.get()=="":
            return False
        elif projectcost.get()=="":
            return False
        elif projectstart.get()=="":
            return False
        elif projectend.get()=="":
            return False
        else:
            try:
                sql="INSERT INTO projects(projectname,projectdescription,projectcost,projectstart,projectexpectedend,recordby) VALUES('%s','%s',%s,'%s','%s','%s')"
                cursor=conn.cursor()
                cursor.execute(sql%(projectname.get(),projectdescription.get(),projectcost.get(),projectstart.get(),projectend.get(),myusername))
                conn.commit()
                clearprojectsdeatails()
                viewprojects()
                get_school_finance_analysis()
                get_reporting_analysis()
                get_resource_analysis()
                messagebox.showinfo("success","project saved successfully")
            except IOError:
                pass
    def clearprojectsdeatails():
        projectname.delete(0,END)
        projectcost.delete(0,END)
        projectdescription.delete(0,END)
        projectstart.delete(0,END)
        projectend.delete(0,END)
        projectstart.insert(END,today)
        projectend.insert(END,today)
    def viewprojects():
        try:
            sql="SELECT * FROM projects"
            cursor=conn.cursor()
            cursor.execute(sql)
            results=cursor.fetchall()
            if results:
                for records in tableprojects.get_children():
                    tableprojects.delete(records)
                for i in results:
                    tableprojects.insert('',END,values=i)
        except IOError:
            pass
    frameaddproject=customtkinter.CTkFrame(fmprojects,fg_color=bg3)
    frameaddproject.pack(fill=X,expand=True)
    lb=customtkinter.CTkLabel(frameaddproject,text="Add Project",text_color=bg1,font=fontlbl2)
    lb.grid(column=0,row=0,columnspan=4)
    lb=customtkinter.CTkLabel(frameaddproject,text="Project Name",text_color=bg1,font=fontlbl)
    lb.grid(column=0,row=1)
    projectname=customtkinter.CTkEntry(frameaddproject,border_width=1,border_color=bg1,width=textboxwidth)
    projectname.grid(column=1,row=1)
    lb=customtkinter.CTkLabel(frameaddproject,text="Description",text_color=bg1,font=fontlbl)
    lb.grid(column=0,row=2)
    projectdescription=customtkinter.CTkEntry(frameaddproject,height=70,border_width=1,border_color=bg1,width=textboxwidth)
    projectdescription.grid(column=1,row=2,rowspan=2)
    lb=customtkinter.CTkLabel(frameaddproject,text="Project Cost",text_color=bg1,font=fontlbl)
    lb.grid(column=2,row=1)
    projectcost=customtkinter.CTkEntry(frameaddproject,border_width=1,border_color=bg1,width=textboxwidth)
    projectcost.grid(column=3,row=1)
    lb=customtkinter.CTkLabel(frameaddproject,text="Project Start",text_color=bg1,font=fontlbl)
    lb.grid(column=2,row=2)
    projectstart=DateEntry(frameaddproject,date_pattern="yyyy/mm/dd",width=19,font=fontlbl)
    projectstart.grid(column=3,row=2)
    lb=customtkinter.CTkLabel(frameaddproject,text="Project End",text_color=bg1,font=fontlbl)
    lb.grid(column=2,row=3)
    projectend=DateEntry(frameaddproject,date_pattern="yyyy/mm/dd",width=19,font=fontlbl)
    projectend.grid(column=3,row=3)
    frambtnprojects=customtkinter.CTkFrame(frameaddproject)
    frambtnprojects.grid(column=0,row=4,columnspan=4,pady=(10,10))
    btnsaveproject=customtkinter.CTkButton(frambtnprojects,text="Save",text_color=fg1,fg_color=bg1,hover_color=hovbg1,cursor="hand2",command=addproject)
    btnsaveproject.grid(column=0,row=0,pady=10,padx=12)
    btncancelproject=customtkinter.CTkButton(frambtnprojects,text="Cancel",text_color=fg1,fg_color=coldark,hover_color=hovbg1,cursor="hand2",command=clearprojectsdeatails)
    btncancelproject.grid(column=1,row=0,pady=10,padx=12)
    tableprojects=ttk.Treeview(fmprojects)
    tableprojects.pack(fill=X,expand=True)
    tableprojects['show']="headings"
    tableprojects['columns']=(0,1,2,3,4,5,6)
    tableprojects.heading(0,text="Id")
    tableprojects.heading(1,text="Name")
    tableprojects.heading(2,text="Description")
    tableprojects.heading(3,text="Cost")
    tableprojects.heading(4,text="Start")
    tableprojects.heading(5,text="Expected End")
    tableprojects.heading(6,text="Record By")
    tableprojects.column(0,width=30,anchor=CENTER)
    tableprojects.column(1,width=90,anchor=CENTER)
    tableprojects.column(2,width=150,anchor=CENTER)
    tableprojects.column(3,width=80,anchor=CENTER)
    tableprojects.column(4,width=80,anchor=CENTER)
    tableprojects.column(5,width=80,anchor=CENTER)
    tableprojects.column(6,width=80,anchor=CENTER)
    viewprojects()
    #------------------------------------------
    #--------------------petty cash------------
    def addpettycash():
        global myusername
        if ptccategory.get()=="":
            return False
        elif ptcnarration.get()=="":
            return False
        elif ptcamount.get()=="":
            return False
        else:
            try:
                sql="INSERT INTO pettycash(ptcategory,ptnarration,ptamount,ptrecordby,ptapproved) VALUES('%s','%s',%s,'%s',%s)"
                cursor=conn.cursor()
                cursor.execute(sql%(ptccategory.get(),ptcnarration.get(),ptcamount.get(),myusername,0))
                conn.commit()
                messagebox.showinfo("success","Your request has been posted successful.\n wait for Approval")
                clearpettycashinput()
                viewpettycash()
                get_petty_cash_toapprove()
                get_school_finance_analysis()
                get_reporting_analysis()
                get_resource_analysis()
            except IOError:
                pass
    def clearpettycashinput():
        ptccategory.set("Tution")
        ptcnarration.delete(0,END)
        ptcamount.delete(0,END)
    def viewpettycash():
        try:
            sql="SELECT * FROM pettycash"
            cursor=conn.cursor()
            cursor.execute(sql)
            results=cursor.fetchall()
            if results:
                for records in tablepettycash.get_children():
                    tablepettycash.delete(records)
                for i in results:
                    tablepettycash.insert('',END,values=i)
        except IOError:
            pass
    
    frameaddpettycash=customtkinter.CTkFrame(fmpettycash,fg_color=bg3)
    frameaddpettycash.pack(side=TOP,fill=X,expand=True)
    lb=customtkinter.CTkLabel(frameaddpettycash,text="Post request",text_color=bg1,font=fontlbl2)
    lb.grid(column=0,row=0)
    lb=customtkinter.CTkLabel(frameaddpettycash,text="Category",text_color=bg1,font=fontlbl2)
    lb.grid(column=0,row=1)
    ptccategory=customtkinter.CTkOptionMenu(frameaddpettycash,width=entry_width,text_color="black",fg_color="white",button_color=bg1,button_hover_color=hovbg1,values=["Tution","Staff-Tea","Remedial","Tour","Overtime","Pricegiving"])
    ptccategory.grid(column=1,row=1)
    lb=customtkinter.CTkLabel(frameaddpettycash,text="Narration",text_color=bg1,font=fontlbl2)
    lb.grid(column=0,row=2)
    ptcnarration=customtkinter.CTkEntry(frameaddpettycash,border_width=1,border_color=bg1,width=entry_width)
    ptcnarration.grid(column=1,row=2)
    lb=customtkinter.CTkLabel(frameaddpettycash,text="Amount",text_color=bg1,font=fontlbl2)
    lb.grid(column=2,row=1)
    ptcamount=customtkinter.CTkEntry(frameaddpettycash,border_width=1,border_color=bg1,width=entry_width)
    ptcamount.grid(column=3,row=1)
    btnsavepettycash=customtkinter.CTkButton(frameaddpettycash,text="Save",text_color=fg1,fg_color=bg1,hover_color=hovbg1,cursor="hand2",command=addpettycash)
    btnsavepettycash.grid(column=4,row=1,pady=10,padx=12)
    btncancelpettycash=customtkinter.CTkButton(frameaddpettycash,text="Cancel",text_color=fg1,fg_color=coldark,hover_color=hovbg1,cursor="hand2",command=clearpettycashinput)
    btncancelpettycash.grid(column=4,row=2,pady=10,padx=12)
    tablepettycash=ttk.Treeview(fmpettycash)
    tablepettycash.pack(fill=BOTH,side=TOP,expand=True)
    tablepettycash['show']="headings"
    tablepettycash['columns']=(0,1,2,3,4,5,6,7)
    tablepettycash.heading(0,text="Id")
    tablepettycash.heading(1,text="Category")
    tablepettycash.heading(2,text="Narration")
    tablepettycash.heading(3,text="Amount")
    tablepettycash.heading(4,text="Recordby")
    tablepettycash.heading(5,text="Approved")
    tablepettycash.heading(6,text="Approved By")
    tablepettycash.heading(7,text="Date Approved")
    tablepettycash.column(0,width=30,anchor=CENTER)
    tablepettycash.column(1,width=70,anchor=CENTER)
    tablepettycash.column(2,width=150,anchor=CENTER)
    tablepettycash.column(3,width=80,anchor=CENTER)
    tablepettycash.column(4,width=80,anchor=CENTER)
    tablepettycash.column(5,width=40,anchor=CENTER)
    tablepettycash.column(6,width=80,anchor=CENTER)
    tablepettycash.column(7,width=80,anchor=CENTER)
    viewpettycash()
    #------------------------------------------
    #==========================================
    #=========framelibary====================
    def toaddbook():
        ntbook6.select(0)
    def toeditbook():
        ntbook6.select(1)
    def toviewbook():
        ntbook6.select(2)
    def toissuebook():
        ntbook6.select(3)
    def toviewissuedbook():
        ntbook6.select(4)

    framesubmenu6=customtkinter.CTkFrame(framelibrary,width=150,fg_color=coldark)
    framesubmenu6.pack(side=LEFT,fill=Y)
    btnaddbook=customtkinter.CTkButton(framesubmenu6,text="Add Books",text_color=fg1,fg_color=bg1,hover_color=hovbg1,cursor="hand2",command=toaddbook)
    btnaddbook.pack(pady=6,padx=6)
    btneditbook=customtkinter.CTkButton(framesubmenu6,text="Edit Books",text_color=fg1,fg_color=bg1,hover_color=hovbg1,cursor="hand2",command=toeditbook)
    btneditbook.pack(pady=6,padx=6)
    btnviewbooks=customtkinter.CTkButton(framesubmenu6,text="View Books",text_color=fg1,fg_color=bg1,hover_color=hovbg1,cursor="hand2",command=toviewbook)
    btnviewbooks.pack(pady=6,padx=6)
    btnissuebook=customtkinter.CTkButton(framesubmenu6,text="Issue Book",text_color=fg1,fg_color=bg1,hover_color=hovbg1,cursor="hand2",command=toissuebook)
    btnissuebook.pack(pady=6,padx=6)
    btnviewissued=customtkinter.CTkButton(framesubmenu6,text="View Issued",text_color=fg1,fg_color=bg1,hover_color=hovbg1,cursor="hand2",command=toviewissuedbook)
    btnviewissued.pack(pady=6,padx=6)
    ntbook6=ttk.Notebook(framelibrary)
    ntbook6.pack(side=LEFT,fill=BOTH,expand=True)
    fmaddbook=customtkinter.CTkFrame(ntbook6,fg_color=bg3,bg_color=coldark)
    fmaddbook.place(x=0,y=0,relheight=1,relwidth=1)
    fmeditbook=customtkinter.CTkFrame(ntbook6,fg_color=bg3,bg_color=coldark)
    fmeditbook.place(x=0,y=0,relheight=1,relwidth=1)
    fmviewbooks=customtkinter.CTkFrame(ntbook6,fg_color=bg3,bg_color=coldark)
    fmviewbooks.place(x=0,y=0,relheight=1,relwidth=1)
    fmissuebook=customtkinter.CTkFrame(ntbook6,fg_color=bg3,bg_color=coldark)
    fmissuebook.place(x=0,y=0,relheight=1,relwidth=1)
    fmviewissued=customtkinter.CTkFrame(ntbook6,fg_color=bg3,bg_color=coldark)
    fmviewissued.place(x=0,y=0,relheight=1,relwidth=1)
    ntbook6.add(fmaddbook,text="Add Book")
    ntbook6.add(fmeditbook,text="Edit Book")
    ntbook6.add(fmviewbooks,text="View Books")
    ntbook6.add(fmissuebook,text="Issue Book")
    ntbook6.add(fmviewissued,text="View Issued")
    #-------------------add books----------------------
    def addbooks():
        if bkcategory.get()=="":
            return False
        elif bktitle.get()=="":
            return False
        elif bkauthor.get()=="":
            return False
        elif bkpublisher.get()=="":
            return False
        elif bkpubdate.get()=="":
            return False
        elif bkquantity.get()=="":
            return False
        elif bkprice.get()=="":
            return False
        elif bkno.get()=="":
            return False
        else:
            global myusername,today
            try:
                sql="INSERT INTO books(bookcategory,booktitle,bookauthor,bookpublisher,publishdate,quantity,bookprice,bookno,recordby,recorddate) VALUES('%s','%s','%s','%s','%s',%s,%s,'%s','%s','%s')"
                cursor=conn.cursor()
                cursor.execute(sql%(bkcategory.get(),bktitle.get(),bkauthor.get(),bkpublisher.get(),bkpubdate.get(),bkquantity.get(),bkprice.get(),bkno.get(),myusername,today))
                conn.commit()
                messagebox.showinfo("success","Book saved successfully")
                clearbookdetails()
                viewbooks()
                get_school_finance_analysis()
                get_reporting_analysis()
                get_resource_analysis()
            except IOError:
                pass
    def clearbookdetails():
        bkcategory.set("Mathematics")
        bktitle.delete(0,END)
        bkauthor.delete(0,END)
        bkpublisher.delete(0,END)
        bkpubdate.delete(0,END)
        bkpubdate.insert(END,today)
        bkquantity.delete(0,END)
        bkprice.delete(0,END)
        bkno.delete(0,END)
    lb=customtkinter.CTkLabel(fmaddbook,text="Add book",font=fontlbl2,text_color=bg1,justify=CENTER)
    lb.grid(column=0,row=0,columnspan=4,pady=12,padx=5)
    lb=customtkinter.CTkLabel(fmaddbook,text="Category:",font=fontlbl,text_color=bg1,justify=RIGHT)
    lb.grid(column=0,row=1,pady=12,padx=5,sticky=E)
    bkcategory=customtkinter.CTkOptionMenu(fmaddbook,fg_color="white",width=150,button_color=bg1,text_color="black",button_hover_color=bg1,values=["Mathematics","English","Kiswahili","Sciences","Humanities","Technical","Moral","Health","Research"])
    bkcategory.grid(column=1,row=1,pady=12,padx=5)
    lb=customtkinter.CTkLabel(fmaddbook,text="Title:",font=fontlbl,text_color=bg1,justify=RIGHT)
    lb.grid(column=2,row=1,pady=12,padx=5,sticky=E)
    bktitle=customtkinter.CTkEntry(fmaddbook,border_width=1,border_color=bg1,width=150)
    bktitle.grid(column=3,row=1,pady=12,padx=5)
    lb=customtkinter.CTkLabel(fmaddbook,text="Author:",font=fontlbl,text_color=bg1,justify=RIGHT)
    lb.grid(column=0,row=2,pady=12,padx=5,sticky=E)
    bkauthor=customtkinter.CTkEntry(fmaddbook,border_width=1,border_color=bg1,width=150)
    bkauthor.grid(column=1,row=2,pady=12,padx=5)
    lb=customtkinter.CTkLabel(fmaddbook,text="Publisher:",font=fontlbl,text_color=bg1,justify=RIGHT)
    lb.grid(column=2,row=2,pady=12,padx=5,sticky=E)
    bkpublisher=customtkinter.CTkEntry(fmaddbook,border_width=1,border_color=bg1,width=150)
    bkpublisher.grid(column=3,row=2,pady=12,padx=5)

    lb=customtkinter.CTkLabel(fmaddbook,text="Date Published:",font=fontlbl,text_color=bg1,justify=RIGHT)
    lb.grid(column=0,row=3,pady=12,padx=5,sticky=E)
    bkpubdate=DateEntry(fmaddbook,width=14,font=fontlbl,date_pattern="yyyy/mm/dd")
    bkpubdate.grid(column=1,row=3,pady=12,padx=5)
    lb=customtkinter.CTkLabel(fmaddbook,text="Quantity:",font=fontlbl,text_color=bg1,justify=RIGHT)
    lb.grid(column=2,row=3,pady=12,padx=5,sticky=E)
    bkquantity=customtkinter.CTkEntry(fmaddbook,border_width=1,border_color=bg1,width=150)
    bkquantity.grid(column=3,row=3,pady=12,padx=5)
    lb=customtkinter.CTkLabel(fmaddbook,text="Price/book:",font=fontlbl,text_color=bg1,justify=RIGHT)
    lb.grid(column=0,row=4,pady=12,padx=5,sticky=E)
    bkprice=customtkinter.CTkEntry(fmaddbook,border_width=1,border_color=bg1,width=150)
    bkprice.grid(column=1,row=4,pady=12,padx=5)
    lb=customtkinter.CTkLabel(fmaddbook,text="Book No:",font=fontlbl,text_color=bg1,justify=RIGHT)
    lb.grid(column=2,row=4,pady=12,padx=5,sticky=E)
    bkno=customtkinter.CTkEntry(fmaddbook,border_width=1,border_color=bg1,width=150)
    bkno.grid(column=3,row=4,pady=12,padx=5)
    framebtnaddbook=customtkinter.CTkFrame(fmaddbook)
    framebtnaddbook.grid(column=0,row=5,columnspan=4)
    btnsavebook=customtkinter.CTkButton(framebtnaddbook,text="Save",fg_color=bg1,text_color=fg1,hover_color=hovbg1,cursor="hand2",command=addbooks)
    btnsavebook.grid(column=0,row=0,pady=5,padx=6)
    btncancelbook=customtkinter.CTkButton(framebtnaddbook,text="Cancel",fg_color=coldark,text_color=fg1,hover_color=hovbg1,cursor="hand2",command=clearbookdetails)
    btncancelbook.grid(column=1,row=0,pady=5,padx=6)
    #-------------------edit book----------------
    def searchbook():
        global bookid
        try:
            sql="SELECT * FROM books WHERE bookno='%s'"%(sbkno.get())
            cursor=conn.cursor()
            cursor.execute(sql)
            results=cursor.fetchall()
            if results:
                cleareditbook()
                bookid=results[0][0]
                ebkcategory.set(results[0][1])
                ebktitle.insert(END,results[0][2])
                ebkauthor.insert(END,results[0][3])
                ebkpublisher.insert(END,results[0][4])
                ebkpubdate.insert(END,results[0][5])
                ebkquantity.insert(END,results[0][6])
                ebkprice.insert(END,results[0][7])
                ebkno.insert(END,results[0][8])
            else:
                messagebox.showwarning("Error","Invalid Book No")
                sbkno.delete(0,END)
                cleareditbook()
        except IOError:
            pass
    def cleareditbook():
        global bookid
        bookid=""
        ebkcategory.set("Mathematics")
        ebktitle.delete(0,END)
        ebkauthor.delete(0,END)
        ebkpublisher.delete(0,END)
        ebkpubdate.delete(0,END)
        ebkquantity.delete(0,END)
        ebkprice.delete(0,END)
        ebkno.delete(0,END)
    def updatebook():
        if ebkcategory.get()=="":
            return False
        elif ebktitle.get()=="":
            return False
        elif ebkauthor.get()=="":
            return False
        elif ebkpublisher.get()=="":
            return False
        elif ebkpubdate.get()=="":
            return False
        elif ebkquantity.get()=="":
            return False
        elif ebkprice.get()=="":
            return False
        elif ebkno.get()=="":
            return False
        else:
            global bookid,myusername
            try:
                sql="UPDATE books SET bookcategory='%s',booktitle='%s',bookauthor='%s',bookpublisher='%s',publishdate='%s',quantity=%s,bookprice=%s,bookno='%s',recordby='%s',recorddate='%s' WHERE bookid=%s"%(ebkcategory.get(),ebktitle.get(),ebkauthor.get(),ebkpublisher.get(),ebkpubdate.get(),ebkquantity.get(),ebkprice.get(),ebkno.get(),myusername,today,bookid)
                cursor=conn.cursor()
                cursor.execute(sql)
                conn.commit()
                messagebox.showinfo("success","Book updated successfully")
                cleareditbook()
                viewbooks()
                get_school_finance_analysis()
                get_reporting_analysis()
                get_resource_analysis()
            except IOError:
                pass
                cleareditbook()
    def deletebook():
        global bookid
        if ebkno.get()=="":
            messagebox.showerror("Error","Search book to delete")
        else:
            conmessage=messagebox.askyesno("Confirm message","Do you want to delete this book "+ebkno.get())
            if conmessage==True:
                try:
                    sql="DELETE FROM books WHERE bookid=%s"%(bookid)
                    cursor=conn.cursor()
                    cursor.execute(sql)
                    conn.commit()
                    messagebox.showinfo("Success","book deleted successful")
                    cleareditbook()
                    viewbooks()
                except IOError:
                    messagebox.showwarning("Error","Unable to delete book.\n Try Again")
                    cleareditbook()
                    pass
    framesearchbook=customtkinter.CTkFrame(fmeditbook,fg_color=bg3)
    framesearchbook.grid(column=0,row=0,columnspan=4)
    lb=customtkinter.CTkLabel(framesearchbook,text="Search book",font=fontlbl,text_color=bg1)
    lb.grid(column=0,row=0)
    lb=customtkinter.CTkLabel(framesearchbook,text="Book No",font=fontlbl,text_color=bg1)
    lb.grid(column=1,row=0)
    sbkno=customtkinter.CTkEntry(framesearchbook,border_width=1,border_color=bg1,width=150)
    sbkno.grid(column=2,row=0,pady=5,padx=5)
    btnsearchbook=customtkinter.CTkButton(framesearchbook,text="Search",fg_color=bg1,text_color=fg1,hover_color=hovbg1,cursor="hand2",command=searchbook)
    btnsearchbook.grid(column=3,row=0,pady=5,padx=6)

    lb=customtkinter.CTkLabel(fmeditbook,text="Edit book",font=fontlbl2,text_color=bg1,justify=CENTER)
    lb.grid(column=0,row=1,columnspan=4,pady=12,padx=5)
    lb=customtkinter.CTkLabel(fmeditbook,text="Category:",font=fontlbl,text_color=bg1,justify=RIGHT)
    lb.grid(column=0,row=2,pady=12,padx=5,sticky=E)
    ebkcategory=customtkinter.CTkOptionMenu(fmeditbook,fg_color="white",width=150,button_color=bg1,text_color="black",button_hover_color=bg1,values=["Mathematics","English","Kiswahili","Sciences","Humanities","Technical","Moral","Health","Research"])
    ebkcategory.grid(column=1,row=2,pady=12,padx=5)
    lb=customtkinter.CTkLabel(fmeditbook,text="Title:",font=fontlbl,text_color=bg1,justify=RIGHT)
    lb.grid(column=2,row=2,pady=12,padx=5,sticky=E)
    ebktitle=customtkinter.CTkEntry(fmeditbook,border_width=1,border_color=bg1,width=150)
    ebktitle.grid(column=3,row=2,pady=12,padx=5)
    lb=customtkinter.CTkLabel(fmeditbook,text="Author:",font=fontlbl,text_color=bg1,justify=RIGHT)
    lb.grid(column=0,row=3,pady=12,padx=5,sticky=E)
    ebkauthor=customtkinter.CTkEntry(fmeditbook,border_width=1,border_color=bg1,width=150)
    ebkauthor.grid(column=1,row=3,pady=12,padx=5)
    lb=customtkinter.CTkLabel(fmeditbook,text="Publisher:",font=fontlbl,text_color=bg1,justify=RIGHT)
    lb.grid(column=2,row=3,pady=12,padx=5,sticky=E)
    ebkpublisher=customtkinter.CTkEntry(fmeditbook,border_width=1,border_color=bg1,width=150)
    ebkpublisher.grid(column=3,row=3,pady=12,padx=5)

    lb=customtkinter.CTkLabel(fmeditbook,text="Date Published:",font=fontlbl,text_color=bg1,justify=RIGHT)
    lb.grid(column=0,row=4,pady=12,padx=5,sticky=E)
    ebkpubdate=customtkinter.CTkEntry(fmeditbook,border_width=1,border_color=bg1,width=150)
    ebkpubdate.grid(column=1,row=4,pady=12,padx=5)
    lb=customtkinter.CTkLabel(fmeditbook,text="Quantity:",font=fontlbl,text_color=bg1,justify=RIGHT)
    lb.grid(column=2,row=4,pady=12,padx=5,sticky=E)
    ebkquantity=customtkinter.CTkEntry(fmeditbook,border_width=1,border_color=bg1,width=150)
    ebkquantity.grid(column=3,row=4,pady=12,padx=5)
    lb=customtkinter.CTkLabel(fmeditbook,text="Price/book:",font=fontlbl,text_color=bg1,justify=RIGHT)
    lb.grid(column=0,row=5,pady=12,padx=5,sticky=E)
    ebkprice=customtkinter.CTkEntry(fmeditbook,border_width=1,border_color=bg1,width=150)
    ebkprice.grid(column=1,row=5,pady=12,padx=5)
    lb=customtkinter.CTkLabel(fmeditbook,text="Book No:",font=fontlbl,text_color=bg1,justify=RIGHT)
    lb.grid(column=2,row=5,pady=12,padx=5,sticky=E)
    ebkno=customtkinter.CTkEntry(fmeditbook,border_width=1,border_color=bg1,width=150)
    ebkno.grid(column=3,row=5,pady=12,padx=5)
    eframebtnaddbook=customtkinter.CTkFrame(fmeditbook)
    eframebtnaddbook.grid(column=0,row=6,columnspan=4)
    btnupdatebook=customtkinter.CTkButton(eframebtnaddbook,text="Update",fg_color=bg1,text_color=fg1,hover_color=hovbg1,cursor="hand2",command=updatebook)
    btnupdatebook.grid(column=0,row=0,pady=5,padx=6)
    btndeletebook=customtkinter.CTkButton(eframebtnaddbook,text="Delete",fg_color=cancelcol,text_color=fg1,hover_color=hovbg1,cursor="hand2",command=deletebook)
    btndeletebook.grid(column=1,row=0,pady=5,padx=6)
    btncancelebook=customtkinter.CTkButton(eframebtnaddbook,text="Cancel",fg_color=coldark,text_color=fg1,hover_color=hovbg1,cursor="hand2",command=cleareditbook)
    btncancelebook.grid(column=2,row=0,pady=5,padx=6)
    #-------------------view books---------------
    def viewbooks():
        try:
            sql="SELECT bookno,bookcategory,booktitle,bookauthor,bookpublisher,quantity,bookprice FROM books"
            cursor=conn.cursor()
            cursor.execute(sql)
            results=cursor.fetchall()
            if results:
                for records in tablebooks.get_children():
                    tablebooks.delete(records)
                for i in results:
                    tablebooks.insert('',END,values=i)
        except IOError:
            pass
    tablebooks=ttk.Treeview(fmviewbooks)
    tablebooks.pack(side=TOP,fill=BOTH,expand=True)
    tablebooks['show']="headings"
    tablebooks['columns']=(0,1,2,3,4,5,6)
    tablebooks.heading(0,text="Book No")
    tablebooks.heading(1,text="Category")
    tablebooks.heading(2,text="Title")
    tablebooks.heading(3,text="Author")
    tablebooks.heading(4,text="Publisher")
    tablebooks.heading(5,text="Quantity")
    tablebooks.heading(6,text="Price")
    tablebooks.column(0,width=50,anchor=CENTER)
    tablebooks.column(1,width=80,anchor=CENTER)
    tablebooks.column(2,width=100,anchor=CENTER)
    tablebooks.column(3,width=90,anchor=CENTER)
    tablebooks.column(4,width=90,anchor=CENTER)
    tablebooks.column(5,width=70,anchor=CENTER)
    tablebooks.column(6,width=80,anchor=CENTER)
    viewbooks()
    #============issue book=======================
    def issuebook():
        if bksadmno.get()=="":
            return False
        elif bksname.get()=="":
            return False
        elif bkbkno.get()=="":
            return False
        elif bkbksrno.get()=="":
            return False
        elif bkbktitle.get()=="":
            return False
        elif bkbkreturndate.get()=="":
            return False
        else:
            fullbookno=str(bkbkno.get())+"/"+str(bkbksrno.get())
            sql="INSERT INTO issuebook(bookno,stadmno,studentname,issuedate,returndate,cleared,recordby) VALUES('%s','%s','%s','%s','%s','%s','%s')"
            values=(str(fullbookno),bksadmno.get(),bksname.get(),today,bkbkreturndate.get(),"Not-cleared",myusername)
            try:
                cursor=conn.cursor()
                cursor.execute(sql%values)
                conn.commit()
                clearissuebook()
                viewissuedbooks()
                messagebox.showinfo("Success","book issued successful")
            except IOError:
                messagebox.showerror("Error","Unable to issue book.\n Try again")
                clearissuebook()
                pass

    def clearissuebook():
        bksadmno.delete(0,END)
        bksname.delete(0,END)
        bkbkno.delete(0,END)
        bkbksrno.delete(0,END)
        bkbktitle.delete(0,END)
        bkbkreturndate.delete(0,END)

    def searchstudent1(*arg):
        if bksadmno.get()!="":
            try:
                cursor=conn.cursor()
                sql="SELECT admno,CONCAT(fname,' ',lname,' ',sname) AS name FROM students WHERE admno='%s'"%(bksadmno.get())
                cursor.execute(sql)
                results=cursor.fetchall()
                if results:
                    bksadmno.delete(0,END)
                    bksadmno.insert(END,results[0][0])
                    bksname.delete(0,END)
                    bksname.insert(END,results[0][1])
                else:
                    messagebox.showerror("Error","Invalid student admno")
                    #bksadmno.focus()
                    clearissuebook()
            except:
                messagebox.showerror("Error","Invalid student admno")
                clearissuebook()
        else:
            pass
    def searchbook1(*arg):
        try:
            cursor=conn.cursor()
            sql="SELECT bookno,booktitle FROM books WHERE bookno='%s'"%(bkbkno.get())
            cursor.execute(sql)
            results=cursor.fetchall()
            if results:
                bkbkno.delete(0,END)
                bkbktitle.delete(0,END)
                bkbkno.insert(END,results[0][0])
                bkbktitle.insert(END,results[0][1])
            else:
                messagebox.showerror("Error","Invalid book No")
                clearissuebook()
        except:
            messagebox.showerror("Error","Invalid book No")
            clearissuebook()
    lb=customtkinter.CTkLabel(fmissuebook,text="Issue Book",text_color=bg1,font=fontlbl2)
    lb.grid(column=0,row=0,columnspan=4)
    lb=customtkinter.CTkLabel(fmissuebook,text="Admno",font=fontlbl,text_color=bg1,justify=RIGHT)
    lb.grid(column=0,row=1,pady=7,padx=5,sticky=E)
    bksadmno=customtkinter.CTkEntry(fmissuebook,border_width=1,border_color=bg1,width=200)
    bksadmno.grid(column=1,row=1,padx=5,pady=7)
    bksadmno.bind("<FocusOut>",searchstudent1)
    lb=customtkinter.CTkLabel(fmissuebook,text="Student Name",font=fontlbl,text_color=bg1,justify=RIGHT)
    lb.grid(column=2,row=1,pady=7,padx=5,sticky=E)
    bksname=customtkinter.CTkEntry(fmissuebook,border_width=1,border_color=bg1,width=200)
    bksname.grid(column=3,row=1,padx=5,pady=7)
    lb=customtkinter.CTkLabel(fmissuebook,text="Book No",font=fontlbl,text_color=bg1,justify=RIGHT)
    lb.grid(column=0,row=2,pady=7,padx=5,sticky=E)
    bkbkno=customtkinter.CTkEntry(fmissuebook,border_width=1,border_color=bg1,width=200)
    bkbkno.grid(column=1,row=2,padx=5,pady=7)
    bkbkno.bind("<FocusOut>",searchbook1)
    lb=customtkinter.CTkLabel(fmissuebook,text="Series No",font=fontlbl,text_color=bg1,justify=RIGHT)
    lb.grid(column=2,row=2,pady=7,padx=5,sticky=E)
    bkbksrno=customtkinter.CTkEntry(fmissuebook,border_width=1,border_color=bg1,width=200)
    bkbksrno.grid(column=3,row=2,padx=5,pady=7)
    lb=customtkinter.CTkLabel(fmissuebook,text="Title",font=fontlbl,text_color=bg1,justify=RIGHT)
    lb.grid(column=0,row=3,pady=7,padx=5,sticky=E)
    bkbktitle=customtkinter.CTkEntry(fmissuebook,border_width=1,border_color=bg1,width=200)
    bkbktitle.grid(column=1,row=3,padx=5,pady=7)
    lb=customtkinter.CTkLabel(fmissuebook,text="Return Date",font=fontlbl,text_color=bg1,justify=RIGHT)
    lb.grid(column=2,row=3,pady=7,padx=5,sticky=E)
    bkbkreturndate=DateEntry(fmissuebook,date_pattern="yyyy/mm/dd",width=19,font=fontlbl)
    bkbkreturndate.grid(column=3,row=3,padx=5,pady=7)
    frameissuebtn=customtkinter.CTkFrame(fmissuebook)
    frameissuebtn.grid(column=0,row=4,pady=10,columnspan=4)
    btnsaveissued=customtkinter.CTkButton(frameissuebtn,text="save",fg_color=bg1,text_color=fg1,hover_color=hovbg1,cursor="hand2",command=issuebook)
    btnsaveissued.grid(column=0,row=0,pady=5,padx=10)
    btncancelissued=customtkinter.CTkButton(frameissuebtn,text="Cancel",fg_color=coldark,text_color=fg1,hover_color=hovbg1,cursor="hand2",command=clearissuebook)
    btncancelissued.grid(column=1,row=0,pady=5,padx=10)
    #===========view issued=======================
    def viewissuedbooks():
        try:
            cursor=conn.cursor()
            sql="SELECT * FROM issuebook ORDER BY issueid DESC"
            cursor.execute(sql)
            results=cursor.fetchall()
            if results:
                for records in tableissued.get_children():
                    tableissued.delete(records)
                for i in results:
                    tableissued.insert('',END,values=i)
        except:
            pass
    def makeclearance():
        global issue_id_clearance,book_no_clearance
        if issue_id_clearance=="":
            messagebox.showwarning("Message","Record not selected")
        elif book_no_clearance=="":
            messagebox.showwarning("Message","Record not selected")
        else:
            msgtoclr=messagebox.askyesno("Clearance message","Do you want to clear this book no "+book_no_clearance)
            if msgtoclr==True:
                try:
                    cursor=conn.cursor()
                    sql="UPDATE issuebook SET cleared='Cleared' WHERE issueid=%s"
                    cursor.execute(sql%(issue_id_clearance))
                    conn.commit()
                    messagebox.showinfo("Success","Book cleared successfully")
                    issue_id_clearance=""
                    book_no_clearance=""
                    viewissuedbooks()
                except IOError:
                    conn.rollback()
                    messagebox.showerror("Error","Unable to clear this book\n Try again")
                    issue_id_clearance=""
                    book_no_clearance=""
    def get_issue_id(event):
        global issue_id_clearance,book_no_clearance
        selected_record=tableissued.focus()
        items=tableissued.item(selected_record,"values")
        issue_id_clearance=items[0]
        book_no_clearance=items[1]

    framefilterissue=customtkinter.CTkFrame(fmviewissued)
    framefilterissue.pack(side=TOP,fill=X)
    lb=customtkinter.CTkLabel(framefilterissue,text="From Date",font=fontlbl,text_color=bg1,justify=LEFT)
    lb.grid(column=0,row=0,pady=5,padx=2,sticky=E)
    issuefrom=DateEntry(framefilterissue,date_pattern="yyyy/mm/dd",width=10,font=("times",9))
    issuefrom.grid(column=1,row=0,padx=2,pady=5)
    lb=customtkinter.CTkLabel(framefilterissue,text="To Date",font=fontlbl,text_color=bg1,justify=LEFT)
    lb.grid(column=2,row=0,pady=5,padx=2,sticky=E)
    issueto=DateEntry(framefilterissue,date_pattern="yyyy/mm/dd",width=10,font=("times",9))
    issueto.grid(column=3,row=0,padx=2,pady=5)
    lb=customtkinter.CTkLabel(framefilterissue,text="Admno",font=("times",10),text_color=bg1,justify=LEFT)
    lb.grid(column=4,row=0,pady=7,padx=2,sticky=E)
    issuedadmno=customtkinter.CTkEntry(framefilterissue,border_width=1,border_color=bg1,width=100)
    issuedadmno.grid(column=5,row=0,padx=2,pady=7)

    tableissued=ttk.Treeview(fmviewissued)
    tableissued.pack(side=TOP,fill=BOTH,expand=True)
    tableissued['show']="headings"
    tableissued['columns']=(0,1,2,3,4,5,6,7)
    tableissued.heading(0,text="Id")
    tableissued.heading(1,text="Book No")
    tableissued.heading(2,text="Admno")
    tableissued.heading(3,text="Name")
    tableissued.heading(4,text="I Date")
    tableissued.heading(5,text="R Date")
    tableissued.heading(6,text="Cleared")
    tableissued.heading(7,text="Recordby")
    tableissued.column(0,width=50,anchor=CENTER)
    tableissued.column(1,width=70,anchor=CENTER)
    tableissued.column(2,width=60,anchor=CENTER)
    tableissued.column(3,width=80,anchor=CENTER)
    tableissued.column(4,width=70,anchor=CENTER)
    tableissued.column(5,width=70,anchor=CENTER)
    tableissued.column(6,width=60,anchor=CENTER)
    tableissued.column(7,width=80,anchor=CENTER)
    tableissued.bind("<Double-1>",get_issue_id)
    viewissuedbooks()
    frameclearance=customtkinter.CTkFrame(fmviewissued)
    frameclearance.pack(side=BOTTOM)
    btnclearance=customtkinter.CTkButton(frameclearance,text="Clearance",text_color=fg1,fg_color=bg1,hover_color=hovbg1,cursor="hand2",command=makeclearance)
    btnclearance.grid(column=0,row=0)
    #=========framesystemusers====================
    def toaddsystemuser():
        ntbook7.select(0)
    def toeditsystemuser():
        ntbook7.select(1)
    def toviewsystemusers():
        ntbook7.select(2)
        viewsystemusers()
        
    def addsystemuser():
        if sysstaffid.get()=="":
            return False
        elif sysstaffname.get()=="":
            return False
        elif sysstaffusername.get()=="":
            return False
        elif sysstaffpassword.get()=="":
            return False
        elif sysstaffrole.get()=="":
            return False
        else:
            global myusername
            try:
                sql="INSERT INTO systemusers(sysstaffid,username,password,role,activestatus,recordby) VALUES(%s,'%s','%s','%s','%s','%s')"
                sysvalues=(sysstaffid.get(),sysstaffusername.get(),sysstaffpassword.get(),sysstaffrole.get(),0,myusername)
                cursor=conn.cursor()
                cursor.execute(sql%sysvalues)
                conn.commit()
                viewsystemusers()
                clearsystemuserform()
                get_system_analysis()
                messagebox.showinfo("Message","Staff added as user successful")    
            except IOError:
                messagebox.showinfo("Message","Unable to process.\n Try again")
                pass
    def viewsystemusers():
        try:
            sql="SELECT sysstaffid,username,password,role,lastlogindate,activestatus,recordby FROM systemusers"
            cursor=conn.cursor()
            cursor.execute(sql)
            results=cursor.fetchall()
            if results:
                for records in tablesystemusers.get_children():
                    tablesystemusers.delete(records)
                for i in results:
                    tablesystemusers.insert('',END,values=i)

        except IOError:
            pass
        
    def clearsystemuserform():
        
        searchstaffno.delete(0,END)
        sysstaffid.delete(0,END)
        sysstaffno.delete(0,END)
        sysstaffname.delete(0,END)
        sysstaffoccupation.delete(0,END)
        sysstaffusername.delete(0,END)
        sysstaffpassword.delete(0,END)
        sysstaffrole.set("")
    def cleareditsystemuserform():
        global sysuid
        sysuid=""
        esearchstaffno.delete(0,END)
        esysstaffid.delete(0,END)
        esysstaffno.delete(0,END)
        esysstaffname.delete(0,END)
        esysstaffoccupation.delete(0,END)
        esysstaffusername.delete(0,END)
        esysstaffpassword.delete(0,END)
        esysstaffrole.set("")
    def updatesystemuser():
        global myusername,sysuid
        if sysuid!="":
            try:
                sql="UPDATE systemusers SET username='%s',password='%s',role='%s',recordby='%s' WHERE sysstaffid=%s"%(esysstaffusername.get(),esysstaffpassword.get(),esysstaffrole.get(),myusername,sysuid)
                cursor=conn.cursor()
                cursor.execute(sql)
                conn.commit()
                cleareditsystemuserform()
                viewsystemusers()
                get_system_analysis()
                messagebox.showinfo("success","User updated successfully")
            except IOError:
                messagebox.showwarning("success","Unable to update.\n Try again")
                pass
        else:
            messagebox.showwarning("Error message","Search staff to update")
    def searchstaffsystemuser():
        try:
            sql="SELECT staffid,staffno,CONCAT(fname,' ',lname,' ',sname) AS name,occupation FROM staffs WHERE staffno='%s'"%(searchstaffno.get())
            cursor=conn.cursor()
            cursor.execute(sql)
            results=cursor.fetchall()
            if results:
                clearsystemuserform()
                sysstaffid.insert(END,results[0][0])
                sysstaffno.insert(END,results[0][1])
                sysstaffname.insert(END,results[0][2])
                sysstaffoccupation.insert(END,results[0][3])
            else:
                messagebox.showwarning("Error message","Invalid staff no. Try again!")
        except IOError:
            pass
    def searchsystemuser():
        global sysuid
        if esearchstaffno.get()!="":
            try:
                sql="SELECT stf.*,sysu.* FROM staffs stf,systemusers sysu WHERE stf.staffid=sysu.sysstaffid AND stf.staffno='%s'"%(esearchstaffno.get())
                cursor=conn.cursor()
                cursor.execute(sql)
                results=cursor.fetchall()
                if results:
                    cleareditsystemuserform()
                    sysuid=results[0][0]
                    esysstaffid.insert(END,results[0][0])
                    esysstaffno.insert(END,results[0][8])
                    esysstaffname.insert(END,results[0][1]+' '+results[0][2]+' '+results[0][3])
                    esysstaffoccupation.insert(END,results[0][9])
                    esysstaffusername.insert(END,results[0][17])
                    esysstaffpassword.insert(END,results[0][18])
                    esysstaffrole.set(results[0][19])
            except IOError:
                pass
        else:
            messagebox.showwarning("Error message","Enter staff no to search!")
    def deletesystemuser():
        global myusername,sysuid
        if sysuid!="":
            todeletemsg=messagebox.askyesno("Confirm message","Do you want to delete system user "+esysstaffno.get())
            if todeletemsg==True:
                try:
                    sql="DELETE FROM systemusers WHERE sysstaffid=%s"%(sysuid)
                    cursor=conn.cursor()
                    cursor.execute(sql)
                    conn.commit()
                    cleareditsystemuserform()
                    viewsystemusers()
                    get_system_analysis()
                    messagebox.showinfo("Success","User deleted successful")
                except IOError:
                    messagebox.showerror("alert message","Unable to delete user\n Try again")
                    pass
        else:
            messagebox.showwarning("Error message","Search staff to update")
    framesubmenu7=customtkinter.CTkFrame(framesystemuser,width=150,fg_color=coldark)
    framesubmenu7.pack(side=LEFT,fill=Y)
    btnadduser=customtkinter.CTkButton(framesubmenu7,text="Add User",text_color=fg1,fg_color=bg1,hover_color=hovbg1,cursor="hand2",command=toaddsystemuser)
    btnadduser.pack(pady=6,padx=6)
    btnedituser=customtkinter.CTkButton(framesubmenu7,text="Edit User",text_color=fg1,fg_color=bg1,hover_color=hovbg1,cursor="hand2",command=toeditsystemuser)
    btnedituser.pack(pady=6,padx=6)
    btnviewuser=customtkinter.CTkButton(framesubmenu7,text="View User",text_color=fg1,fg_color=bg1,hover_color=hovbg1,cursor="hand2",command=toviewsystemusers)
    btnviewuser.pack(pady=6,padx=6)
    ntbook7=ttk.Notebook(framesystemuser)
    ntbook7.pack(side=LEFT,fill=BOTH,expand=True)
    fmadduser=customtkinter.CTkFrame(ntbook7,fg_color=bg3,bg_color=coldark)
    fmadduser.place(x=0,y=0,relheight=1,relwidth=1)
    fmedituser=customtkinter.CTkFrame(ntbook7,fg_color=bg3,bg_color=coldark)
    fmedituser.place(x=0,y=0,relheight=1,relwidth=1)
    fmviewuser=customtkinter.CTkFrame(ntbook7,fg_color=bg3,bg_color=coldark)
    fmviewuser.place(x=0,y=0,relheight=1,relwidth=1)
    ntbook7.add(fmadduser,text="Add User")
    ntbook7.add(fmedituser,text="Edit User")
    ntbook7.add(fmviewuser,text="View Users")
    #-------------addsystem user---------------
    lb=customtkinter.CTkLabel(fmadduser,text="Add System Form",text_color=bg1,font=fontlbl2,justify=CENTER,anchor=CENTER)
    lb.grid(column=0,row=0,pady=(5,5),columnspan=4)
    fmsearchstaff=customtkinter.CTkFrame(fmadduser,fg_color=bg3)
    fmsearchstaff.grid(column=0,row=1)
    lb=customtkinter.CTkLabel(fmsearchstaff,text="Search Staff By No",text_color=bg1,font=fontlbl,justify=RIGHT)
    lb.grid(column=0,row=0,pady=(5,5))
    searchstaffno=customtkinter.CTkEntry(fmsearchstaff,border_width=1,border_color=bg1,font=fontentries,width=textboxwidth)
    searchstaffno.grid(column=1,row=0,pady=(5,5))
    btnsearchstfno=customtkinter.CTkButton(fmsearchstaff,fg_color=bg1,hover_color=hovbg1,text_color=fg1,text="Search",command=searchstaffsystemuser)
    btnsearchstfno.grid(column=2,row=0,pady=(5,5),padx=(3,3))
    fmstaffdetails=customtkinter.CTkFrame(fmadduser,fg_color=bg3)
    fmstaffdetails.grid(column=0,row=2)
    lb=customtkinter.CTkLabel(fmstaffdetails,text="Staff ID",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=CENTER)
    lb.grid(column=0,row=0)
    sysstaffid=customtkinter.CTkEntry(fmstaffdetails,border_width=1,border_color=bg1,font=fontentries,width=textboxwidth)
    sysstaffid.grid(column=1,row=0,pady=(5,5))
    lb=customtkinter.CTkLabel(fmstaffdetails,text="Staff No",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=CENTER)
    lb.grid(column=2,row=0)
    sysstaffno=customtkinter.CTkEntry(fmstaffdetails,border_width=1,border_color=bg1,font=fontentries,width=textboxwidth)
    sysstaffno.grid(column=3,row=0,pady=(5,5))
    lb=customtkinter.CTkLabel(fmstaffdetails,text="Name",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=CENTER)
    lb.grid(column=0,row=1)
    sysstaffname=customtkinter.CTkEntry(fmstaffdetails,border_width=1,border_color=bg1,font=fontentries,width=textboxwidth)
    sysstaffname.grid(column=1,row=1,pady=(5,5))
    lb=customtkinter.CTkLabel(fmstaffdetails,text="Occupation",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=CENTER)
    lb.grid(column=2,row=1)
    sysstaffoccupation=customtkinter.CTkEntry(fmstaffdetails,border_width=1,border_color=bg1,font=fontentries,width=textboxwidth)
    sysstaffoccupation.grid(column=3,row=1,pady=(5,5))
    lb=customtkinter.CTkLabel(fmstaffdetails,text="Username",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=CENTER)
    lb.grid(column=0,row=2)
    sysstaffusername=customtkinter.CTkEntry(fmstaffdetails,border_width=1,border_color=bg1,font=fontentries,width=textboxwidth)
    sysstaffusername.grid(column=1,row=2,pady=(5,5))
    lb=customtkinter.CTkLabel(fmstaffdetails,text="Password",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=CENTER)
    lb.grid(column=0,row=3)
    sysstaffpassword=customtkinter.CTkEntry(fmstaffdetails,border_width=1,border_color=bg1,show='*',font=fontentries,width=textboxwidth)
    sysstaffpassword.grid(column=1,row=3,pady=(5,5))
    lb=customtkinter.CTkLabel(fmstaffdetails,text="Role",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=CENTER)
    lb.grid(column=0,row=4)
    sysstaffrole=customtkinter.CTkOptionMenu(fmstaffdetails,font=fontentries,width=textboxwidth,button_color=bg1,button_hover_color=bg1,text_color="black",fg_color="white",values=["Principal","Deputy-Principal","Teacher","Secretary","Bursar","Librarian","ICT"])
    sysstaffrole.grid(column=1,row=4,pady=(5,5))
    framebtnaddsysuser=customtkinter.CTkFrame(fmstaffdetails)
    framebtnaddsysuser.grid(column=0,row=5,pady=(5,5),columnspan=4)
    btnsavesysuser=customtkinter.CTkButton(framebtnaddsysuser,text="Save",text_color=fg1,fg_color=bg1,hover_color=hovbg1,command=addsystemuser)
    btnsavesysuser.grid(column=0,row=0,padx=5,pady=(5,5))
    btncancelsysuser=customtkinter.CTkButton(framebtnaddsysuser,text="Cancel",text_color=fg1,fg_color=cancelcol,hover_color=hovbg1,command=clearsystemuserform)
    btncancelsysuser.grid(column=1,row=0,padx=5,pady=(5,5))
    #------------------------------------------
    #-------------editsystem user---------------
    lb=customtkinter.CTkLabel(fmedituser,text="Edit System Form",text_color=bg1,font=fontlbl2,justify=CENTER,anchor=CENTER)
    lb.grid(column=0,row=0,pady=(5,5),columnspan=4)
    efmsearchstaff=customtkinter.CTkFrame(fmedituser,fg_color=bg3)
    efmsearchstaff.grid(column=0,row=1)
    lb=customtkinter.CTkLabel(efmsearchstaff,text="Search Staff",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=CENTER)
    lb.grid(column=0,row=0,pady=(5,5))
    esearchstaffno=customtkinter.CTkEntry(efmsearchstaff,border_width=1,border_color=bg1,font=fontentries,width=textboxwidth)
    esearchstaffno.grid(column=1,row=0,pady=(5,5))
    ebtnsearchstfno=customtkinter.CTkButton(efmsearchstaff,fg_color=bg1,hover_color=hovbg1,text_color=fg1,text="Search",command=searchsystemuser)
    ebtnsearchstfno.grid(column=2,row=0,pady=(5,5),padx=(3,3))
    efmstaffdetails=customtkinter.CTkFrame(fmedituser,fg_color=bg3)
    efmstaffdetails.grid(column=0,row=2)
    lb=customtkinter.CTkLabel(efmstaffdetails,text="Staff ID",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=CENTER)
    lb.grid(column=0,row=0)
    esysstaffid=customtkinter.CTkEntry(efmstaffdetails,border_width=1,border_color=bg1,font=fontentries,width=textboxwidth)
    esysstaffid.grid(column=1,row=0,pady=(5,5))
    lb=customtkinter.CTkLabel(efmstaffdetails,text="Staff No",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=CENTER)
    lb.grid(column=2,row=0)
    esysstaffno=customtkinter.CTkEntry(efmstaffdetails,border_width=1,border_color=bg1,font=fontentries,width=textboxwidth)
    esysstaffno.grid(column=3,row=0,pady=(5,5))
    lb=customtkinter.CTkLabel(efmstaffdetails,text="Name",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=CENTER)
    lb.grid(column=0,row=1)
    esysstaffname=customtkinter.CTkEntry(efmstaffdetails,border_width=1,border_color=bg1,font=fontentries,width=textboxwidth)
    esysstaffname.grid(column=1,row=1,pady=(5,5))
    lb=customtkinter.CTkLabel(efmstaffdetails,text="Occupation",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=CENTER)
    lb.grid(column=2,row=1)
    esysstaffoccupation=customtkinter.CTkEntry(efmstaffdetails,border_width=1,border_color=bg1,font=fontentries,width=textboxwidth)
    esysstaffoccupation.grid(column=3,row=1,pady=(5,5))
    lb=customtkinter.CTkLabel(efmstaffdetails,text="Username",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=CENTER)
    lb.grid(column=0,row=2)
    esysstaffusername=customtkinter.CTkEntry(efmstaffdetails,border_width=1,border_color=bg1,font=fontentries,width=textboxwidth)
    esysstaffusername.grid(column=1,row=2,pady=(5,5))
    lb=customtkinter.CTkLabel(efmstaffdetails,text="Password",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=CENTER)
    lb.grid(column=0,row=3)
    esysstaffpassword=customtkinter.CTkEntry(efmstaffdetails,border_width=1,border_color=bg1,show='*',font=fontentries,width=textboxwidth)
    esysstaffpassword.grid(column=1,row=3,pady=(5,5))
    lb=customtkinter.CTkLabel(efmstaffdetails,text="Role",text_color=bg1,font=fontlbl,justify=RIGHT,anchor=CENTER)
    lb.grid(column=0,row=4)
    esysstaffrole=customtkinter.CTkOptionMenu(efmstaffdetails,fg_color="white",button_color=bg1,button_hover_color=bg1,text_color="black",font=fontentries,width=textboxwidth,values=["Principal","Deputy-Principal","Teacher","Secretary","Bursar","Librarian","ICT"])
    esysstaffrole.grid(column=1,row=4,pady=(5,5))
    eframebtnaddsysuser=customtkinter.CTkFrame(efmstaffdetails)
    eframebtnaddsysuser.grid(column=0,row=5,pady=(5,5),columnspan=4)
    ebtnsavesysuser=customtkinter.CTkButton(eframebtnaddsysuser,text="Update",text_color=fg1,fg_color=bg1,hover_color=hovbg1,command=updatesystemuser)
    ebtnsavesysuser.grid(column=0,row=0,pady=(5,5),padx=5)
    ebtndeletesysuser=customtkinter.CTkButton(eframebtnaddsysuser,text="Delete",text_color=fg1,fg_color=cancelcol,hover_color=hovbg1,command=deletesystemuser)
    ebtndeletesysuser.grid(column=1,row=0,pady=(5,5),padx=5)
    ebtncancelsysuser=customtkinter.CTkButton(eframebtnaddsysuser,text="Cancel",text_color=fg1,fg_color=coldark,hover_color=hovbg1,command=cleareditsystemuserform)
    ebtncancelsysuser.grid(column=2,row=0,pady=(5,5),padx=5)
    #------------------------------------------
    #-----------------view system users---------
    tablesystemusers=ttk.Treeview(fmviewuser)
    tablesystemusers.pack(side=LEFT,fill=BOTH,expand=True)
    tablesystemusers['columns']=(0,1,2,3,4,5,6)
    tablesystemusers['show']="headings"
    tablesystemusers.heading(0,text="Staff ID")
    tablesystemusers.heading(1,text="Username")
    tablesystemusers.heading(2,text="Password")
    tablesystemusers.heading(3,text="Role")
    tablesystemusers.heading(4,text="LastLogin")
    tablesystemusers.heading(5,text="Status")
    tablesystemusers.heading(6,text="Record By")
    tablesystemusers.column(0,width=60)
    tablesystemusers.column(1,width=130)
    tablesystemusers.column(2,width=60)
    tablesystemusers.column(3,width=80)
    tablesystemusers.column(4,width=110)
    tablesystemusers.column(5,width=80)
    tablesystemusers.column(6,width=80)
    viewsystemusers()
    #-------------------------------------------
    #==========================================
    toremovealltabs()
    todashboard()
    win.protocol("WM_DELETE_WINDOW",tologout)
    win.mainloop()
def tosplashscreen():
    global icon1,icon
    def run_progress():
        myprogrees['maximum']=100
        for i in range(101):
            sleep(0.1)
            myprogrees['value']=i
            myprogrees.update()
            if i==100:
                mysplashwin.destroy()
                
                tomainwindow()
            else:
                pass
                
            
    mysplashwin=Tk()
    width=500
    height=220
    #========icons====================
    iconc1=Image.open("icons/schoolmis.png")
    iconc1=iconc1.resize((130,130))
    iconc1=ImageTk.PhotoImage(iconc1)
    #=================================
    scr_w=mysplashwin.winfo_screenwidth()
    scr_h=mysplashwin.winfo_screenheight()
    x_cord=(scr_w/2)-(width/2)
    y_cord=(scr_h/2)-(height/2)
    mysplashwin.overrideredirect(True)
    mysplashwin.geometry("%dx%d+%d+%d"%(width,height,x_cord,y_cord))
    frame1=customtkinter.CTkFrame(mysplashwin,height=height,width=width,bg_color=bg1,fg_color=bg1)
    frame1.pack(fill=BOTH,expand=True)
    label=customtkinter.CTkLabel(frame1,image=iconc1,text="",height=height,width=(width/3))
    label.pack(side=LEFT)
    frame2=customtkinter.CTkFrame(frame1,border_width=1,border_color=bg1,corner_radius=0,height=height,width=(width/2),fg_color="aliceblue")
    frame2.pack(side=LEFT,fill=BOTH,expand=True)
    label1=customtkinter.CTkLabel(frame2,text="School Management information \nsystem",text_color=bg1,font=("times",16,"bold"))
    label1.pack(side=TOP,pady=(9,5))
    label2=customtkinter.CTkLabel(frame2,text="The system manages school information with ease",text_color=fg2,font=("times",11,"italic"))
    label2.pack(side=TOP,pady=(3,4))
    label3=customtkinter.CTkLabel(frame2,text="Modules: Students, Staffs,\n Exams,Library, Reporting, Bulk SMS",text_color=fg2,font=("times",11,"italic"))
    label3.pack(side=TOP,pady=(3,4))

    s=ttk.Style()
    s.theme_use('clam')
    s.configure("darkblue.Horizontal.TProgressbar",foreground=bg1,background=bg1)


    myprogrees=ttk.Progressbar(frame2,style="darkblue.Horizontal.TProgressbar",orient=HORIZONTAL,length=(width/20),mode="determinate")
    myprogrees.pack(side=BOTTOM,fill=X,pady=(20,20),padx=(10,10))
    run_progress()
    mysplashwin.mainloop()
#=================login window============
loginwin=Tk()
mywid=650
myhei=370
icon=PhotoImage(file="icons/schoolmis.png")
#========icons====================
icon1=Image.open("icons/schoolmis.png")
icon1=icon1.resize((150,150))
icon1=ImageTk.PhotoImage(icon1)
icon2=Image.open("icons/admin.png")
icon2=icon2.resize((50,50))
icon2=ImageTk.PhotoImage(icon2)
#=================================
loginwin.title("School Management Information System")
loginwin.iconphoto(False,icon)
loginwin.resizable(False,False)

scr_w=loginwin.winfo_screenwidth()
scr_h=loginwin.winfo_screenheight()
x_cord=(scr_w/2)-(mywid/2)
y_cord=(scr_h/2)-(myhei/2)
loginwin.geometry("%dx%d+%d+%d"%(mywid,myhei,x_cord,y_cord))
loginwin.config(bg=bg1)

def tologin():
    global conn,myusername
    username=entryusername.get()
    password=entrypassword.get().encode('utf-8')
    sql="SELECT username,password FROM systemusers WHERE username='%s'"%(username)
    try:
        cursor=conn.cursor()
        cursor.execute(sql)
        results=cursor.fetchall()
        conn.commit()
        if results:
            dbpass=results[0][1].encode('utf-8')
            if bcrypt.checkpw(password,dbpass)==True:
                today=time.strftime("%Y/%m/%d")
                curtime=time.strftime("%H:%M:%S")
                activestatus=1
                sql2="UPDATE systemusers SET lastlogindate='%s',lastlogintime='%s',activestatus=%s WHERE username='%s'"%(today,curtime,activestatus,username)
                try:
                    cursor=conn.cursor()
                    cursor.execute(sql2)
                    conn.commit()
                    myusername=username
                    loginwin.destroy()
                    tosplashscreen()
                except IOError:
                    pass
        else:
            messagebox.showerror("error message","Username or password is incorrect")
        
    except IOError:
        conn.rollback()
        pass

frm1=customtkinter.CTkFrame(loginwin,height=myhei,width=mywid,bg_color=bg1,fg_color=bg1)
frm1.pack(fill=BOTH,expand=True)
lbl=customtkinter.CTkLabel(frm1,image=icon1,text="",height=myhei,width=(mywid/2))
lbl.pack(side=LEFT)
frmform=customtkinter.CTkFrame(frm1,height=myhei,width=(mywid/2),fg_color=bg2)
frmform.pack(side=LEFT,fill=Y)
lbl1=Label(frmform,text="-- Admin Login --",compound=TOP,image=icon2,width=(mywid/2),font=("times",14,"bold"),bg=bg2)
lbl1.place(x=5,y=0)
lblusername=customtkinter.CTkLabel(frmform,text="Username: ",font=("times",12,"bold"),anchor=customtkinter.W)
lblusername.place(relx=0.1,y=100)
entryusername=customtkinter.CTkEntry(frmform,width=(mywid/3),text_color=bg1,border_width=1,border_color=bg1)
entryusername.place(relx=0.1,y=130)
entryusername.focus()
lbl2=customtkinter.CTkLabel(frmform,text="Password:",font=("times",12,"bold"),anchor=customtkinter.W)
lbl2.place(relx=0.1,y=160)
entrypassword=customtkinter.CTkEntry(frmform,width=(mywid/3),show="*",text_color=bg1,border_width=1,border_color=bg1)
entrypassword.place(relx=0.1,y=190)
btnlogin=customtkinter.CTkButton(master=frmform,text="Login",cursor="hand2",corner_radius=7,
text_color=bg2,fg_color=bg1,hover_color=bg1,command=tologin)
btnlogin.place(relx=0.3,y=240)
loginwin.mainloop()
